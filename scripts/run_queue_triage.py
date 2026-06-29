#!/usr/bin/env python3
"""
run_queue_triage.py — Stage 5 Run 1: Queue-driven triage orchestrator.

Reads the 'Queue' tab in sources/youtube/ingest_queue.xlsx.
For each pending row, enumerates videos into the named source tab via the
existing triage pipeline (youtube_triage.process_sheet).

Gate: this script only writes metadata (video titles, guesses, ingest flags).
No transcripts are downloaded, no DB writes occur. Run 2 (run_queue_ingest.py)
is a separate, explicit step — review source tabs before running it.

Queue tab columns:
  url         — channel, playlist, or single video URL
  source_name — tab name to write into (and label for whitelist files)
  review      — "yes" (user reviews before ingest) | "no" (auto-approved via whitelist)
  filter      — "min5", "whitelist", or blank; comma-separate multiple, e.g. "whitelist,min5"
  limit       — integer cap on videos enumerated (blank = no cap)
  status      — pending → triaged (updated by this script)

Filter logic:
  min5      → --min-duration 300 (skip clips ≤5 min)
  whitelist → title-whitelist mode; reads sources/youtube/whitelist_{slug}.txt
              where slug = source_name lowercased with spaces→underscores

review=no + whitelist filter → rows pre-classified sermon=TRUE, Groq skipped
review=yes                   → Groq classifies; user reviews ingest=TRUE rows

Fault-tolerant: one bad Queue row logs and continues, never kills the run.

Usage:
    python3 scripts/run_queue_triage.py
    python3 scripts/run_queue_triage.py --dry-run
"""

import os
import re
import sys
import time
from pathlib import Path
from typing import List, Optional

import openpyxl
from dotenv import load_dotenv
from groq import Groq

ROOT       = Path(__file__).resolve().parent.parent
QUEUE_PATH = ROOT / "sources" / "youtube" / "ingest_queue.xlsx"

load_dotenv(ROOT / "backend" / "app" / ".env")

_SCRIPTS = Path(__file__).resolve().parent
if str(_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS))

from youtube_triage import (  # noqa: E402
    find_ytdlp, load_whitelist, process_sheet, load_or_create_wb,
    COLUMNS as DETAIL_COLUMNS,
)

QUEUE_COLS  = ["url", "source_name", "review", "filter", "limit", "status"]
QUEUE_COL   = {name: i + 1 for i, name in enumerate(QUEUE_COLS)}
SKIP_TABS   = {"Queue", "HowToRun"}
MIN5_SECS   = 300


# ── Queue sheet helpers ────────────────────────────────────────────────────────

def qcell(ws, row: int, col: str):
    return ws.cell(row=row, column=QUEUE_COL[col]).value

def sqcell(ws, row: int, col: str, value) -> None:
    ws.cell(row=row, column=QUEUE_COL[col], value=value)


def _slug(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")


def _whitelist_path(source_name: str) -> Path:
    return ROOT / "scripts" / f"whitelist_{_slug(source_name)}.txt"


def _channel_slug_from_url(url: str) -> str:
    """Derive a readable label from a YouTube channel URL handle.
    @sermonindex/videos → 'sermonindex'
    Used as the workbook tab label for whitelist rows where source_name is blank.
    """
    from urllib.parse import urlparse
    path = urlparse(url.strip()).path.rstrip("/")
    m = re.search(r'/@([^/?]+)', path)
    if m:
        return m.group(1)
    skip = {"videos", "playlists", "streams", "shorts", "community", "about"}
    parts = [p for p in path.split("/") if p and p.lower() not in skip]
    return parts[-1] if parts else "unknown_channel"


def _normalize_playlist_url(url: str) -> str:
    """Convert watch?v=X&list=Y to the cleaner playlist?list=Y form."""
    if "list=" in url and "watch?v=" in url:
        from urllib.parse import urlparse, parse_qs
        list_id = parse_qs(urlparse(url).query).get("list", [None])[0]
        if list_id:
            return f"https://www.youtube.com/playlist?list={list_id}"
    return url


_KNOWN_FILTER_TOKENS = {"min5", "whitelist"}


def _parse_filters(filter_cell: str):
    """Parse the filter cell string. Returns (min_duration, whitelist_mode).
    Warns on unrecognised tokens so typos in the filter column are visible.
    Valid values: blank, 'min5', 'whitelist', or comma-combos thereof.
    """
    raw = str(filter_cell or "").strip()
    parts = [p.strip().lower() for p in raw.split(",") if p.strip()]
    unknown = [p for p in parts if p not in _KNOWN_FILTER_TOKENS]
    if unknown:
        print(f"  ⚠  filter: unrecognised token(s) {unknown!r} — ignored. "
              f"Valid tokens: {sorted(_KNOWN_FILTER_TOKENS)}")
    min_duration   = MIN5_SECS if "min5" in parts else 0
    whitelist_mode = "whitelist" in parts
    return min_duration, whitelist_mode


def _parse_limit(limit_cell) -> Optional[int]:
    """Parse the limit cell. Returns int or None (no cap).
    Warns loudly if the value is non-integer — defends against filter/limit
    column mix-ups (e.g. 'min5' accidentally entered in the limit cell).
    """
    raw = str(limit_cell or "").strip()
    if not raw or raw.lower() in ("none", "null", "na"):
        return None
    try:
        v = int(float(raw))
        return v if v > 0 else None
    except (ValueError, TypeError):
        print(f"  ⚠  limit: expected an integer but got {raw!r} — treating as unset (no cap). "
              f"Did you put a filter token in the limit column by mistake?")
        return None


def _ensure_tab(wb: openpyxl.Workbook, tab_name: str) -> None:
    """Create source detail tab with standard headers if it doesn't exist."""
    if tab_name not in wb.sheetnames:
        ws = wb.create_sheet(tab_name)
        ws.append(DETAIL_COLUMNS)
        wb.save(QUEUE_PATH)
        print(f"  Created tab: {tab_name!r}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    import argparse
    parser = argparse.ArgumentParser(
        description="Run 1 — Queue-driven triage: enumerate videos into source tabs"
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="Print what would happen without writing to the sheet or DB")
    parser.add_argument("--only", metavar="LABEL",
                        help="Process only the Queue row whose tab label matches LABEL "
                             "(case-insensitive). All other pending rows are skipped.")
    parser.add_argument("--time-limit", type=float, default=None, metavar="MINUTES",
                        help="Stop gracefully at the next Queue-row boundary once this many "
                             "minutes have elapsed. Workbook is always saved before stopping.")
    args = parser.parse_args()

    if not QUEUE_PATH.exists():
        print(f"ERROR: queue not found at {QUEUE_PATH}")
        sys.exit(1)

    _wb_check = openpyxl.load_workbook(QUEUE_PATH, read_only=True)
    if "Queue" not in _wb_check.sheetnames:
        print("ERROR: 'Queue' tab not found in ingest_queue.xlsx")
        _wb_check.close()
        sys.exit(1)
    _wb_check.close()

    ytdlp = find_ytdlp()
    if not ytdlp:
        print("ERROR: yt-dlp not found. Run: pip3 install yt-dlp")
        sys.exit(1)

    # Init Groq once — used only for review=yes rows (non-whitelist)
    _groq_key = os.environ.get("GROQ_API_KEY")
    groq_client = Groq(api_key=_groq_key) if _groq_key else None
    if not groq_client:
        print("⚠  GROQ_API_KEY not set — review=yes rows will skip Groq classification")

    wb = openpyxl.load_workbook(QUEUE_PATH)
    ws_q = wb["Queue"]

    # Collect pending Queue rows
    pending = [
        r for r in range(2, ws_q.max_row + 1)
        if qcell(ws_q, r, "url") and str(qcell(ws_q, r, "status") or "").strip().lower()
           in ("pending", "")
    ]

    mode = "(DRY-RUN) " if args.dry_run else ""
    print(f"\n── Queue Triage {mode}────────────────────────────────────────────")
    print(f"  {len(pending)} pending row(s) in Queue")
    if args.time_limit:
        print(f"  Time limit: {args.time_limit} min")

    summary = []   # [(source_name, written, error)]
    start_time = time.time()

    for row_idx in pending:
        if args.time_limit:
            elapsed_min = (time.time() - start_time) / 60
            if elapsed_min >= args.time_limit:
                print(f"\nTime limit reached ({elapsed_min:.1f} min) — stopping before next row.")
                break
        url_raw     = str(qcell(ws_q, row_idx, "url") or "").strip()
        source_name = str(qcell(ws_q, row_idx, "source_name") or "").strip()
        review      = str(qcell(ws_q, row_idx, "review") or "yes").strip().lower()
        filter_raw  = str(qcell(ws_q, row_idx, "filter") or "").strip()
        limit_raw   = qcell(ws_q, row_idx, "limit")

        # Parse filters early — needed before the source_name guard so that
        # whitelist rows (which intentionally have blank source_name) are not skipped.
        min_duration, whitelist_mode = _parse_filters(filter_raw)
        limit = _parse_limit(limit_raw)

        if not url_raw:
            print(f"\n  row {row_idx}: skipping — no URL")
            continue
        if not source_name and not whitelist_mode:
            # Blank source_name is only valid for whitelist rows (self-attribute per speaker).
            print(f"\n  row {row_idx}: skipping — no source_name "
                  f"(set filter=whitelist if this is a multi-speaker whitelist row)")
            continue

        # Tab label: use source_name when present; for whitelist rows with blank
        # source_name, derive from the URL handle (@sermonindex → 'Sermonindex').
        tab_label = source_name if source_name else _channel_slug_from_url(url_raw).capitalize()

        if args.only and tab_label.lower() != args.only.lower():
            continue

        print(f"\n{'─' * 64}")
        print(f"  Queue row {row_idx}: {tab_label!r}")
        print(f"  url:    {url_raw[:80]}")
        print(f"  review: {review}  filter: {filter_raw!r}  limit: {limit_raw}")

        try:
            # Load whitelist if needed — file resolved by tab_label slug, not source_name,
            # so blank source_name whitelist rows find the right file.
            whitelist_names: List[str] = []
            if whitelist_mode:
                wl_path = _whitelist_path(tab_label)
                if not wl_path.exists():
                    raise FileNotFoundError(
                        f"whitelist file not found: {wl_path}\n"
                        f"  Create it with one speaker name per line."
                    )
                whitelist_names = load_whitelist(str(wl_path))
                print(f"  whitelist: {len(whitelist_names)} names from {wl_path.name}")

            # Groq required for review=yes non-whitelist rows
            effective_groq = groq_client if (review == "yes" and not whitelist_mode) else None

            # Normalize playlist URL (watch?v=X&list=Y → playlist?list=Y)
            url = _normalize_playlist_url(url_raw)
            if url != url_raw:
                print(f"  url normalized → {url[:80]}")

            # Ensure source tab exists (named by tab_label)
            if not args.dry_run:
                _ensure_tab(wb, tab_label)
            else:
                print(f"  [DRY-RUN] would ensure tab: {tab_label!r}")

            # Run triage on the source tab
            if not args.dry_run:
                wb_cur = openpyxl.load_workbook(QUEUE_PATH)
                ws_src = wb_cur[tab_label]
                stats = process_sheet(
                    wb_cur, ws_src, ytdlp, effective_groq,
                    whitelist_names=whitelist_names,
                    min_duration=min_duration,
                    limit=limit,
                    add_url=url,
                    dry_run=False,
                )
                # Reload Queue sheet to update status
                wb_cur2 = openpyxl.load_workbook(QUEUE_PATH)
                ws_q2   = wb_cur2["Queue"]
                ws_q2.cell(row=row_idx, column=QUEUE_COL["status"]).value = "triaged"
                wb_cur2.save(QUEUE_PATH)
                written = stats.get("written", 0)
                print(f"  ✓ triaged → {written} video row(s) written to {tab_label!r}")
                summary.append((tab_label, written, None))
            else:
                print(f"  [DRY-RUN] would run process_sheet on {tab_label!r} "
                      f"(min_duration={min_duration}, whitelist={len(whitelist_names)}, limit={limit})")
                summary.append((tab_label, 0, None))

        except Exception as exc:
            msg = str(exc)
            print(f"  ✗ ERROR: {msg}")
            if not args.dry_run:
                try:
                    wb_err = openpyxl.load_workbook(QUEUE_PATH)
                    ws_err = wb_err["Queue"]
                    ws_err.cell(row=row_idx, column=QUEUE_COL["status"]).value = f"error: {msg[:60]}"
                    wb_err.save(QUEUE_PATH)
                except Exception:
                    pass
            summary.append((tab_label, 0, msg))

    # ── Summary ───────────────────────────────────────────────────────────────
    elapsed = (time.time() - start_time) / 60
    print(f"\n{'═' * 64}")
    print("QUEUE TRIAGE SUMMARY")
    print(f"{'═' * 64}")
    total_written = 0
    for name, written, err in summary:
        status = f"ERROR: {err[:50]}" if err else f"{written} video(s) written"
        print(f"  {name:<30} {status}")
        total_written += written
    print(f"{'─' * 64}")
    print(f"  Total rows processed: {len(summary)}")
    print(f"  Total videos written: {total_written}")
    print(f"  Elapsed: {elapsed:.1f} min")
    print()


if __name__ == "__main__":
    main()
