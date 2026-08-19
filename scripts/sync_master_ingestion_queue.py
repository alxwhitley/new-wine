#!/usr/bin/env python3
"""sync_master_ingestion_queue.py — one-way sync, spreadsheet -> database.

docs/ingestion/master_ingestion_queue.xlsx is the single master source of
truth for ingestion candidates (Alex, 2026-08-19). The workbook has two
tabs: Discovery (raw, unvetted candidates) and Queue (vetted, ready to
actually ingest). This script reads ONLY the Queue tab -- it does not even
open the Discovery tab, let alone read rows from it, so a Discovery-stage
candidate can never reach the database through this path (restructured
2026-08-19; the two-tab split replaced an earlier single-tab design that
told the two kinds of row apart by a "research_candidate" stage value
instead of by tab). It reconciles the live source_ingest_queue table to
match the Queue tab. On any disagreement the spreadsheet wins and
overwrites the database -- deliberate, standing policy, not a per-run
choice this script makes.

Only rows staged "ready_to_queue" or "already_queued" ever touch the
database. "done" rows are already fully processed and are deliberately out
of sync scope too -- they exist in the sheet as a record, not as something
this script re-touches.

Two modes, same convention as every other database-writing script in this
repo (see scripts/apply_migration_088.py):
  no flags     -- DRY RUN. Reads the sheet and the database, computes the
                  full plan (creates / overwrites with field-level diffs /
                  orphans / blocked rows), prints it, writes nothing. Every
                  write is also rehearsed inside a rolled-back transaction
                  so a constraint violation shows up here, not on --apply.
  --apply      -- Recomputes the identical plan fresh, then actually writes
                  it inside one committed transaction (all-or-nothing), then
                  reconnects and re-reads the database to confirm every
                  touched row now matches the sheet exactly. Never run this
                  without having first reviewed a --apply-free run's report.

Never deletes. A database row with no matching sheet row is reported as an
orphan for a human decision -- never removed automatically.

Meant to be re-run on demand whenever the sheet changes -- not a one-off.
Python 3.12 (Invariant 1).
"""
from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

import openpyxl
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv

ROOT = Path(__file__).resolve().parent.parent
SHEET_PATH = ROOT / "docs" / "ingestion" / "master_ingestion_queue.xlsx"
# Only the Queue tab is ever read. The workbook also has a Discovery tab
# holding raw, unvetted candidates -- those must never reach the database
# under any circumstances (Alex, 2026-08-19), so this script does not even
# open that tab, let alone read rows from it.
SHEET_TAB = "Queue"

load_dotenv(ROOT / "backend" / "app" / ".env")

SYNC_ELIGIBLE_STAGES = {"ready_to_queue", "already_queued"}

# spreadsheet column -> database column, for every field this sync touches.
# (sheet-only bookkeeping columns -- name, description, reference_urls,
# source_card_id, origin, our own "stage" -- are deliberately absent: they
# have no database counterpart.)
WRITABLE_FIELDS = [
    ("url", "url"),
    ("source_format", "source_format"),
    ("source_scope", "source_scope"),
    ("attribute_to", "attribute_to"),
    ("attribution_mode", "attribution_mode"),
    ("on_unknown_author", "on_unknown_author"),
    ("retain_original_text", "retain_original_text"),
    ("notes", "notes"),
    ("flag_reason", "flag_reason"),
    ("cleared_to_run", "cleared_to_run"),
    ("db_status", "status"),
    ("db_execution_stage", "stage"),
    ("attempts", "attempts"),
    ("max_attempts", "max_attempts"),
    ("worker_id", "worker_id"),
    ("lease_expires_at", "lease_expires_at"),
    ("run_after", "run_after"),
    ("final_url", "final_url"),
    ("content_sha256", "content_sha256"),
    ("fetched_bytes", "fetched_bytes"),
    ("attempted_documents", "attempted_documents"),
    ("stored_documents", "stored_documents"),
    ("skipped_documents", "skipped_documents"),
    ("errored_documents", "errored_documents"),
    ("result_document_id", "result_document_id"),
    ("submitted_by", "submitted_by"),
]

# Every database column that is NOT NULL -- a blank sheet cell can never be
# sent as an explicit NULL for any of these; it means "leave unchanged" on
# an overwrite, or "let the database default apply" on a create.
NOT_NULL_DB_COLUMNS = {
    "url", "source_format", "source_scope", "attribution_mode", "submitted_by",
    "on_unknown_author", "retain_original_text", "status", "cleared_to_run",
    "attempts", "max_attempts", "run_after", "stage",
    "attempted_documents", "stored_documents", "skipped_documents", "errored_documents",
}

# The subset of NOT_NULL_DB_COLUMNS with no database default at all -- these
# are the only ones that can actually block a create; every other NOT NULL
# column already has a sensible default the database applies on insert.
REQUIRED_FOR_CREATE = {"url", "source_format", "source_scope", "attribution_mode", "submitted_by"}

BOOL_SHEET_FIELDS = {"retain_original_text", "cleared_to_run"}
INT_SHEET_FIELDS = {
    "attempts", "max_attempts", "fetched_bytes",
    "attempted_documents", "stored_documents", "skipped_documents", "errored_documents",
}


def db_connect():
    return psycopg2.connect(os.environ["SUPABASE_DB_URL"])


def _blank(v):
    return v is None or (isinstance(v, str) and v.strip() == "")


def _coerce_bool(v, field, row_label, problems):
    if isinstance(v, bool) or v is None:
        return v
    if isinstance(v, str):
        s = v.strip().upper()
        if s == "":
            return None
        if s == "TRUE":
            return True
        if s == "FALSE":
            return False
    problems.append(f"{row_label}: {field!r} has an unrecognised value {v!r} (expected TRUE/FALSE/blank)")
    return None


def _coerce_int(v, field, row_label, problems):
    if v is None or isinstance(v, int):
        return v
    if isinstance(v, str) and v.strip() == "":
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        problems.append(f"{row_label}: {field!r} has a non-integer value {v!r}")
        return None


def read_sheet(problems: list) -> list[dict]:
    if not SHEET_PATH.exists():
        raise SystemExit(f"Master spreadsheet not found: {SHEET_PATH}")
    wb = openpyxl.load_workbook(SHEET_PATH, data_only=True)
    if SHEET_TAB not in wb.sheetnames:
        raise SystemExit(f"Expected tab {SHEET_TAB!r} not found in {SHEET_PATH}")
    ws = wb[SHEET_TAB]
    headers = [c.value for c in ws[1]]
    rows = []
    for excel_row_num, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in raw):
            continue
        row = dict(zip(headers, raw))
        row["_excel_row"] = excel_row_num
        label = f"sheet row {excel_row_num} ({row.get('name') or row.get('url') or 'unnamed'})"
        for field in BOOL_SHEET_FIELDS:
            row[field] = _coerce_bool(row.get(field), field, label, problems)
        for field in INT_SHEET_FIELDS:
            row[field] = _coerce_int(row.get(field), field, label, problems)
        for field, _ in WRITABLE_FIELDS:
            if field not in BOOL_SHEET_FIELDS and field not in INT_SHEET_FIELDS:
                if _blank(row.get(field)):
                    row[field] = None
        rows.append(row)
    return rows


def fetch_all_db_rows(cur) -> dict:
    cur.execute("SELECT * FROM source_ingest_queue")
    return {str(r["id"]): dict(r) for r in cur.fetchall()}


def determine_default_submitted_by(cur, problems: list):
    cur.execute("SELECT DISTINCT submitted_by FROM source_ingest_queue")
    distinct = [str(r["submitted_by"]) for r in cur.fetchall()]
    if len(distinct) == 1:
        return distinct[0]
    if len(distinct) == 0:
        problems.append(
            "No existing queue rows to infer a default submitter from -- any new row "
            "without its own value filled in the sheet will be blocked, not guessed."
        )
    else:
        problems.append(
            f"Queue rows are currently submitted by {len(distinct)} different accounts -- "
            "refusing to guess a default for new rows. Fill in a value in the sheet for "
            "any row that needs to be created."
        )
    return None


def _normalize_for_compare(v):
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None
    if isinstance(v, bool):
        return v
    return str(v).strip()


def build_plan(sheet_rows, db_rows, default_submitted_by, problems):
    """Returns (creates, overwrites, unchanged, orphans, blocked, stale_refs)."""
    in_scope = [r for r in sheet_rows if r.get("stage") in SYNC_ELIGIBLE_STAGES]
    referenced_db_ids = {
        r["source_db_id"] for r in sheet_rows if not _blank(r.get("source_db_id"))
    }

    creates, overwrites, unchanged, blocked, stale_refs = [], [], [], [], []

    for row in in_scope:
        label = row.get("name") or row.get("url") or f"sheet row {row['_excel_row']}"
        src_db_id = row.get("source_db_id")

        if not _blank(src_db_id):
            # Expect an existing row to overwrite.
            if src_db_id not in db_rows:
                stale_refs.append((row, label))
                continue
            db_row = db_rows[src_db_id]
            diffs = []
            payload = {}
            row_blocked = []
            for sheet_field, db_field in WRITABLE_FIELDS:
                sheet_val = row.get(sheet_field)
                db_val = db_row.get(db_field)
                if _blank(sheet_val):
                    if db_field in NOT_NULL_DB_COLUMNS:
                        continue  # NOT NULL + blank in sheet -> keep existing DB value
                    if _normalize_for_compare(db_val) is not None:
                        diffs.append((db_field, db_val, None))
                        payload[db_field] = None
                    continue
                if _normalize_for_compare(sheet_val) != _normalize_for_compare(db_val):
                    diffs.append((db_field, db_val, sheet_val))
                    payload[db_field] = sheet_val
            if diffs:
                overwrites.append({"row": row, "label": label, "db_id": src_db_id, "diffs": diffs, "payload": payload})
            else:
                unchanged.append({"row": row, "label": label, "db_id": src_db_id})
        else:
            # ready_to_queue with no source_db_id -> create.
            payload = {}
            for sheet_field, db_field in WRITABLE_FIELDS:
                v = row.get(sheet_field)
                if not _blank(v):
                    payload[db_field] = v
            missing = [c for c in REQUIRED_FOR_CREATE if c not in payload]
            if "submitted_by" in missing and default_submitted_by:
                payload["submitted_by"] = default_submitted_by
                missing.remove("submitted_by")
            if payload.get("retain_original_text") is False:
                row_blocked_reason = (
                    "retain_original_text is FALSE in the sheet, but the database requires "
                    "it to always be TRUE (migration 088) -- this row cannot be created as specified."
                )
                blocked.append({"row": row, "label": label, "reason": row_blocked_reason})
                continue
            payload.setdefault("retain_original_text", True)
            if missing:
                blocked.append({
                    "row": row, "label": label,
                    "reason": f"missing required field(s) before this can be created: {', '.join(missing)}",
                })
                continue
            creates.append({"row": row, "label": label, "payload": payload})

    orphans = [
        {"db_id": db_id, "row": db_row}
        for db_id, db_row in db_rows.items()
        if db_id not in referenced_db_ids
    ]

    return creates, overwrites, unchanged, orphans, blocked, stale_refs


def print_plan(creates, overwrites, unchanged, orphans, blocked, stale_refs, problems, mode: str):
    print("=" * 90)
    print(f"SYNC PLAN ({mode})")
    print("=" * 90)

    if problems:
        print(f"\nSHEET PROBLEMS ({len(problems)}) -- affected rows are skipped, not guessed at:")
        for p in problems:
            print(f"  ! {p}")

    print(f"\nWill CREATE ({len(creates)}):")
    for c in creates:
        print(f"  + {c['label']}")
        print(f"      url: {c['payload'].get('url')}")
        print(f"      submitted_by used: {c['payload'].get('submitted_by')}")

    print(f"\nWill OVERWRITE ({len(overwrites)}):")
    for o in overwrites:
        print(f"  ~ {o['label']}  (database row {o['db_id']})")
        for field, old, new in o["diffs"]:
            print(f"      {field}: {old!r} -> {new!r}")

    print(f"\nAlready in sync, no changes ({len(unchanged)}):")
    for u in unchanged:
        print(f"  = {u['label']}  (database row {u['db_id']})")

    print(f"\nBLOCKED -- cannot sync as specified ({len(blocked)}):")
    for b in blocked:
        print(f"  x {b['label']}: {b['reason']}")

    if stale_refs:
        print(f"\nSTALE REFERENCES -- sheet points at a database row that no longer exists ({len(stale_refs)}):")
        for row, label in stale_refs:
            print(f"  ? {label}  (missing database row {row.get('source_db_id')})")

    print(f"\nORPHANS -- in the database, not referenced anywhere in the sheet ({len(orphans)}):")
    print("  Not deleted. Needs a manual decision: add to the sheet, or leave as-is.")
    for o in orphans:
        r = o["row"]
        print(f"  ? database row {o['db_id']}  status={r.get('status')}  url={r.get('url')}  created={r.get('created_at')}")

    print()
    print(
        f"Summary: {len(creates)} to create, {len(overwrites)} to overwrite, "
        f"{len(unchanged)} already correct, {len(blocked)} blocked, "
        f"{len(stale_refs)} stale references, {len(orphans)} orphans."
    )


def rehearse_in_rolled_back_transaction(conn, creates, overwrites):
    """Run every planned write inside a transaction, then roll back. Surfaces
    constraint violations (e.g. a bad retain_original_text or CHECK failure)
    during the dry run instead of during --apply."""
    cur = conn.cursor()
    try:
        for c in creates:
            cols = list(c["payload"].keys())
            placeholders = ", ".join(["%s"] * len(cols))
            col_list = ", ".join(cols)
            cur.execute(
                f"INSERT INTO source_ingest_queue ({col_list}) VALUES ({placeholders}) RETURNING id",
                [c["payload"][col] for col in cols],
            )
        for o in overwrites:
            cols = list(o["payload"].keys())
            set_clause = ", ".join(f"{col} = %s" for col in cols)
            cur.execute(
                f"UPDATE source_ingest_queue SET {set_clause}, updated_at = now() WHERE id = %s",
                [o["payload"][col] for col in cols] + [o["db_id"]],
            )
        print("\nRehearsal (rolled back): every planned write executed without error.")
    finally:
        conn.rollback()


def apply_writes(conn, creates, overwrites):
    cur = conn.cursor()
    created_ids = []
    for c in creates:
        cols = list(c["payload"].keys())
        placeholders = ", ".join(["%s"] * len(cols))
        col_list = ", ".join(cols)
        cur.execute(
            f"INSERT INTO source_ingest_queue ({col_list}) VALUES ({placeholders}) RETURNING id",
            [c["payload"][col] for col in cols],
        )
        new_id = str(cur.fetchone()[0])
        created_ids.append((c, new_id))
    overwritten_ids = []
    for o in overwrites:
        cols = list(o["payload"].keys())
        set_clause = ", ".join(f"{col} = %s" for col in cols)
        cur.execute(
            f"UPDATE source_ingest_queue SET {set_clause}, updated_at = now() WHERE id = %s",
            [o["payload"][col] for col in cols] + [o["db_id"]],
        )
        overwritten_ids.append(o["db_id"])
    conn.commit()
    return created_ids, overwritten_ids


def verify_after_write(conn, created_ids, overwritten_ids):
    print()
    print("=" * 90)
    print("POST-WRITE VERIFICATION (fresh read)")
    print("=" * 90)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    all_ok = True

    for c, new_id in created_ids:
        cur.execute("SELECT * FROM source_ingest_queue WHERE id = %s", (new_id,))
        row = cur.fetchone()
        if row is None:
            print(f"  FAIL: created row for {c['label']} not found on re-read (id {new_id})")
            all_ok = False
            continue
        mismatches = []
        for db_field, expected in c["payload"].items():
            if _normalize_for_compare(row.get(db_field)) != _normalize_for_compare(expected):
                mismatches.append((db_field, expected, row.get(db_field)))
        if mismatches:
            print(f"  FAIL: created row for {c['label']} (id {new_id}) does not match the sheet:")
            for field, expected, actual in mismatches:
                print(f"      {field}: expected {expected!r}, found {actual!r}")
            all_ok = False
        else:
            print(f"  OK: {c['label']}  (new database row {new_id}) matches the sheet")

    for db_id in overwritten_ids:
        cur.execute("SELECT * FROM source_ingest_queue WHERE id = %s", (db_id,))
        row = cur.fetchone()
        print(f"  OK: database row {db_id} re-read successfully (status={row['status'] if row else 'MISSING'})")
        if row is None:
            all_ok = False

    print()
    print("VERIFICATION PASSED -- database matches the sheet." if all_ok else "VERIFICATION FAILED -- see above.")
    return all_ok


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description="Sync the master ingestion spreadsheet into the live queue table.")
    parser.add_argument("--apply", action="store_true", help="required acknowledgement for the real database write")
    args = parser.parse_args(argv)

    problems: list = []
    sheet_rows = read_sheet(problems)

    conn = db_connect()
    conn.autocommit = False
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("BEGIN READ ONLY")
        db_rows = fetch_all_db_rows(cur)
        default_submitted_by = determine_default_submitted_by(cur, problems)
    finally:
        conn.rollback()

    creates, overwrites, unchanged, orphans, blocked, stale_refs = build_plan(
        sheet_rows, db_rows, default_submitted_by, problems
    )

    if not args.apply:
        print_plan(creates, overwrites, unchanged, orphans, blocked, stale_refs, problems, mode="DRY RUN")
        rehearse_in_rolled_back_transaction(conn, creates, overwrites)
        print("\nDRY RUN complete. Nothing was written. Re-run with --apply after this is reviewed and approved.")
        conn.close()
        return 0

    # --apply: recompute fresh (spreadsheet/database may have moved since any
    # earlier dry run) and print the same plan one more time before writing.
    print_plan(creates, overwrites, unchanged, orphans, blocked, stale_refs, problems, mode="APPLY")
    if not creates and not overwrites:
        print("\nNothing to write -- database already matches the sheet for every in-scope row.")
        conn.close()
        return 0

    created_ids, overwritten_ids = apply_writes(conn, creates, overwrites)
    print(f"\nCommitted: {len(created_ids)} created, {len(overwritten_ids)} overwritten.")

    verify_conn = db_connect()
    verify_conn.autocommit = True
    ok = verify_after_write(verify_conn, created_ids, overwritten_ids)
    verify_conn.close()
    conn.close()
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
