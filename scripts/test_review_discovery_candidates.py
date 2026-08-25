#!/usr/bin/env python3
"""Regression tests for review_discovery_candidates.py's pure filtering
logic and its Discovery/Approved Sites read-modify-write behavior. Runs
entirely against a throwaway temp workbook -- never
docs/ingestion/master_ingestion_queue.xlsx.

Run: python3.12 scripts/test_review_discovery_candidates.py
"""
from __future__ import annotations

import shutil
import sys
import tempfile
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))

import review_discovery_candidates as tool

_checks = []
_failures = []


def check(label, condition):
    _checks.append(label)
    if not condition:
        _failures.append(label)
        print(f"FAIL: {label}")


# ---------------------------------------------------------------------------
# _candidate_link -- pure, no I/O
# ---------------------------------------------------------------------------
check(
    "prefers claimed_blog_or_articles_url",
    tool._candidate_link({"claimed_blog_or_articles_url": "https://a.example.com", "claimed_main_url": "https://b.example.com"}) == "https://a.example.com",
)
check(
    "falls back to claimed_main_url",
    tool._candidate_link({"claimed_main_url": "https://b.example.com"}) == "https://b.example.com",
)
check(
    "falls back to first of other_urls",
    tool._candidate_link({"other_urls": "https://c.example.com; https://d.example.com"}) == "https://c.example.com",
)
check("no link anywhere yields None", tool._candidate_link({}) is None)
check(
    "blank strings treated as absent, falls through to next field",
    tool._candidate_link({"claimed_main_url": "   ", "other_urls": "https://d.example.com"}) == "https://d.example.com",
)

# ---------------------------------------------------------------------------
# build_queue -- pure, no I/O
# ---------------------------------------------------------------------------
_rows = [
    {"name": "Eligible", "verification_status": "unverified", "already_in_corpus": False, "claimed_written_content_exists": True, "claimed_main_url": "https://ok.example.com"},
    {"name": "Already Verified", "verification_status": "verified", "already_in_corpus": False, "claimed_written_content_exists": True, "claimed_main_url": "https://x.example.com"},
    {"name": "Already Rejected", "verification_status": "rejected", "already_in_corpus": False, "claimed_written_content_exists": True, "claimed_main_url": "https://x.example.com"},
    {"name": "In Corpus", "verification_status": "unverified", "already_in_corpus": True, "claimed_written_content_exists": True, "claimed_main_url": "https://x.example.com"},
    {"name": "No Content", "verification_status": "unverified", "already_in_corpus": False, "claimed_written_content_exists": False, "claimed_main_url": "https://x.example.com"},
    {"name": "No Link", "verification_status": "unverified", "already_in_corpus": False, "claimed_written_content_exists": True},
    {"name": "Unknown Content Flag", "verification_status": "unverified", "already_in_corpus": False, "claimed_written_content_exists": None, "claimed_main_url": "https://ok2.example.com"},
]
_queue = tool.build_queue(_rows)
_queue_names = [r["name"] for r, _ in _queue]
check("eligible row included", "Eligible" in _queue_names)
check("already-verified row excluded", "Already Verified" not in _queue_names)
check("already-rejected row excluded", "Already Rejected" not in _queue_names)
check("already-in-corpus row excluded", "In Corpus" not in _queue_names)
check("confirmed-no-content row excluded", "No Content" not in _queue_names)
check("row with no usable link excluded", "No Link" not in _queue_names)
check("unknown/blank content-exists flag included (not confirmed absent)", "Unknown Content Flag" in _queue_names)
check("exactly the two eligible rows returned", len(_queue) == 2)
check("sheet order preserved", _queue_names == ["Eligible", "Unknown Content Flag"])

# ---------------------------------------------------------------------------
# approve_candidate / reject_candidate / next_candidate -- end to end
# against a throwaway workbook. SHEET_PATH / LOCK_PATH are monkeypatched.
# ---------------------------------------------------------------------------
_DISCOVERY_HEADER = ["verification_status", "already_in_corpus", "name", "claimed_main_url", "claimed_blog_or_articles_url", "other_urls", "claimed_written_content_exists"]
_APPROVED_HEADER = ["approved", "name", "attribute_to", "blog_url", "authorship_confidence", "scale_note", "proposal_notes", "approved_at"]


def _build_test_workbook(discovery_rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tool.DISCOVERY_TAB
    ws.append(_DISCOVERY_HEADER)
    for row in discovery_rows:
        ws.append([row.get(h) for h in _DISCOVERY_HEADER])
    approved = wb.create_sheet(tool.APPROVED_TAB)
    approved.append(_APPROVED_HEADER)
    tmpdir = Path(tempfile.mkdtemp())
    path = tmpdir / "test_master_ingestion_queue.xlsx"
    wb.save(path)
    return path


_original_sheet_path = tool.SHEET_PATH
_original_lock_path = tool.LOCK_PATH
_wb_path = _build_test_workbook(
    [
        {"verification_status": "unverified", "already_in_corpus": False, "name": "Alpha", "claimed_main_url": "https://alpha.example.com"},
        {"verification_status": "unverified", "already_in_corpus": False, "name": "Beta", "claimed_main_url": "https://beta.example.com"},
        {"verification_status": "unverified", "already_in_corpus": False, "name": "Already Approved Elsewhere", "claimed_main_url": "https://gamma.example.com"},
    ]
)

try:
    tool.SHEET_PATH = _wb_path
    tool.LOCK_PATH = _wb_path.parent / f"~${_wb_path.name}"

    # Pre-seed Approved Sites with a name that will collide, to exercise the dedupe guard.
    wb = openpyxl.load_workbook(_wb_path)
    approved_ws = wb[tool.APPROVED_TAB]
    approved_ws.append(["TRUE", "Already Approved Elsewhere", "Already Approved Elsewhere", "https://gamma.example.com", None, None, None, "2026-01-01 -- prior manual approval"])
    wb.save(_wb_path)

    first = tool.next_candidate()
    check("next_candidate returns the first eligible row", first is not None and first[0]["name"] == "Alpha")
    check("next_candidate reports the full remaining count", first is not None and first[2] == 3)

    tool.approve_candidate("Alpha", "https://alpha.example.com")

    wb2 = openpyxl.load_workbook(_wb_path, data_only=True)
    discovery_ws2 = wb2[tool.DISCOVERY_TAB]
    header2 = [c.value for c in discovery_ws2[1]]
    check("reviewed_at column was added", "reviewed_at" in header2)
    check("review_notes column was added", "review_notes" in header2)
    idx2 = {h: i for i, h in enumerate(header2)}
    alpha_row = next(r for r in discovery_ws2.iter_rows(min_row=2, values_only=True) if r[idx2["name"]] == "Alpha")
    check("approved Discovery row marked verified", alpha_row[idx2["verification_status"]] == "verified")
    check("approved Discovery row got a reviewed_at stamp", bool(alpha_row[idx2["reviewed_at"]]))
    check("approved Discovery row got a review_notes stamp", bool(alpha_row[idx2["review_notes"]]))

    approved_ws2 = wb2[tool.APPROVED_TAB]
    approved_header2 = [c.value for c in approved_ws2[1]]
    aidx2 = {h: i for i, h in enumerate(approved_header2)}
    alpha_approved = next(r for r in approved_ws2.iter_rows(min_row=2, values_only=True) if r[aidx2["name"]] == "Alpha")
    check("approve wrote a new Approved Sites row with the right blog_url", alpha_approved[aidx2["blog_url"]] == "https://alpha.example.com")
    check("approve set approved=TRUE", str(alpha_approved[aidx2["approved"]]).strip().upper() == "TRUE")

    second = tool.next_candidate()
    check("next_candidate skips the now-approved row", second is not None and second[0]["name"] == "Beta")
    check("remaining count dropped by one", second is not None and second[2] == 2)

    tool.reject_candidate("Beta")
    wb3 = openpyxl.load_workbook(_wb_path, data_only=True)
    discovery_ws3 = wb3[tool.DISCOVERY_TAB]
    header3 = [c.value for c in discovery_ws3[1]]
    idx3 = {h: i for i, h in enumerate(header3)}
    beta_row = next(r for r in discovery_ws3.iter_rows(min_row=2, values_only=True) if r[idx3["name"]] == "Beta")
    check("rejected Discovery row marked rejected", beta_row[idx3["verification_status"]] == "rejected")

    # Approve a candidate whose name already exists in Approved Sites -- should not duplicate.
    tool.approve_candidate("Already Approved Elsewhere", "https://gamma.example.com")
    wb4 = openpyxl.load_workbook(_wb_path, data_only=True)
    approved_ws4 = wb4[tool.APPROVED_TAB]
    approved_header4 = [c.value for c in approved_ws4[1]]
    aidx4 = {h: i for i, h in enumerate(approved_header4)}
    gamma_rows = [r for r in approved_ws4.iter_rows(min_row=2, values_only=True) if r[aidx4["name"]] == "Already Approved Elsewhere"]
    check("approving an already-approved-elsewhere name does not duplicate the Approved Sites row", len(gamma_rows) == 1)
    discovery_ws4 = wb4[tool.DISCOVERY_TAB]
    header4 = [c.value for c in discovery_ws4[1]]
    idx4 = {h: i for i, h in enumerate(header4)}
    gamma_discovery = next(r for r in discovery_ws4.iter_rows(min_row=2, values_only=True) if r[idx4["name"]] == "Already Approved Elsewhere")
    check("but the Discovery row is still marked verified", gamma_discovery[idx4["verification_status"]] == "verified")

    final = tool.next_candidate()
    check("queue is empty once every row is decided", final is None)

    # Lock-file check: both writers must refuse while Excel's own lock file exists.
    tool.LOCK_PATH.write_text("")
    try:
        tool.approve_candidate("Alpha", "https://alpha.example.com")
        check("approve_candidate refuses while the workbook is locked", False)
    except RuntimeError:
        check("approve_candidate refuses while the workbook is locked", True)
    try:
        tool.reject_candidate("Alpha")
        check("reject_candidate refuses while the workbook is locked", False)
    except RuntimeError:
        check("reject_candidate refuses while the workbook is locked", True)
    tool.LOCK_PATH.unlink()

    try:
        tool.approve_candidate("Someone Not In The Sheet", "https://nowhere.example.com")
        check("approving an unknown name raises RuntimeError", False)
    except RuntimeError:
        check("approving an unknown name raises RuntimeError", True)
finally:
    tool.SHEET_PATH = _original_sheet_path
    tool.LOCK_PATH = _original_lock_path
    shutil.rmtree(_wb_path.parent, ignore_errors=True)


print(f"\n{len(_checks) - len(_failures)}/{len(_checks)} checks passed")
if _failures:
    print("\nFAILED:")
    for f in _failures:
        print(f"  - {f}")
    raise SystemExit(1)
