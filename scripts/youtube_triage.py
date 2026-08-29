#!/usr/bin/env python3
"""
youtube_triage.py — Stage 2: metadata triage pass for the unified YouTube ingest pipeline.

Reads/writes sources/youtube/ingest_queue.xlsx. For each row where status is blank:
  - Channel URL (@handle, /channel/, /c/): enumerate all videos via yt-dlp flat-playlist
    (titles + video IDs only, no downloads). Expand into one row per video; mark the
    original channel row status=expanded.
  - Single video URL (watch?v=, youtu.be/): fetch title + channel name via yt-dlp
    (no download).

Then classifies each video_title via Groq (llama-3.3-70b-versatile) into:
    sermon | worship | promo | other

Sets ingest=TRUE when guess=sermon, else ingest=FALSE.
Sets status=triaged on completed video rows.
Leaves resolved_source blank — Stage 3 fills it.

Idempotent: rows with any non-blank status are skipped.

Sheet: sources/youtube/ingest_queue.xlsx
Columns: url | video_title | channel_name | guess | ingest | status | resolved_source

Usage:
    python3 scripts/youtube_triage.py                      # process all new rows in sheet
    python3 scripts/youtube_triage.py --add URL            # add URL then triage
    python3 scripts/youtube_triage.py --add URL --limit 20 # cap videos from channel
    python3 scripts/youtube_triage.py --retry-unknown      # re-classify guess=unknown rows
    python3 scripts/youtube_triage.py --dry-run            # print actions, no writes
"""

import argparse
import json
import os
import re
import shutil
import subprocess
import sys
from pathlib import Path
from typing import List, Optional, Tuple

import openpyxl
from dotenv import load_dotenv
from groq import Groq

ROOT         = Path(__file__).resolve().parent.parent
QUEUE_PATH   = ROOT / "sources" / "youtube" / "ingest_queue.xlsx"
COOKIES_PATH = ROOT / "scripts" / "youtube_cookies.txt"

load_dotenv(ROOT / "backend" / "app" / ".env")

# normalize_alias_key: the sole normalization contract for source_aliases.
# MUST be imported from source_resolver — do not re-implement here.
_SCRIPTS = Path(__file__).resolve().parent
if str(_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS))
from source_resolver import normalize_alias_key  # noqa: E402

MODEL      = "openai/gpt-oss-120b"
BATCH_SIZE = 10   # titles per Groq call (smaller = less blast radius on mismatch)

COLUMNS = ["url", "video_title", "channel_name", "guess", "ingest", "status", "resolved_source"]
COL     = {name: i + 1 for i, name in enumerate(COLUMNS)}   # 1-based for openpyxl ws.cell()

TRIAGE_SYSTEM = (
    "You are classifying YouTube video titles for a theological research tool. "
    "Classify each title as exactly one of: sermon, worship, promo, other.\n\n"
    "sermon  — teaching, preaching, Bible study, Q&A, conference session, lecture, interview about theology\n"
    "worship — music, song, praise, worship set, hymn, choir\n"
    "promo   — trailer, announcement, highlights, teaser, channel intro, behind-the-scenes, event promo\n"
    "other   — anything not clearly covered above\n\n"
    'Input: a JSON array of objects: [{"i": 1, "title": "..."}, {"i": 2, "title": "..."}, ...]\n'
    'Output: a JSON array of objects: [{"i": 1, "label": "sermon"}, {"i": 2, "label": "other"}, ...]\n'
    "Use the SAME \"i\" value from each input object. Include one output object per input. "
    "No explanation. Return the JSON array only."
)


# ── yt-dlp helpers ────────────────────────────────────────────────────────────

def find_ytdlp() -> Optional[str]:
    candidates = [
        shutil.which("yt-dlp"),
        os.path.expanduser("~/Library/Python/3.9/bin/yt-dlp"),
        os.path.expanduser("~/Library/Python/3.10/bin/yt-dlp"),
        os.path.expanduser("~/Library/Python/3.11/bin/yt-dlp"),
        os.path.expanduser("~/Library/Python/3.12/bin/yt-dlp"),
        os.path.expanduser("~/.local/bin/yt-dlp"),
    ]
    for c in candidates:
        if c and os.path.isfile(c):
            return c
    return None


def _ytdlp_base_args(ytdlp: str) -> List[str]:
    """Common yt-dlp flags: force IPv4, set player client, cookies if present."""
    args = [ytdlp, "-4", "--extractor-args", "youtube:player_client=android_vr,web_safari"]
    if COOKIES_PATH.exists():
        args += ["--cookies", str(COOKIES_PATH)]
    return args


def is_channel_url(url: str) -> bool:
    """True if the URL points to a channel or playlist (enumerable via flat-playlist).
    list= is checked BEFORE watch?v= so that watch?v=X&list=Y is treated as a playlist."""
    url = url.strip()
    if "list=" in url:   # playlist URL — may also contain watch?v=, but it's enumerable
        return True
    return "watch?v=" not in url and "youtu.be/" not in url


_NA_VALS = {"na", "none", "null", ""}


# ── Whitelist helpers ─────────────────────────────────────────────────────────

def load_whitelist(spec: str) -> List[str]:
    """
    Load speaker names from a comma-separated string or a file path (one name per line).
    Lines starting with '#' are treated as comments and ignored.
    Returns list of display strings; normalization is applied at match time.
    """
    p = Path(spec)
    if p.exists():
        names = [
            line.strip()
            for line in p.read_text(encoding="utf-8").splitlines()
            if line.strip() and not line.strip().startswith("#")
        ]
    else:
        names = [n.strip() for n in spec.split(",") if n.strip()]
    return names


def match_whitelist(title: str, wl_entries: List[str]) -> Optional[str]:
    """
    Return the first whitelist entry whose normalized form appears as a substring
    of the normalized title, or None if no entry matches.

    normalize_alias_key() is used for both sides — consistent with alias resolution.
    A period-strip fallback is also tried so that "AW Tozer" (norm: "aw tozer") can
    match a title rendered "A.W. Tozer - ..." (norm strips to "aw tozer - ...").
    This handles the punctuation variants registered for A.W. Tozer and T. Austin-Sparks.
    """
    norm_title        = normalize_alias_key(title)
    norm_title_nodots = norm_title.replace(".", "")
    for name in wl_entries:
        norm_name = normalize_alias_key(name)
        if not norm_name:
            continue
        if norm_name in norm_title:
            return name
        # Period-strip fallback: "aw tozer" in "aw tozer - prayer" when title had "A.W."
        if norm_name.replace(".", "") in norm_title_nodots:
            return name
    return None


def enumerate_channel(
    ytdlp: str, channel_url: str, limit: Optional[int] = None
) -> Tuple[str, List[dict]]:
    """
    Enumerate video IDs + titles + durations from a channel via flat-playlist. No downloads.
    Returns (channel_name, [{"url": ..., "title": ..., "duration": int|None}, ...]).
    duration is None when missing/zero/live (caller must treat None as keep).
    channel_name is taken from the first video entry's channel field.
    """
    cmd = _ytdlp_base_args(ytdlp) + [
        "--flat-playlist",
        "--print", "%(id)s\t%(title)s\t%(channel)s\t%(duration)s",
    ]
    if limit:
        cmd += ["--playlist-items", f"1:{limit}"]
    cmd.append(channel_url)
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
    videos = []
    channel_name = ""
    for line in result.stdout.strip().splitlines():
        parts = line.split("\t", 3)
        if len(parts) < 2:
            continue
        vid_id = parts[0].strip()
        title  = parts[1].strip()
        ch     = parts[2].strip() if len(parts) > 2 else ""
        dur_raw = parts[3].strip() if len(parts) > 3 else ""
        if not vid_id or vid_id in ("NA", "None", "null"):
            continue
        if ch and not channel_name:
            channel_name = ch
        # Parse duration; treat missing/zero/NA as None (fail toward inclusion)
        duration: Optional[int] = None
        if dur_raw and dur_raw.lower() not in _NA_VALS:
            try:
                d = int(float(dur_raw))
                if d > 0:
                    duration = d
            except ValueError:
                pass
        videos.append({
            "url":      f"https://www.youtube.com/watch?v={vid_id}",
            "title":    title,
            "duration": duration,
        })
        if limit and len(videos) >= limit:
            break
    return channel_name, videos


def fetch_video_info(ytdlp: str, video_url: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Fetch (title, channel_name) for a single video. No download.
    Returns (None, None) on failure.
    """
    cmd = _ytdlp_base_args(ytdlp) + [
        "--flat-playlist",
        "--print", "%(title)s\t%(channel)s",
        video_url,
    ]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
    except subprocess.TimeoutExpired:
        return None, None
    for line in result.stdout.strip().splitlines():
        parts = line.split("\t", 1)
        if not parts:
            continue
        title = parts[0].strip()
        ch    = parts[1].strip() if len(parts) > 1 else ""
        if title and title not in ("NA", "None", "null"):
            return title, ch or None
    return None, None


# ── Groq classification ────────────────────────────────────────────────────────

def classify_batch(client: Groq, titles: List[str]) -> List[str]:
    """
    Classify a batch of video titles using keyed matching.

    Sends each title as {"i": N, "title": "..."} so Groq returns {"i": N, "label": "..."}.
    Matches verdicts back by i — a missing verdict for one title yields "unknown" for
    that title only, never discarding verdicts for the rest of the batch.

    On whole-batch parse failure: all titles in the batch → "unknown". Never raises.
    """
    keyed_input = [{"i": i + 1, "title": t} for i, t in enumerate(titles)]
    try:
        resp = client.chat.completions.create(
            model=MODEL,
            max_tokens=512,
            messages=[
                {"role": "system", "content": TRIAGE_SYSTEM},
                {"role": "user",   "content": json.dumps(keyed_input)},
            ],
        )
        raw = (resp.choices[0].message.content or "").strip()
        # Strip markdown code fences — known Groq behavior
        raw = re.sub(r"^```(?:json)?\s*\n?", "", raw)
        raw = re.sub(r"\n?```\s*$", "", raw)
        verdicts = json.loads(raw.strip())
        if not isinstance(verdicts, list):
            raise ValueError(f"expected list, got {type(verdicts).__name__}: {raw[:80]}")
    except Exception as e:
        print(f"  ⚠  Groq batch parse failed ({len(titles)} titles): {e}")
        return ["unknown"] * len(titles)

    # Build i → label map; tolerate i as int or string
    valid = {"sermon", "worship", "promo", "other"}
    label_map = {}
    for item in verdicts:
        if not isinstance(item, dict):
            continue
        try:
            idx = int(item["i"])
        except (KeyError, TypeError, ValueError):
            continue
        label = item.get("label", "")
        if label in valid:
            label_map[idx] = label

    # Resolve by 1-based index; any gap → "unknown" for that title only
    result = [label_map.get(i + 1, "unknown") for i in range(len(titles))]
    missing = sum(1 for g in result if g == "unknown")
    if missing:
        print(f"  ⚠  {missing}/{len(titles)} title(s) had no verdict → unknown")
    return result


# ── Summary printer ──────────────────────────────────────────────────────────

def _print_summary(ws_r) -> None:
    """Print the full results table and guess-count breakdown."""
    video_rows = []
    for r in range(2, ws_r.max_row + 1):
        url = str(gcell(ws_r, r, "url") or "")
        if not url or is_channel_url(url):
            continue
        video_rows.append({
            "title":  str(gcell(ws_r, r, "video_title") or ""),
            "guess":  str(gcell(ws_r, r, "guess") or ""),
            "ingest": str(gcell(ws_r, r, "ingest") or ""),
            "status": str(gcell(ws_r, r, "status") or ""),
        })

    if not video_rows:
        print("  No video rows in queue.")
        return

    W = [54, 9, 6, 8]
    hdr = f"{'TITLE':<{W[0]}} {'GUESS':<{W[1]}} {'INGEST':<{W[2]}} STATUS"
    print(f"\n── Results ──────────────────────────────────────────────────────")
    print(hdr)
    print("─" * len(hdr))
    for row in video_rows:
        print(
            f"{row['title'][:W[0]]:<{W[0]}} "
            f"{row['guess']:<{W[1]}} "
            f"{row['ingest']:<{W[2]}} "
            f"{row['status']}"
        )

    counts = {}
    for row in video_rows:
        g = row["guess"]
        counts[g] = counts.get(g, 0) + 1

    print(f"\nTotal: {len(video_rows)} video(s)")
    for label, n in sorted(counts.items()):
        tick = " ← pre-ticked TRUE" if label == "sermon" else ""
        print(f"  {label:<8}: {n}{tick}")


# ── Sheet helpers ─────────────────────────────────────────────────────────────

def load_or_create_wb(path: Path, sheet_name: str) -> Tuple[openpyxl.Workbook, openpyxl.worksheet.worksheet.Worksheet]:
    """Load workbook and select the named sheet. Creates the file if absent. Returns (wb, ws)."""
    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb[sheet_name]
    else:
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(COLUMNS)
        wb.save(path)
        print(f"Created {path}")
    return wb, ws


def gcell(ws, row_idx: int, col_name: str):
    return ws.cell(row=row_idx, column=COL[col_name]).value


def scell(ws, row_idx: int, col_name: str, value) -> None:
    ws.cell(row=row_idx, column=COL[col_name], value=value)


def is_blank_status(ws, row_idx: int) -> bool:
    v = gcell(ws, row_idx, "status")
    return v is None or str(v).strip() == ""


def all_urls_in_sheet(ws) -> set:
    return {
        str(ws.cell(row=r, column=COL["url"]).value or "").strip()
        for r in range(2, ws.max_row + 1)
    }


# ── Callable pipeline (used by CLI and queue orchestrator) ───────────────────

def process_sheet(
    wb: openpyxl.Workbook,
    ws,
    ytdlp: str,
    groq_client,
    whitelist_names: List[str],
    min_duration: int = 0,
    limit: Optional[int] = None,
    add_url: Optional[str] = None,
    dry_run: bool = False,
) -> dict:
    """
    Run triage Phases 1-3 on the given (wb, ws) pair.
    Called by CLI main() and by run_queue_triage.py.
    Returns stats dict: {written, classified, dur_skipped, wl_no_match, speaker_counts}.
    """
    stats: dict = {
        "written": 0, "classified": 0,
        "dur_skipped": 0, "wl_no_match": 0, "speaker_counts": {},
    }

    # ── Pre-step: add URL if requested ───────────────────────────────────────
    if add_url:
        url_to_add = add_url.strip()
        if url_to_add in all_urls_in_sheet(ws):
            print(f"Already in queue: {url_to_add}")
        else:
            ws.append([url_to_add] + [""] * (len(COLUMNS) - 1))
            if not dry_run:
                wb.save(QUEUE_PATH)
            print(f"Added: {url_to_add}")

    # ── Phase 1: Channel expansion ────────────────────────────────────────────
    print("\n── Phase 1: Channel expansion ──────────────────────────────────")
    channel_rows = [
        r for r in range(2, ws.max_row + 1)
        if gcell(ws, r, "url") and is_blank_status(ws, r)
        and is_channel_url(str(gcell(ws, r, "url")))
    ]
    print(f"  {len(channel_rows)} channel row(s) to expand")

    dur_total_skipped   = 0
    wl_total_enumerated = 0
    wl_total_no_match   = 0
    wl_total_written    = 0
    wl_speaker_counts: dict = {}

    for row_idx in channel_rows:
        channel_url = str(gcell(ws, row_idx, "url")).strip()
        print(f"\n  {channel_url}")
        try:
            channel_name, videos = enumerate_channel(ytdlp, channel_url, limit=limit)
        except Exception as e:
            print(f"  ✗ enumeration failed: {e}")
            continue

        print(f"  → {len(videos)} video(s) enumerated | channel: {channel_name!r}")
        wl_total_enumerated += len(videos)

        # Duration pre-filter: skip videos whose duration is known and <= threshold.
        if min_duration > 0:
            kept, n_skipped = [], 0
            for v in videos:
                d = v.get("duration")
                if d is not None and d <= min_duration:
                    n_skipped += 1
                else:
                    kept.append(v)
            if n_skipped:
                print(f"  ↳ duration filter (≤{min_duration}s): kept {len(kept)}, skipped {n_skipped}")
            dur_total_skipped += n_skipped
            videos = kept

        # Whitelist filter: keep only videos whose title contains a whitelisted name.
        if whitelist_names:
            matched_videos, n_no_match = [], 0
            for v in videos:
                hit = match_whitelist(v["title"], whitelist_names)
                if hit:
                    v["wl_match"] = hit
                    matched_videos.append(v)
                    wl_speaker_counts[hit] = wl_speaker_counts.get(hit, 0) + 1
                else:
                    n_no_match += 1
            if n_no_match:
                print(f"  ↳ whitelist filter: kept {len(matched_videos)}, skipped {n_no_match} (no speaker match)")
            wl_total_no_match += n_no_match
            videos = matched_videos
            # wl_total_written updated below after dedup

        if dry_run:
            for v in videos[:5]:
                if whitelist_names:
                    print(f"    [{v.get('wl_match', '?')}] {v['title'][:60]}")
                else:
                    print(f"    {v['title'][:72]}")
            if len(videos) > 5:
                print(f"    … and {len(videos) - 5} more")
            continue

        # Dedup: skip URLs already in the sheet so re-runs with a larger limit
        # are purely additive and never double-write existing rows.
        existing_urls = all_urls_in_sheet(ws)
        new_videos = [v for v in videos if v["url"] not in existing_urls]
        n_dupes = len(videos) - len(new_videos)
        if n_dupes:
            print(f"  ↳ dedup: {n_dupes} already in tab, skipping")

        if whitelist_names:
            for v in new_videos:
                ws.append([v["url"], v["title"], v["wl_match"], "sermon", "TRUE", "triaged", ""])
        else:
            for v in new_videos:
                ws.append([v["url"], v["title"], channel_name, "", "", "", ""])
        wl_total_written += len(new_videos)

        scell(ws, row_idx, "channel_name", channel_name)
        scell(ws, row_idx, "status", "expanded")
        wb.save(QUEUE_PATH)
        print(f"  ✓ appended {len(new_videos)} new row(s), marked expanded")

    if min_duration > 0 and channel_rows and not dry_run and not whitelist_names:
        print(f"\n  Duration filter summary: {wl_total_written} written, {dur_total_skipped} skipped (≤{min_duration}s)")

    if whitelist_names and channel_rows and not dry_run:
        print(f"\n  ── Whitelist filter summary ──────────────────────────────")
        print(f"     enumerated           : {wl_total_enumerated}")
        print(f"     skipped (≤{min_duration}s duration) : {dur_total_skipped}")
        print(f"     skipped (no match)   : {wl_total_no_match}")
        print(f"     written              : {wl_total_written}")
        if wl_speaker_counts:
            print(f"\n     Per speaker:")
            for spk, cnt in sorted(wl_speaker_counts.items(), key=lambda x: -x[1]):
                print(f"       {spk}: {cnt}")

    stats["written"]        = wl_total_written
    stats["dur_skipped"]    = dur_total_skipped
    stats["wl_no_match"]    = wl_total_no_match
    stats["speaker_counts"] = wl_speaker_counts

    # ── Phase 2: Title fetch for bare single-video rows ───────────────────────
    print("\n── Phase 2: Fetch titles for untitled video rows ───────────────")
    needs_title = [
        r for r in range(2, ws.max_row + 1)
        if gcell(ws, r, "url") and is_blank_status(ws, r)
        and not is_channel_url(str(gcell(ws, r, "url")))
        and not gcell(ws, r, "video_title")
    ]
    print(f"  {len(needs_title)} row(s) need title fetch")

    for row_idx in needs_title:
        url = str(gcell(ws, row_idx, "url")).strip()
        title, ch = fetch_video_info(ytdlp, url)
        if title:
            print(f"  ✓ {title[:70]}")
            if not dry_run:
                scell(ws, row_idx, "video_title",  title)
                scell(ws, row_idx, "channel_name", ch or "")
        else:
            print(f"  ⚠  could not fetch title: {url[:70]}")

    if needs_title and not dry_run:
        wb.save(QUEUE_PATH)

    # ── Phase 3: Groq classification ──────────────────────────────────────────
    print("\n── Phase 3: Groq classification ────────────────────────────────")
    if whitelist_names:
        print("  Whitelist mode — Groq skipped (matched rows pre-classified as sermon).")
        return stats

    to_triage = [
        r for r in range(2, ws.max_row + 1)
        if gcell(ws, r, "url") and is_blank_status(ws, r)
        and not is_channel_url(str(gcell(ws, r, "url")))
        and gcell(ws, r, "video_title")
        and not gcell(ws, r, "guess")
    ]
    print(f"  {len(to_triage)} row(s) to classify")

    for batch_start in range(0, len(to_triage), BATCH_SIZE):
        batch = to_triage[batch_start: batch_start + BATCH_SIZE]
        titles = [str(gcell(ws, r, "video_title")) for r in batch]
        n = len(titles)
        print(f"  Groq: classifying {n} title(s) (rows {batch_start+1}–{batch_start+n})…")
        guesses = classify_batch(groq_client, titles)

        for row_idx, guess in zip(batch, guesses):
            ingest_val = "TRUE" if guess == "sermon" else "FALSE"
            if not dry_run:
                scell(ws, row_idx, "guess",  guess)
                scell(ws, row_idx, "ingest", ingest_val)
                scell(ws, row_idx, "status", "triaged")
            else:
                t = str(gcell(ws, row_idx, "video_title"))[:55]
                print(f"    {guess:<8} {ingest_val}  {t}")

    if to_triage and not dry_run:
        wb.save(QUEUE_PATH)
        print(f"  ✓ saved {len(to_triage)} classified row(s)")

    stats["classified"] = len(to_triage)
    return stats


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="YouTube triage — Stage 2 of the unified ingest pipeline"
    )
    parser.add_argument("--sheet",        metavar="NAME",           help="Tab name to operate on (required)")
    parser.add_argument("--add",          metavar="URL",            help="Add a channel or video URL before triaging")
    parser.add_argument("--limit",        metavar="N",   type=int,  help="Max videos to enumerate per channel (applied before duration filter)")
    parser.add_argument("--min-duration", metavar="N",   type=int,  default=0,
                        help="Skip videos whose duration is known and <= N seconds (0 = off)")
    parser.add_argument("--whitelist",    metavar="NAMES_OR_FILE",
                        help="Comma-separated speaker names, or path to a text file (one name per line). "
                             "Only videos whose title contains a whitelisted name are written; "
                             "matched rows are pre-classified as sermon=TRUE. Groq is skipped.")
    parser.add_argument("--retry-unknown", action="store_true",     help="Re-classify rows where status=triaged AND guess=unknown")
    parser.add_argument("--dry-run",       action="store_true",     help="Print actions without writing")
    args = parser.parse_args()

    # ── Sheet validation ──────────────────────────────────────────────────────
    if QUEUE_PATH.exists():
        _wb_check = openpyxl.load_workbook(QUEUE_PATH, read_only=True)
        _available = _wb_check.sheetnames
        _wb_check.close()
        if not args.sheet:
            print("ERROR: --sheet is required. Available tabs: {}".format(_available))
            sys.exit(1)
        if args.sheet not in _available:
            print("ERROR: sheet {!r} not found. Available tabs: {}".format(args.sheet, _available))
            sys.exit(1)
    elif not args.sheet:
        print("ERROR: --sheet is required.")
        sys.exit(1)

    ytdlp = find_ytdlp()
    if not ytdlp:
        print("ERROR: yt-dlp not found. Run: pip3 install yt-dlp")
        sys.exit(1)

    # Load whitelist if provided
    whitelist_names: List[str] = []
    if args.whitelist:
        whitelist_names = load_whitelist(args.whitelist)
        if not whitelist_names:
            print("ERROR: --whitelist resolved to an empty list")
            sys.exit(1)
        print(f"Whitelist: {len(whitelist_names)} name(s): {', '.join(whitelist_names)}")

    api_key = os.environ.get("GROQ_API_KEY")
    if not api_key:
        print("ERROR: GROQ_API_KEY not set in backend/app/.env")
        sys.exit(1)
    groq_client = Groq(api_key=api_key)

    wb, ws = load_or_create_wb(QUEUE_PATH, args.sheet)

    # ── Retry-unknown mode ────────────────────────────────────────────────────
    if args.retry_unknown:
        unknown_rows = [
            r for r in range(2, ws.max_row + 1)
            if str(gcell(ws, r, "status") or "").strip() == "triaged"
            and str(gcell(ws, r, "guess") or "").strip() == "unknown"
            and gcell(ws, r, "video_title")
        ]
        print(f"\n── Retry-unknown: {len(unknown_rows)} row(s) with guess=unknown ──────")
        if not unknown_rows:
            print("  Nothing to retry.")
        else:
            for batch_start in range(0, len(unknown_rows), BATCH_SIZE):
                batch = unknown_rows[batch_start: batch_start + BATCH_SIZE]
                titles = [str(gcell(ws, r, "video_title")) for r in batch]
                n = len(titles)
                print(f"  Groq: re-classifying {n} title(s) (rows {batch_start+1}–{batch_start+n})…")
                guesses = classify_batch(groq_client, titles)
                for row_idx, guess in zip(batch, guesses):
                    ingest_val = "TRUE" if guess == "sermon" else "FALSE"
                    if not args.dry_run:
                        scell(ws, row_idx, "guess",  guess)
                        scell(ws, row_idx, "ingest", ingest_val)
                        # status stays "triaged" — only guess+ingest change
                    else:
                        t = str(gcell(ws, row_idx, "video_title"))[:55]
                        print(f"    {guess:<8} {ingest_val}  {t}")
            if not args.dry_run:
                wb.save(QUEUE_PATH)
                print(f"  ✓ saved {len(unknown_rows)} updated row(s)")
        # Fall through to summary table
        wb_r = openpyxl.load_workbook(QUEUE_PATH) if (not args.dry_run and QUEUE_PATH.exists()) else wb
        ws_r = wb_r[args.sheet]
        _print_summary(ws_r)
        return

    process_sheet(
        wb, ws, ytdlp, groq_client,
        whitelist_names=whitelist_names,
        min_duration=args.min_duration,
        limit=args.limit,
        add_url=args.add,
        dry_run=args.dry_run,
    )
    wb_r = openpyxl.load_workbook(QUEUE_PATH) if (not args.dry_run and QUEUE_PATH.exists()) else wb
    _print_summary(wb_r[args.sheet])


if __name__ == "__main__":
    main()
