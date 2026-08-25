#!/usr/bin/env python3
"""site_ingest_crawler.py -- autonomous web-article discovery and ingestion.

Alex's authorized exception (2026-08-25 session) to the standing "DB writes
are attended-only" hard rule: production writes through THIS path may run
unattended, with no per-item human review, provided every deterministic
gate below passes. Any gate failure hard-refuses the write -- never a
guess. Scope of the exception is narrow and does not extend anywhere else:
migrations, admin actions, and every other write path in this repo remain
attended.

Only input surface: the "Approved Sites" tab of
docs/ingestion/master_ingestion_queue.xlsx, and only a row whose `approved`
column is literally TRUE. Alex requires this and controls it directly --
this script does not open the Discovery or Queue tabs at all.

`--site NAME` runs exactly that row. Omitting `--site` runs every row with
approved=TRUE in one invocation, one after another -- each site still goes
through the exact same per-site gates below independently (its own crawl,
its own --max-candidates cap, its own byline check per URL). This is what
lets Alex just check boxes in the Approved Sites tab and rerun the script,
with no separate per-site command and no promotion/sync step in between.

Pipeline, per approved site (looped automatically when --site is omitted):
  1. Crawl the site's blog_url (with pagination, bounded by --max-pages)
     for candidate post URLs -- source_ingest_queue.link_discovery.
  2. Drop any URL that already exists in source_ingest_queue for this
     domain, in any status -- never re-propose already-known work.
  3. Fetch each remaining candidate (source_ingest_queue.fetcher.fetch_html,
     the same SSRF-safe pinned fetch the attended path already uses) and
     run the byline gate (source_ingest_queue.byline_verify): CONFIRMED
     clears it, MISMATCH or UNCONFIRMED quarantines it -- this is the
     automated version of the Vlad/Lana check done by hand earlier this
     session, and it is what makes unattended safe instead of just fast.
  4. Cap at --max-candidates newly-confirmed documents for this run
     (default 1 -- the first real run for any site should prove one URL
     before a larger batch, matching this repo's standing dry-run ->
     isolated-proof -> batch discipline; raise it deliberately afterward).
  5. --apply only: insert each confirmed candidate into source_ingest_queue
     with cleared_to_run=true, then invoke the exact same sanctioned
     single-row path the attended pipeline uses --
     `source_ingest_worker.py --once --row-id <id>` (Invariant 16's
     documented shape) -- as a subprocess, so this script never reaches
     into corpus-writing internals directly and stays byte-identical to
     the already-proven attended path for the actual document write. All
     of the processor's existing gates (license, hidden visibility,
     format/scope, source resolution) still apply unchanged; this script
     only removes the per-item human click in front of them.

Two modes, same convention as every other database-writing script in this
repo:
  no flags   -- DRY RUN. Crawls, fetches, and byline-checks every
                candidate; prints the plan (what would be queued, what
                would be quarantined and why); writes nothing.
  --apply    -- Recomputes the identical plan fresh, then actually inserts
                and processes confirmed rows.

Every run writes a full JSON log under local/<YYYY-MM>/ -- what was
crawled, what cleared, what was quarantined and why, and (on --apply) each
row's final processing outcome. Nobody reads this routinely; it exists so
an unattended run's decisions are reconstructable afterward, the same
posture as quote_verification_log.

Python 3.12 (Invariant 1).
"""

from __future__ import annotations

import argparse
import json
import os
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Optional
from urllib.parse import urlsplit

import openpyxl
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv

ROOT = Path(__file__).resolve().parent.parent
SHEET_PATH = ROOT / "docs" / "ingestion" / "master_ingestion_queue.xlsx"
APPROVED_TAB = "Approved Sites"
WORKER_SCRIPT = ROOT / "scripts" / "source_ingest_worker.py"

sys.path.insert(0, str(Path(__file__).resolve().parent))

from source_ingest_queue.byline_verify import verify_byline  # noqa: E402
from source_ingest_queue.fetcher import (  # noqa: E402
    FetchRejected,
    FetchTransient,
    fetch_html,
)
from source_ingest_queue.html_extract import HtmlRejected, extract_article_bounded  # noqa: E402
from source_ingest_queue.link_discovery import discover_links, same_registrable_host  # noqa: E402
from sync_master_ingestion_queue import determine_default_submitted_by  # noqa: E402

load_dotenv(ROOT / "backend" / "app" / ".env")

_TRUTHY = {"true", "yes", "1"}


def db_connect():
    return psycopg2.connect(os.environ["SUPABASE_DB_URL"])


def _is_approved(match: dict) -> bool:
    return str(match.get("approved") or "").strip().lower() in _TRUTHY


def _has_required_fields(match: dict) -> bool:
    return bool(match.get("blog_url")) and bool(match.get("attribute_to"))


def _read_approved_tab() -> List[dict]:
    """Shared by load_approved_site and load_all_approved_sites -- one place
    that opens the Approved Sites tab, so the two loaders can't drift on how
    they read it."""
    wb = openpyxl.load_workbook(SHEET_PATH, data_only=True)
    if APPROVED_TAB not in wb.sheetnames:
        raise SystemExit(f"'{APPROVED_TAB}' tab not found in {SHEET_PATH}")
    ws = wb[APPROVED_TAB]
    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}
    rows = [{h: row[idx[h]] for h in header} for row in ws.iter_rows(min_row=2, values_only=True) if row[idx["name"]]]
    return rows


def load_approved_site(site_name: str) -> dict:
    """A single named row. Refuses anything not explicitly approved=TRUE by
    Alex in the Approved Sites tab -- never the Discovery or Queue tabs,
    which this function does not even open. Explicit-name lookup fails hard
    on a bad row, since Alex named it deliberately."""
    match = None
    for row in _read_approved_tab():
        if str(row["name"]).strip().lower() == site_name.strip().lower():
            match = row
            break

    if match is None:
        raise SystemExit(
            f"No row named '{site_name}' in the '{APPROVED_TAB}' tab. "
            "This script only reads that tab -- add the site there first."
        )
    if not _is_approved(match):
        raise SystemExit(
            f"'{site_name}' is not approved (approved={match.get('approved')!r}). "
            "Flip the 'approved' column to TRUE in the Approved Sites tab first -- "
            "this script will not crawl an unapproved site."
        )
    if not _has_required_fields(match):
        raise SystemExit(f"'{site_name}' row is missing blog_url or attribute_to.")
    return match


def load_all_approved_sites() -> List[dict]:
    """Every Approved Sites row with approved literally TRUE, in sheet
    order -- the crawler's entire input surface when --site is omitted,
    never the Discovery or Queue tabs. A checked-but-incomplete row is
    skipped with a printed warning rather than aborting the whole run --
    unlike the single-name lookup above, one bad row here shouldn't block
    every other approved site."""
    sites = []
    for row in _read_approved_tab():
        if not _is_approved(row):
            continue
        if not _has_required_fields(row):
            print(f"  [skip] '{row['name']}' is approved but missing blog_url or attribute_to -- fix the row and rerun.")
            continue
        sites.append(row)
    return sites


def crawl_candidate_urls(blog_url: str, *, max_pages: int) -> List[str]:
    seen: List[str] = []
    seen_set = set()
    page_url = blog_url
    page_number = 1
    pages_crawled = 0
    while page_url and pages_crawled < max_pages:
        try:
            fetched = fetch_html(page_url)
        except (FetchRejected, FetchTransient) as exc:
            print(f"  [crawl] page {page_number} ({page_url}) fetch failed: {exc}")
            break
        result = discover_links(fetched.content, fetched.final_url, current_page_number=page_number)
        for url in result.post_urls:
            if url not in seen_set:
                seen_set.add(url)
                seen.append(url)
        pages_crawled += 1
        page_url = result.next_page_url
        page_number += 1
    return seen


def existing_urls_for_domain(cur, sample_url: str) -> set:
    cur.execute("SELECT url FROM source_ingest_queue")
    return {
        row["url"]
        for row in cur.fetchall()
        if same_registrable_host(row["url"], sample_url)
    }


def check_candidate(url: str, declared_author: str) -> dict:
    """Fetch one candidate and run the byline gate. Never raises -- every
    outcome (fetch failure, extraction failure, mismatch, unconfirmed,
    confirmed) is a normal, reportable result."""
    try:
        fetched = fetch_html(url)
    except (FetchRejected, FetchTransient) as exc:
        return {"url": url, "outcome": "fetch_failed", "detail": str(exc)}

    article_text = ""
    try:
        article = extract_article_bounded(fetched.content)
        article_text = article.text
    except HtmlRejected as exc:
        # Byline can still clear via <meta>/JSON-LD even when the article
        # body itself couldn't be isolated -- only the "By <Name>" text
        # fallback needs article_text, so this is not a hard failure here.
        article_extraction_note = str(exc)
    else:
        article_extraction_note = None

    verdict = verify_byline(fetched.content, article_text, declared_author)
    return {
        "url": url,
        "final_url": fetched.final_url,
        "outcome": verdict.status,
        "found_name": verdict.found_name,
        "signal_source": verdict.signal_source,
        "article_extraction_note": article_extraction_note,
    }


def insert_queue_row(cur, *, url: str, attribute_to: str, submitted_by: str, notes: str) -> str:
    cur.execute(
        """
        INSERT INTO source_ingest_queue (
          url, source_format, source_scope, attribute_to, attribution_mode,
          on_unknown_author, retain_original_text, status, cleared_to_run,
          notes, submitted_by
        ) VALUES (
          %s, 'web_page', 'single', %s, 'declared',
          'flag', true, 'waiting', true,
          %s, %s
        )
        RETURNING id::text
        """,
        (url, attribute_to, notes, submitted_by),
    )
    return cur.fetchone()["id"]


def run_worker_once(row_id: str) -> dict:
    proc = subprocess.run(
        [sys.executable, str(WORKER_SCRIPT), "--once", "--row-id", row_id],
        cwd=str(ROOT),
        capture_output=True,
        text=True,
        timeout=300,
    )
    return {"returncode": proc.returncode, "stdout": proc.stdout[-4000:], "stderr": proc.stderr[-4000:]}


def fetch_row_final_state(cur, row_id: str) -> dict:
    cur.execute(
        "SELECT status, stage, flag_reason, stored_documents, skipped_documents, "
        "errored_documents, result_document_id::text FROM source_ingest_queue WHERE id = %s",
        (row_id,),
    )
    row = cur.fetchone()
    return dict(row) if row else {}


def write_run_log(report: dict) -> Path:
    month_dir = ROOT / "local" / datetime.now(timezone.utc).strftime("%Y-%m")
    month_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    site_slug = "".join(c if c.isalnum() else "-" for c in report["site"]).strip("-").lower()
    path = month_dir / f"site_ingest_crawler_{site_slug}_{stamp}.json"
    path.write_text(json.dumps(report, indent=2, default=str))
    return path


def run_for_site(site: dict, args: argparse.Namespace) -> int:
    declared_author = str(site["attribute_to"]).strip()
    blog_url = str(site["blog_url"]).strip()
    print(f"Site: {site['name']}  |  attribute_to: {declared_author}  |  blog_url: {blog_url}")

    print(f"\nCrawling (max {args.max_pages} pages)...")
    all_candidates = crawl_candidate_urls(blog_url, max_pages=args.max_pages)
    print(f"  found {len(all_candidates)} same-domain post-shaped links")

    conn = db_connect()
    conn.autocommit = False
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("BEGIN READ ONLY")
        known = existing_urls_for_domain(cur, blog_url)
        problems: list = []
        submitted_by = determine_default_submitted_by(cur, problems)
    finally:
        conn.rollback()

    unknown_candidates = [u for u in all_candidates if u not in known]
    check_budget = max(args.max_candidates * 4, args.max_candidates)
    candidates_to_check = unknown_candidates[:check_budget]
    not_checked_this_run = len(unknown_candidates) - len(candidates_to_check)
    print(f"  {len(all_candidates) - len(unknown_candidates)} already known to source_ingest_queue, skipped")
    print(f"  checking byline on {len(candidates_to_check)} of {len(unknown_candidates)} new candidate(s) this run, capped at {args.max_candidates} confirmed write(s)")
    if not_checked_this_run:
        print(f"  {not_checked_this_run} more new candidate(s) not checked this run -- available on a later run")

    confirmed: List[dict] = []
    quarantined: List[dict] = []
    for url in candidates_to_check:
        if len(confirmed) >= args.max_candidates:
            break
        result = check_candidate(url, declared_author)
        if result["outcome"] == "confirmed":
            confirmed.append(result)
            print(f"  [confirmed] {url}  (byline: {result['found_name']!r} via {result['signal_source']})")
        else:
            quarantined.append(result)
            found_note = f" (found: {result['found_name']!r})" if result.get("found_name") else ""
            print(f"  [{result['outcome']}] {url}{found_note}")

    report = {
        "site": site["name"],
        "attribute_to": declared_author,
        "blog_url": blog_url,
        "run_started_at": datetime.now(timezone.utc).isoformat(),
        "mode": "apply" if args.apply else "dry_run",
        "total_links_found": len(all_candidates),
        "already_known_skipped": len(all_candidates) - len(unknown_candidates),
        "new_candidates_not_checked_this_run": not_checked_this_run,
        "confirmed": confirmed,
        "quarantined": quarantined,
        "queued_rows": [],
    }

    if not args.apply:
        print(f"\nDRY RUN complete. {len(confirmed)} would be queued and processed, {len(quarantined)} would be quarantined (no write).")
        log_path = write_run_log(report)
        print(f"Log: {log_path}")
        conn.close()
        return 0

    if not confirmed:
        print("\nNothing confirmed -- nothing to write.")
        log_path = write_run_log(report)
        print(f"Log: {log_path}")
        conn.close()
        return 0
    if not submitted_by:
        print("\nRefusing to write: could not determine a single default submitted_by (see problems above).")
        conn.close()
        return 1

    print(f"\nAPPLY: inserting and processing {len(confirmed)} confirmed row(s)...")
    conn.autocommit = True
    write_cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    for item in confirmed:
        notes = (
            f"site_ingest_crawler run, {datetime.now(timezone.utc).date().isoformat()} -- "
            f"byline confirmed ({item['found_name']!r} via {item['signal_source']})."
        )
        row_id = insert_queue_row(
            write_cur,
            url=item["final_url"],
            attribute_to=declared_author,
            submitted_by=submitted_by,
            notes=notes,
        )
        print(f"  queued {row_id} -> {item['final_url']}")
        worker_result = run_worker_once(row_id)
        final_state = fetch_row_final_state(write_cur, row_id)
        print(f"    worker exit={worker_result['returncode']} -> status={final_state.get('status')} stage={final_state.get('stage')} flag_reason={final_state.get('flag_reason')}")
        report["queued_rows"].append(
            {
                "row_id": row_id,
                "url": item["final_url"],
                "worker_returncode": worker_result["returncode"],
                "worker_stderr_tail": worker_result["stderr"],
                "final_state": final_state,
            }
        )

    conn.close()
    log_path = write_run_log(report)
    print(f"\nLog: {log_path}")
    return 0


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--site", help="exact 'name' value in the Approved Sites tab; omit to run every row with approved=TRUE")
    parser.add_argument("--apply", action="store_true", help="required acknowledgement for the real database write")
    parser.add_argument("--max-candidates", type=int, default=1, help="cap on new documents PER SITE this run may write (default 1)")
    parser.add_argument("--max-pages", type=int, default=3, help="cap on index-page pagination depth, per site (default 3)")
    args = parser.parse_args(argv)

    if args.max_candidates < 1 or args.max_pages < 1:
        parser.error("--max-candidates and --max-pages must be at least 1")

    if args.site:
        sites = [load_approved_site(args.site)]
    else:
        sites = load_all_approved_sites()
        if not sites:
            print(f"No rows in '{APPROVED_TAB}' have approved=TRUE -- nothing to do.")
            return 0
        print(f"No --site given -- running every approved site ({len(sites)}): {', '.join(str(s['name']) for s in sites)}")

    exit_code = 0
    for site in sites:
        exit_code = max(exit_code, run_for_site(site, args))
    return exit_code


if __name__ == "__main__":
    raise SystemExit(main())
