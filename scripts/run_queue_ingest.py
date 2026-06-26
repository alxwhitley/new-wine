#!/usr/bin/env python3
"""
run_queue_ingest.py — Stage 5 Run 2: Queue-driven ingest orchestrator.

Walks every source detail tab in sources/youtube/ingest_queue.xlsx (skips Queue
and HowToRun). For each tab, ingests rows where ingest=TRUE AND status=triaged
via the existing ingest pipeline (youtube_ingest.ingest_sheet).

Gate: this script SPENDS MONEY (Whisper fallback) and WRITES TO THE DB.
Review source tabs and confirm ingest=TRUE rows before running.

done / done_prior rows are excluded by the existing ingest_sheet() logic.
Sentinel miss → status=needs_source, skip — never ingest unattributed content.

Per-video fault-tolerant: one bad video logs and continues, never kills the run.
Per-tab fault-tolerant: one broken tab logs and moves to the next.

Usage:
    python3 scripts/run_queue_ingest.py
    python3 scripts/run_queue_ingest.py --dry-run
    python3 scripts/run_queue_ingest.py --sheet "Sam Storms"   # single tab
"""

import sys
from pathlib import Path
from typing import Optional

import openpyxl
from dotenv import load_dotenv

ROOT       = Path(__file__).resolve().parent.parent
QUEUE_PATH = ROOT / "sources" / "youtube" / "ingest_queue.xlsx"

load_dotenv(ROOT / "backend" / "app" / ".env")

_SCRIPTS = Path(__file__).resolve().parent
if str(_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS))

from youtube_ingest import find_ytdlp, ingest_sheet  # noqa: E402

SKIP_TABS = {"Queue", "HowToRun"}


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    import argparse
    parser = argparse.ArgumentParser(
        description="Run 2 — Queue-driven ingest: ingest checked rows across all source tabs"
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="Resolve sources only — no transcript fetch, no DB writes")
    parser.add_argument("--sheet", metavar="NAME",
                        help="Process only this tab (default: all source tabs)")
    args = parser.parse_args()

    if not QUEUE_PATH.exists():
        print(f"ERROR: queue not found at {QUEUE_PATH}")
        sys.exit(1)

    ytdlp = find_ytdlp()
    if not ytdlp:
        print("ERROR: yt-dlp not found. Run: pip3 install yt-dlp")
        sys.exit(1)

    wb        = openpyxl.load_workbook(QUEUE_PATH)
    all_tabs  = wb.sheetnames

    if args.sheet:
        if args.sheet not in all_tabs:
            print(f"ERROR: sheet {args.sheet!r} not found. Available: {all_tabs}")
            sys.exit(1)
        tabs_to_run = [args.sheet]
    else:
        tabs_to_run = [t for t in all_tabs if t not in SKIP_TABS]

    mode = "(DRY-RUN) " if args.dry_run else ""
    print(f"\n── Queue Ingest {mode}────────────────────────────────────────────")
    print(f"  {len(tabs_to_run)} source tab(s) to process: {tabs_to_run}")

    grand = {"done": 0, "failed": 0, "needs_source": 0}
    tab_results = []   # [(tab_name, stats, error)]

    for tab_name in tabs_to_run:
        print(f"\n{'─' * 64}")
        print(f"  Tab: {tab_name!r}")

        try:
            # Reload workbook per tab so each tab sees fresh state after prior saves
            wb_tab = openpyxl.load_workbook(QUEUE_PATH)
            ws_tab = wb_tab[tab_name]

            stats = ingest_sheet(wb_tab, ws_tab, ytdlp, dry_run=args.dry_run)
            tab_results.append((tab_name, stats, None))

            for k in ("done", "failed", "needs_source"):
                grand[k] = grand.get(k, 0) + stats.get(k, 0)

        except Exception as exc:
            msg = str(exc)
            print(f"  ✗ Tab error: {msg}")
            tab_results.append((tab_name, {}, msg))

    # ── Final tally ───────────────────────────────────────────────────────────
    print(f"\n{'═' * 64}")
    print("QUEUE INGEST SUMMARY")
    print(f"{'═' * 64}")
    W = 30
    print(f"  {'TAB':<{W}} DONE   FAILED  NEEDS_SOURCE  ERROR")
    print(f"  {'─' * (W + 50)}")
    for tab_name, stats, err in tab_results:
        if err:
            print(f"  {tab_name:<{W}} —      —       —             {err[:40]}")
        else:
            print(
                f"  {tab_name:<{W}} "
                f"{stats.get('done', 0):<7}"
                f"{stats.get('failed', 0):<8}"
                f"{stats.get('needs_source', 0):<14}"
            )
    print(f"  {'─' * (W + 50)}")
    print(f"  {'TOTAL':<{W}} {grand.get('done', 0):<7}{grand.get('failed', 0):<8}{grand.get('needs_source', 0)}")
    print()


if __name__ == "__main__":
    main()
