#!/usr/bin/env python3
"""review_discovery_candidates.py -- minimal one-candidate-at-a-time local
review tool for the Discovery tab of docs/ingestion/master_ingestion_queue.xlsx.

Alex's explicit requirement (2026-08-25): find links to blogs/material,
review them, approve them for ingestion -- nothing else. So this page shows
exactly two things per candidate: their name, and a link to their site.
Two buttons, no forms:
  Yes -- approved for ingestion. Writes a new row to the Approved Sites tab
         automatically (name, attribute_to, blog_url, approved=TRUE,
         approved_at) -- nothing to type. If that name already has an
         Approved Sites row, no duplicate is written, but this Discovery
         row is still marked reviewed.
  No  -- passed on, permanently. Never shown again.
Both actions mark the Discovery row (verification_status, reviewed_at,
review_notes) and immediately show the next unreviewed candidate. There is
no session/queue state anywhere -- every page load re-reads the Discovery
tab fresh and shows the first row that still qualifies, so nothing is lost
if the tool is restarted mid-review and a Yes/No can never apply to a stale
row.

A candidate qualifies if: verification_status == "unverified", it is not
already_in_corpus, claimed_written_content_exists is not False,
auto_link_check is not "no_blog_detected" (see
check_discovery_blog_links.py -- a separate one-shot script that actually
fetches each candidate's link and checks for real post-shaped content
before you ever see it here; run it first so this tool only shows
candidates that already look like real blogs), and it has at least one
claimed URL to show (claimed_blog_or_articles_url, else claimed_main_url,
else the first of other_urls).

Local-only (127.0.0.1), single-user, no auth -- same trust posture as every
other script in scripts/. Reads and writes
docs/ingestion/master_ingestion_queue.xlsx directly via openpyxl; no
database involved, and the Queue tab / sync_master_ingestion_queue.py /
site_ingest_crawler.py are untouched by this tool.

Refuses to start, and refuses every write, while
docs/ingestion/~$master_ingestion_queue.xlsx (Excel's own lock file) exists
-- close the workbook in Excel first.

Run: python3.12 scripts/review_discovery_candidates.py [--port 8765]
Opens your browser automatically.

Python 3.12 (Invariant 1).
"""
from __future__ import annotations

import argparse
import html
import threading
import webbrowser
from datetime import date, datetime, timezone
from pathlib import Path
from typing import List, Optional, Tuple

import openpyxl
import uvicorn
from fastapi import FastAPI, Form
from fastapi.responses import HTMLResponse, RedirectResponse

ROOT = Path(__file__).resolve().parent.parent
SHEET_PATH = ROOT / "docs" / "ingestion" / "master_ingestion_queue.xlsx"
LOCK_PATH = SHEET_PATH.parent / f"~${SHEET_PATH.name}"
DISCOVERY_TAB = "Discovery"
APPROVED_TAB = "Approved Sites"

app = FastAPI()


# ---------------------------------------------------------------------------
# Pure read logic -- no network, no write, safe to unit-test directly.
# ---------------------------------------------------------------------------
def _candidate_link(row: dict) -> Optional[str]:
    for key in ("claimed_blog_or_articles_url", "claimed_main_url"):
        value = row.get(key)
        if value and str(value).strip():
            return str(value).strip()
    other = row.get("other_urls")
    if other:
        first = str(other).split(";")[0].strip()
        if first:
            return first
    return None


def load_discovery_rows() -> List[dict]:
    wb = openpyxl.load_workbook(SHEET_PATH, data_only=True)
    ws = wb[DISCOVERY_TAB]
    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header) if h}
    return [
        {h: row[i] for h, i in idx.items()}
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row[idx["name"]]
    ]


def build_queue(rows: List[dict]) -> List[Tuple[dict, str]]:
    """The rows still worth showing Alex, in sheet order, each paired with
    the one link to display. Excludes: already decided, already in corpus,
    confirmed to have no written content, or nothing to click at all."""
    queue = []
    for row in rows:
        if str(row.get("verification_status") or "").strip().lower() != "unverified":
            continue
        if row.get("already_in_corpus") is True:
            continue
        if row.get("claimed_written_content_exists") is False:
            continue
        if str(row.get("auto_link_check") or "").strip().lower() == "no_blog_detected":
            continue
        link = _candidate_link(row)
        if not link:
            continue
        queue.append((row, link))
    return queue


def next_candidate() -> Optional[Tuple[dict, str, int]]:
    """(row, link, remaining_count) for the first still-eligible Discovery
    row, or None if there's nothing left to review."""
    queue = build_queue(load_discovery_rows())
    if not queue:
        return None
    row, link = queue[0]
    return row, link, len(queue)


# ---------------------------------------------------------------------------
# Write logic.
# ---------------------------------------------------------------------------
def _header_index(ws) -> dict:
    return {cell.value: cell.column for cell in ws[1] if cell.value}


def _ensure_discovery_review_columns(ws) -> dict:
    """Adds reviewed_at / review_notes to the Discovery header if they
    aren't there yet -- invisible bookkeeping, never shown to Alex, kept
    separate from the automated-research notes columns. Idempotent."""
    idx = _header_index(ws)
    next_col = ws.max_column
    for col_name in ("reviewed_at", "review_notes"):
        if col_name not in idx:
            next_col += 1
            ws.cell(row=1, column=next_col, value=col_name)
            idx[col_name] = next_col
    return idx


def _find_row_by_name(ws, name_col: int, name: str) -> int:
    target = name.strip().lower()
    for r in range(2, ws.max_row + 1):
        value = ws.cell(row=r, column=name_col).value
        if value and str(value).strip().lower() == target:
            return r
    raise RuntimeError(f"'{name}' was not found -- nothing was written. Reload the page and try again.")


def _mark_discovery_reviewed(ws, idx: dict, name: str, *, status: str, note: str) -> None:
    r = _find_row_by_name(ws, idx["name"], name)
    ws.cell(row=r, column=idx["verification_status"], value=status)
    ws.cell(row=r, column=idx["reviewed_at"], value=datetime.now(timezone.utc).isoformat())
    ws.cell(row=r, column=idx["review_notes"], value=note)


def _approved_sites_has_name(ws, idx: dict, name: str) -> bool:
    target = name.strip().lower()
    for r in range(2, ws.max_row + 1):
        value = ws.cell(row=r, column=idx["name"]).value
        if value and str(value).strip().lower() == target:
            return True
    return False


def _append_approved_site(ws, idx: dict, *, name: str, link: str) -> None:
    r = ws.max_row + 1
    today = date.today().isoformat()
    values = {
        "approved": "TRUE",
        "name": name,
        "attribute_to": name,
        "blog_url": link,
        "approved_at": f"{today} -- review tool approval (Discovery: {name})",
    }
    for col_name, value in values.items():
        if col_name in idx:
            ws.cell(row=r, column=idx[col_name], value=value)


def approve_candidate(name: str, link: str) -> None:
    if LOCK_PATH.exists():
        raise RuntimeError(f"Excel has the workbook open ('{LOCK_PATH.name}' exists) -- close it, then reload this page.")
    wb = openpyxl.load_workbook(SHEET_PATH, data_only=False)
    discovery = wb[DISCOVERY_TAB]
    approved = wb[APPROVED_TAB]
    d_idx = _ensure_discovery_review_columns(discovery)
    a_idx = _header_index(approved)

    if not _approved_sites_has_name(approved, a_idx, name):
        _append_approved_site(approved, a_idx, name=name, link=link)

    _mark_discovery_reviewed(discovery, d_idx, name, status="verified", note="Approved -> Approved Sites")
    wb.save(SHEET_PATH)


def reject_candidate(name: str) -> None:
    if LOCK_PATH.exists():
        raise RuntimeError(f"Excel has the workbook open ('{LOCK_PATH.name}' exists) -- close it, then reload this page.")
    wb = openpyxl.load_workbook(SHEET_PATH, data_only=False)
    discovery = wb[DISCOVERY_TAB]
    d_idx = _ensure_discovery_review_columns(discovery)
    _mark_discovery_reviewed(discovery, d_idx, name, status="rejected", note="Passed on via review tool")
    wb.save(SHEET_PATH)


# ---------------------------------------------------------------------------
# HTML -- plain string templates, no templating engine dependency.
# ---------------------------------------------------------------------------
def _esc(value) -> str:
    return html.escape(str(value), quote=True)


_STYLE = """
<style>
  body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
         background: #f5f5f4; color: #1c1917; display: flex; min-height: 100vh;
         align-items: center; justify-content: center; margin: 0; }
  .wrap { text-align: center; max-width: 640px; padding: 2rem; }
  .count { color: #78716c; font-size: 0.95rem; margin-bottom: 1.5rem; }
  h1 { font-size: 2rem; margin: 0 0 1rem; word-break: break-word; }
  a.link { font-size: 1.1rem; word-break: break-all; }
  .actions { margin-top: 2.5rem; display: flex; gap: 1rem; justify-content: center; }
  button { font-size: 1.25rem; padding: 0.9rem 2.5rem; border-radius: 0.75rem;
           border: none; cursor: pointer; font-weight: 600; }
  .yes { background: #16a34a; color: white; }
  .yes:hover { background: #15803d; }
  .no { background: #e7e5e4; color: #1c1917; }
  .no:hover { background: #d6d3d1; }
</style>
"""


def _page(body: str) -> str:
    return f"<!doctype html><html><head><meta charset='utf-8'><title>Review candidates</title>{_STYLE}</head><body><div class='wrap'>{body}</div></body></html>"


def _render_candidate(row: dict, link: str, remaining: int) -> str:
    name = row["name"]
    return _page(
        f"""
        <p class="count">{remaining} left to review</p>
        <h1>{_esc(name)}</h1>
        <p><a class="link" href="{_esc(link)}" target="_blank" rel="noopener">{_esc(link)}</a></p>
        <div class="actions">
          <form method="post" action="/yes">
            <input type="hidden" name="name" value="{_esc(name)}">
            <input type="hidden" name="link" value="{_esc(link)}">
            <button class="yes" type="submit">Yes</button>
          </form>
          <form method="post" action="/no">
            <input type="hidden" name="name" value="{_esc(name)}">
            <button class="no" type="submit">No</button>
          </form>
        </div>
        """
    )


def _render_done() -> str:
    return _page("<h1>You're all caught up.</h1><p class=\"count\">Nothing left to review right now.</p>")


def _render_error(message: str) -> str:
    return _page(f"<h1>Couldn't save that</h1><p>{_esc(message)}</p><p><a href='/'>Back</a></p>")


# ---------------------------------------------------------------------------
# Routes.
# ---------------------------------------------------------------------------
@app.get("/", response_class=HTMLResponse)
def index() -> HTMLResponse:
    if LOCK_PATH.exists():
        return HTMLResponse(_render_error(f"Excel has the workbook open ('{LOCK_PATH.name}' exists) -- close it, then reload this page."), status_code=409)
    found = next_candidate()
    if found is None:
        return HTMLResponse(_render_done())
    row, link, remaining = found
    return HTMLResponse(_render_candidate(row, link, remaining))


@app.post("/yes")
def yes(name: str = Form(...), link: str = Form(...)):
    try:
        approve_candidate(name, link)
    except RuntimeError as exc:
        return HTMLResponse(_render_error(str(exc)), status_code=409)
    return RedirectResponse("/", status_code=303)


@app.post("/no")
def no(name: str = Form(...)):
    try:
        reject_candidate(name)
    except RuntimeError as exc:
        return HTMLResponse(_render_error(str(exc)), status_code=409)
    return RedirectResponse("/", status_code=303)


# ---------------------------------------------------------------------------
# Entrypoint.
# ---------------------------------------------------------------------------
def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--port", type=int, default=8765)
    args = parser.parse_args(argv)

    if LOCK_PATH.exists():
        print(f"'{LOCK_PATH.name}' exists -- the workbook looks open in Excel. Close it, then rerun this command.")
        return 1

    url = f"http://127.0.0.1:{args.port}/"
    threading.Timer(1.0, lambda: webbrowser.open(url)).start()
    print(f"Opening {url} ...")
    uvicorn.run(app, host="127.0.0.1", port=args.port, log_level="warning")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
