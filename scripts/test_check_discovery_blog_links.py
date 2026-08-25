#!/usr/bin/env python3
"""Regression tests for check_discovery_blog_links.py. classify() is tested
directly via its injectable fetch parameter (no network). main()'s
row-scanning/write/caching/limit/lock behavior is tested end to end against
a throwaway workbook with classify() itself monkeypatched at the module
level (still no network) -- never docs/ingestion/master_ingestion_queue.xlsx.

Run: python3.12 scripts/test_check_discovery_blog_links.py
"""
from __future__ import annotations

import shutil
import sys
import tempfile
from pathlib import Path
from types import SimpleNamespace

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))

import check_discovery_blog_links as checker
import review_discovery_candidates as review
from source_ingest_queue.fetcher import FetchRejected, FetchTransient

_checks = []
_failures = []


def check(label, condition):
    _checks.append(label)
    if not condition:
        _failures.append(label)
        print(f"FAIL: {label}")


# ---------------------------------------------------------------------------
# classify() -- pure apart from the injected fetch, no network.
# ---------------------------------------------------------------------------
_BLOG_HTML = b"""<html><body>
<main>
<a href="/first-post/">First post</a>
<a href="/second-post/">Second post</a>
</main>
</body></html>"""

_NO_POSTS_HTML = b"""<html><body>
<header><nav><a href="/about/">About</a><a href="/contact/">Contact</a></nav></header>
</body></html>"""


def _fetch_returning(html_bytes, final_url="https://example.com/blog"):
    def _fetch(url):
        return SimpleNamespace(content=html_bytes, final_url=final_url)
    return _fetch


def _fetch_raising(exc):
    def _fetch(url):
        raise exc
    return _fetch


status, detail = checker.classify("https://example.com/blog", fetch=_fetch_returning(_BLOG_HTML))
check("classify: real post links -> looks_like_blog", status == "looks_like_blog")
check("classify: detail mentions the count", "2" in detail)

status, detail = checker.classify("https://example.com/blog", fetch=_fetch_returning(_NO_POSTS_HTML))
check("classify: only nav links -> no_blog_detected", status == "no_blog_detected")

status, detail = checker.classify("https://example.com/blog", fetch=_fetch_raising(FetchRejected("unsafe_url", "blocked")))
check("classify: FetchRejected -> check_failed, not no_blog_detected", status == "check_failed")

status, detail = checker.classify("https://example.com/blog", fetch=_fetch_raising(FetchTransient("connect_failure", "down")))
check("classify: FetchTransient -> check_failed, not no_blog_detected", status == "check_failed")

# ---------------------------------------------------------------------------
# rows_needing_check() / _ensure_column() -- against a throwaway workbook.
# ---------------------------------------------------------------------------
_DISCOVERY_HEADER = ["verification_status", "already_in_corpus", "name", "claimed_main_url", "claimed_blog_or_articles_url", "other_urls", "claimed_written_content_exists"]


def _build_discovery_workbook(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = review.DISCOVERY_TAB
    ws.append(_DISCOVERY_HEADER)
    for row in rows:
        ws.append([row.get(h) for h in _DISCOVERY_HEADER])
    tmpdir = Path(tempfile.mkdtemp())
    path = tmpdir / "test_master_ingestion_queue.xlsx"
    wb.save(path)
    return path


_original_sheet_path = review.SHEET_PATH
_original_lock_path = review.LOCK_PATH
_wb_path = _build_discovery_workbook(
    [
        {"verification_status": "unverified", "already_in_corpus": False, "name": "Alpha", "claimed_main_url": "https://alpha.example.com/blog"},
        {"verification_status": "unverified", "already_in_corpus": False, "name": "Beta", "claimed_main_url": "https://beta.example.com/blog"},
        {"verification_status": "unverified", "already_in_corpus": False, "name": "No Link", "claimed_main_url": None},
        {"verification_status": "verified", "already_in_corpus": False, "name": "Already Decided", "claimed_main_url": "https://decided.example.com/blog"},
    ]
)

try:
    review.SHEET_PATH = _wb_path
    review.LOCK_PATH = _wb_path.parent / f"~${_wb_path.name}"

    wb = openpyxl.load_workbook(_wb_path)
    ws = wb[review.DISCOVERY_TAB]
    idx = review._header_index(ws)
    checker._ensure_column(ws, idx, "auto_link_check")
    checker._ensure_column(ws, idx, "auto_link_check_at")
    check("_ensure_column added auto_link_check", "auto_link_check" in idx)
    before_max_col = ws.max_column
    checker._ensure_column(ws, idx, "auto_link_check")
    check("_ensure_column is idempotent, no duplicate column", ws.max_column == before_max_col)

    pending = checker.rows_needing_check(ws, idx)
    pending_names = [name for _, name, _ in pending]
    check("rows_needing_check includes an unchecked row with a link", "Alpha" in pending_names)
    check("rows_needing_check includes a second unchecked row", "Beta" in pending_names)
    check("rows_needing_check excludes a row with no usable link", "No Link" not in pending_names)
    check("rows_needing_check excludes an already-decided row -- checking it would be wasted work", "Already Decided" not in pending_names)
    check("exactly 2 rows need checking", len(pending) == 2)
    wb.save(_wb_path)

    # --- main(): dry run writes nothing ---
    fake_map = {
        "https://alpha.example.com/blog": ("looks_like_blog", "2 post-shaped link(s) found"),
        "https://beta.example.com/blog": ("no_blog_detected", "fetched fine, no post-shaped links found"),
    }
    _original_classify = checker.classify
    checker.classify = lambda link, **kw: fake_map[link]
    try:
        rc = checker.main([])
        check("dry run exits 0", rc == 0)
        wb2 = openpyxl.load_workbook(_wb_path, data_only=True)
        ws2 = wb2[review.DISCOVERY_TAB]
        idx2 = review._header_index(ws2)
        alpha_val = next(ws2.cell(row=r, column=idx2["auto_link_check"]).value for r in range(2, ws2.max_row + 1) if ws2.cell(row=r, column=idx2["name"]).value == "Alpha")
        check("dry run does not write auto_link_check", alpha_val is None)

        # --- main(): --apply writes results ---
        rc = checker.main(["--apply"])
        check("--apply exits 0", rc == 0)
        wb3 = openpyxl.load_workbook(_wb_path, data_only=True)
        ws3 = wb3[review.DISCOVERY_TAB]
        idx3 = review._header_index(ws3)

        def _val(name, col):
            return next(ws3.cell(row=r, column=idx3[col]).value for r in range(2, ws3.max_row + 1) if ws3.cell(row=r, column=idx3["name"]).value == name)

        check("apply: Alpha labeled looks_like_blog", _val("Alpha", "auto_link_check") == "looks_like_blog")
        check("apply: Beta labeled no_blog_detected", _val("Beta", "auto_link_check") == "no_blog_detected")
        check("apply: already-decided row was never checked at all", _val("Already Decided", "auto_link_check") is None)
        check("apply: verification_status is never touched by this script", _val("Already Decided", "verification_status") == "verified")
        check("apply: auto_link_check_at stamped", bool(_val("Alpha", "auto_link_check_at")))

        # --- rerun: everything already checked, nothing left to do ---
        rc = checker.main([])
        check("second dry run reports nothing pending (caching works)", rc == 0)
    finally:
        checker.classify = _original_classify

    # --- --limit caps a fresh batch ---
    wb4 = openpyxl.load_workbook(_wb_path)
    ws4 = wb4[review.DISCOVERY_TAB]
    ws4.append(["unverified", False, "Gamma", "https://gamma.example.com/blog", None, None, None, None, None])
    ws4.append(["unverified", False, "Delta", "https://delta.example.com/blog", None, None, None, None, None])
    wb4.save(_wb_path)

    checker.classify = lambda link, **kw: ("looks_like_blog", "1 post-shaped link(s) found")
    try:
        rc = checker.main(["--apply", "--limit", "1"])
        check("--limit exits 0", rc == 0)
        wb5 = openpyxl.load_workbook(_wb_path, data_only=True)
        ws5 = wb5[review.DISCOVERY_TAB]
        idx5 = review._header_index(ws5)
        checked_count = sum(
            1 for r in range(2, ws5.max_row + 1)
            if ws5.cell(row=r, column=idx5["name"]).value in ("Gamma", "Delta")
            and ws5.cell(row=r, column=idx5["auto_link_check"]).value
        )
        check("--limit 1 checks exactly one of the two new candidates", checked_count == 1)
    finally:
        checker.classify = _original_classify

    # --- lock file refusal ---
    review.LOCK_PATH.write_text("")
    rc = checker.main(["--apply"])
    check("main refuses (nonzero exit) while the workbook is locked", rc != 0)
    review.LOCK_PATH.unlink()
finally:
    review.SHEET_PATH = _original_sheet_path
    review.LOCK_PATH = _original_lock_path
    shutil.rmtree(_wb_path.parent, ignore_errors=True)


print(f"\n{len(_checks) - len(_failures)}/{len(_checks)} checks passed")
if _failures:
    print("\nFAILED:")
    for f in _failures:
        print(f"  - {f}")
    raise SystemExit(1)
