#!/usr/bin/env python3
"""
youtube_ingest.py — Stage 3: ingest pass for the unified YouTube ingest pipeline.

Reads sources/youtube/ingest_queue.xlsx. For each row where ingest=TRUE AND
status=triaged:

  1. Resolve source via source_resolver (channel_name → alias lookup).
     Sentinel hit → status=needs_source, skip (no unattributed content).
  2. Fetch transcript: yt-dlp auto-captions in NATIVE json3 format, with a
     Whisper-medium fallback. Caption completeness is verified against the
     video's real duration — a truncated track falls back to Whisper rather
     than being stored short. No LLM rewriting happens anywhere on this path;
     stored text is the speaker's verbatim words (see try_auto_captions).
  3. Write .txt to sources/youtube/cleaned/ with correct metadata headers.
  4. Call ingest_file() directly (same path as all other documents), which
     internally handles, in order:
       - chunk → embed → document row → chunks table
       - propositions.process_document() — gate lives in propositions.py
         (fires for unlicensed/licensed sources; skips owned/public_domain).
         Not called directly in this file — inherited via ingest_file().
       - topic tagging (Groq)
  5. RETAIN the .txt by moving it to sources/youtube/ingested/; write
     status=done in sheet. A stored document must never exist only in
     Supabase. This step used to delete the file instead, which is why 301
     of 374 YouTube documents had no local copy as of 2026-08-30. Only a
     FAILED ingest deletes, so ingested/ means "this is in the corpus".

Idempotent: rows with status != "triaged" are skipped.
One bad video never kills the run — exceptions captured as status=failed.

Sheet: sources/youtube/ingest_queue.xlsx
Columns: url | video_title | channel_name | guess | ingest | status | resolved_source

Usage:
    python3 scripts/youtube_ingest.py              # process all ticked rows
    python3 scripts/youtube_ingest.py --limit 2    # demo: process first N rows only
    python3 scripts/youtube_ingest.py --dry-run    # resolve sources only, no writes
"""

import argparse
import json
import os
import re
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Optional, Tuple
from urllib.parse import urlparse, parse_qs

import openpyxl
from dotenv import load_dotenv

ROOT = Path(__file__).resolve().parent.parent
load_dotenv(ROOT / "backend" / "app" / ".env")

# scripts/ must be in sys.path before heavy imports so that ingest, propositions,
# source_resolver, taxonomy, bible_refs etc. all resolve correctly.
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

# Heavy imports — after load_dotenv and sys.path setup
from ingest import ingest_file, DB_PARAMS, supabase  # noqa: E402
from source_resolver import resolve_source_id, SENTINEL_SOURCE_ID, normalize_alias_key  # noqa: E402

QUEUE_PATH   = ROOT / "sources" / "youtube" / "ingest_queue.xlsx"
CLEANED_DIR  = ROOT / "sources" / "youtube" / "cleaned"
# Every stored document must also exist as a local file -- Supabase is not the
# only copy. A successful ingest MOVES its transcript here; it used to be
# unlinked outright, which left 301 of 374 YouTube documents with no local
# copy at all between 2026-06-03 and 2026-08-30.
#
# Files written from here are video-id prefixed ("{video_id}_{slug}.txt").
# NEVER recover that id by splitting the filename on "_" -- YouTube ids
# legitimately contain underscores ("Al_a7taOEo0", "Icli_wYAfTo"), and doing
# so silently truncates them and reports files as missing. Match on the
# "{video_id}_" prefix instead. Older files here predate this convention and
# carry a slug only, with no id at all.
INGESTED_DIR = ROOT / "sources" / "youtube" / "ingested"
COOKIES_PATH = ROOT / "scripts" / "youtube_cookies.txt"

COLUMNS = ["url", "video_title", "channel_name", "guess", "ingest", "status", "resolved_source"]
COL     = {name: i + 1 for i, name in enumerate(COLUMNS)}

# ── Transcript utilities (inlined from scrape_individual_videos.py) ───────────

# Auto-caption completeness guardrail. This is the check whose absence let
# ~38%-length transcripts reach production unnoticed (2026-08-29): a caption
# track that stops well before the end of the video is now rejected, and the
# video falls back to Whisper instead of being stored short and silent.
_MIN_CAPTION_COVERAGE = 0.85
_LARGE_GAP_WARN_S = 180

# Auto-captions mark non-speech as bracketed tokens -- [music], [applause],
# [cheering], [laughter], [snorts] all observed live. Brackets never carry
# legitimate spoken words in this format, so a length-bounded generic pattern
# is both safe and future-proof.
_NON_SPEECH_MARKER_RE = re.compile(r"\[[^\]]{1,30}\]")

AUDIO_EXTENSIONS = {".m4a", ".mp3", ".opus", ".ogg", ".webm", ".wav"}


def find_ytdlp():
    import shutil as _shutil
    candidates = [
        _shutil.which("yt-dlp"),
        os.path.expanduser("~/Library/Python/3.9/bin/yt-dlp"),
        os.path.expanduser("~/Library/Python/3.10/bin/yt-dlp"),
        os.path.expanduser("~/Library/Python/3.11/bin/yt-dlp"),
        os.path.expanduser("~/Library/Python/3.12/bin/yt-dlp"),
        os.path.expanduser("~/.local/bin/yt-dlp"),
    ]
    for c in candidates:
        if c and os.path.isfile(c):
            return c
    return None


def _ytdlp_base_args(ytdlp):
    args = [ytdlp, "-4", "--extractor-args", "youtube:player_client=android_vr,web_safari"]
    if COOKIES_PATH.exists():
        args += ["--cookies", str(COOKIES_PATH)]
    return args


def _parse_json3(path) -> Optional[Tuple[str, int, int]]:
    """Extract transcript text + timing coverage from a yt-dlp json3 caption file.

    Returns (text, last_caption_end_ms, largest_internal_gap_ms), or None if
    the file is unparseable or carries no text.

    Every `segs` entry is concatenated in document order, INCLUDING the
    `aAppend` events: those carry the "\\n" line separators, and dropping them
    glues words together across cue boundaries ("everybody.Well,").
    """
    try:
        data = json.loads(Path(path).read_text(encoding="utf-8", errors="replace"))
    except Exception:
        return None

    parts = []
    spans = []  # (start_ms, end_ms) for text-bearing events only
    for ev in data.get("events", []):
        piece = "".join(seg.get("utf8", "") for seg in (ev.get("segs") or []))
        parts.append(piece)
        if piece.strip():
            start = ev.get("tStartMs") or 0
            spans.append((start, start + (ev.get("dDurationMs") or 0)))

    if not spans:
        return None

    text = _NON_SPEECH_MARKER_RE.sub(" ", "".join(parts))
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return None

    spans.sort()
    last_end = max(end for _, end in spans)
    max_gap = max(
        (spans[i + 1][0] - spans[i][1] for i in range(len(spans) - 1)),
        default=0,
    )
    return text, last_end, max_gap


def try_auto_captions(ytdlp, url, tmp_dir):
    """Download YouTube auto-captions as native json3 and return the transcript.

    json3 is requested deliberately INSTEAD OF `--convert-subs srt`. The SRT
    conversion flattens YouTube's rolling-window cue format into literally
    triplicated text -- proven live 2026-08-29: a 9,327-word sermon came back
    as 27,783 words, and the Groq "cleaning" pass meant to undo that instead
    discarded ~62% of the real sermon. json3 carries each word exactly once,
    so there is no duplication to remove, no dedup regex, and no cleaning
    model anywhere on this path.

    Returns transcript text, or None -- which triggers the Whisper fallback --
    when captions are absent, too short, or stop well before the end of the
    video.
    """
    import os as _os
    out_template = _os.path.join(tmp_dir, "%(id)s.%(ext)s")
    cmd = _ytdlp_base_args(ytdlp) + [
        "--write-auto-sub", "--sub-lang", "en",
        "--skip-download", "--sub-format", "json3",
        # `--print` implies `--simulate` in yt-dlp, which silently suppresses
        # the subtitle write -- `--no-simulate` is required alongside it, or
        # captions are never saved and every video falls through to Whisper.
        "--print", "%(duration)s", "--no-simulate",
        "-o", out_template,
        url,
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)

    # Same invocation reports the real duration, so completeness costs no
    # extra network round-trip.
    duration_s = None
    for line in result.stdout.strip().splitlines():
        if line.strip().isdigit():
            duration_s = int(line.strip())
            break

    for f in _os.listdir(tmp_dir):
        if not f.endswith(".json3"):
            continue
        parsed = _parse_json3(_os.path.join(tmp_dir, f))
        if not parsed:
            return None
        text, last_end_ms, max_gap_ms = parsed
        if len(text.split()) <= 100:
            return None

        if duration_s:
            coverage = (last_end_ms / 1000.0) / duration_s
            if coverage < _MIN_CAPTION_COVERAGE:
                print("    captions stop at {:.0f}s of {:,.0f}s ({:.0%}) — "
                      "rejecting as truncated, will try Whisper".format(
                          last_end_ms / 1000.0, duration_s, coverage))
                return None
            print("    captions cover {:.0%} of {:,.0f}s".format(coverage, duration_s))
        else:
            print("    WARNING: could not read video duration — "
                  "caption completeness unverified")

        if max_gap_ms > _LARGE_GAP_WARN_S * 1000:
            print("    WARNING: largest caption gap is {:.0f}s".format(
                max_gap_ms / 1000.0))
        return text
    return None


def download_and_whisper(ytdlp, url, tmp_dir):
    """Download audio and transcribe with Whisper medium. Returns text or None."""
    import os as _os
    out_template = _os.path.join(tmp_dir, "%(id)s.%(ext)s")
    cmd = _ytdlp_base_args(ytdlp) + ["-x", "-o", out_template, url]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        print("     yt-dlp audio error: {}".format(result.stderr.strip()[:200]))
        return None
    audio_path = None
    for f in _os.listdir(tmp_dir):
        if _os.path.splitext(f)[1].lower() in AUDIO_EXTENSIONS:
            audio_path = _os.path.join(tmp_dir, f)
            break
    if not audio_path:
        return None
    import whisper
    print("     Loading Whisper model: medium")
    model = whisper.load_model("medium")
    print("     Transcribing...")
    result = model.transcribe(audio_path, fp16=False, language="en")
    return result["text"].strip()


# NOTE -- there is deliberately no clean_transcript()/CLEANING_PROMPT here any
# more (removed 2026-08-29, Alex's decision). Do not reintroduce an LLM pass
# over transcript text on this path. The only job that pass genuinely did was
# undoing SRT triplication, which json3 makes impossible to create. Measured
# on one real 9,327-word sermon, every currently-available Groq model rewrites
# rather than copies: gpt-oss-120b kept 38%, gpt-oss-20b 7.5%, and qwen3.6-27b
# truncated mid-transcript after burning its budget on hidden reasoning. Stored
# chunk text must remain the speaker's actual words -- quote verification does
# exact-substring matching against it (CLAUDE.md Settled decisions #16-19).


# ── Sheet helpers ─────────────────────────────────────────────────────────────

def gcell(ws, row_idx: int, col_name: str):
    return ws.cell(row=row_idx, column=COL[col_name]).value


def scell(ws, row_idx: int, col_name: str, value) -> None:
    ws.cell(row=row_idx, column=COL[col_name], value=value)


# ── Helpers ───────────────────────────────────────────────────────────────────

def extract_video_id(url: str) -> str:
    """Extract YouTube video ID from watch?v= or youtu.be/ URL."""
    parsed = urlparse(url.strip())
    qs = parse_qs(parsed.query)
    if "v" in qs:
        return qs["v"][0]
    if "youtu.be" in parsed.netloc:
        return parsed.path.lstrip("/").split("?")[0]
    return re.sub(r"[^\w]", "", url)[-16:]


def _slugify(title: str, max_len: int = 50) -> str:
    s = re.sub(r"[^\w\s]", "", title.lower())
    s = re.sub(r"\s+", "_", s.strip())
    return s[:max_len].rstrip("_")


_SPEAKER_WORD = r"(?:[A-Z][a-z]+|[A-Z]{2,4}(?![a-z]))"
_SPEAKER_NAME = rf"{_SPEAKER_WORD}(?:\s+{_SPEAKER_WORD})+"
_SPEAKER_MULTI = rf"{_SPEAKER_NAME}(?:\s*(?:&|\band\b)\s*{_SPEAKER_NAME})*"

# Trailing words that look name-shaped (Title-Case) but are video-title noise,
# not part of a speaker's name — e.g. "| Paul Kidd Sermon", "| ... | CLF Raleigh".
# Found live 2026-08-29 on the CLF Church playlist batch (2 of 50 titles).
_SPEAKER_TRAILING_JUNK = {
    "sermon", "service", "church", "message", "teaching", "session",
    "morning", "evening", "night", "raleigh",
}

# Known misspellings in source video titles, corrected on extraction rather
# than left to propagate into documents.author. Add an entry only when a
# typo is independently confirmed (e.g. the correct spelling appears
# elsewhere in the same batch) — never a guess.
_SPEAKER_NAME_FIXES = {
    "shabaka willliams": "Shabaka Williams",  # CLF Church playlist, 2026-08-29
}

# Resolution paths that PROVE the title-extracted string is a real, known
# person: both mean the speaker text itself matched a `source_aliases` row.
# 'channel_name'/'source_name' resolved from the CHANNEL instead, and 'MISS'
# resolved to nothing — in those cases the speaker string is unverified and
# must never become documents.author. See _verified_speaker().
_SPEAKER_VERIFIED_VIA = frozenset({"title_speaker", "author"})


def _verified_speaker(extracted_speaker: str, via: str) -> str:
    """The speaker to record as documents.author, or '' if unverified.

    Root cause this closes (audit 2026-08-31): `_extract_speaker()` matches any
    run of two-or-more Title-Case words after a '|' or '-', which YouTube titles
    produce constantly — "Do This Instead", "Your Porn Battle Plan". Five such
    fragments reached `documents.author` as CITABLE under the Vlad Savchuk
    source, entering the permitted-name set the answer writer may attribute
    claims to.

    The signal to prevent it was already being computed and then thrown away:
    when the extracted string fails alias lookup, resolution falls back to the
    channel name and `via` records that — yet the speaker was written as author
    regardless.

    Returning '' leaves documents.author NULL, and citation falls back to the
    source name. That is not a degradation: it is the behavior already proven
    correct on the 119 Savchuk documents whose titles yielded no speaker at all.
    """
    if not extracted_speaker:
        return ""
    return extracted_speaker if via in _SPEAKER_VERIFIED_VIA else ""


def _extract_speaker(title: str) -> str:
    """Best-effort speaker extraction from a video title.
    Handles patterns like 'by Sam Storms', '| Sam Storms', '- Michael Rowntree',
    all-caps initials ('| Bishop JB Masinde'), co-speakers ('| Paul Kidd &
    Alex Whitley' -> 'Paul Kidd, Alex Whitley'), and a trailing noise word
    ('| Paul Kidd Sermon' -> 'Paul Kidd').
    Returns '' if not found.
    """
    # "by Firstname Lastname[, & Firstname Lastname]" — most common in Convergence titles
    m = re.search(rf"\bby\s+({_SPEAKER_MULTI})", title)
    if not m:
        # "- Firstname Lastname" or "| Firstname Lastname" at end
        m = re.search(rf"[|\-—]\s*({_SPEAKER_MULTI})\s*(?:[|\-]|$)", title)
    if not m:
        return ""

    names = []
    for raw_name in re.split(r"\s*(?:&|\band\b)\s*", m.group(1)):
        words = raw_name.split()
        while len(words) > 1 and words[-1].strip(".,").lower() in _SPEAKER_TRAILING_JUNK:
            words.pop()
        if words:
            clean = " ".join(words)
            clean = _SPEAKER_NAME_FIXES.get(clean.lower(), clean)
            names.append(clean)
    return ", ".join(names)


def _write_transcript_file(
    path: Path,
    title: str,
    speaker: str,
    channel: str,
    url: str,
    cleaned: str,
) -> None:
    header = (
        "TITLE: {}\n"
        "SPEAKER: {}\n"
        "CHANNEL: {}\n"
        "SOURCE: {}\n"
        "URL: {}\n"
        "SOURCE_URL: {}\n"
        "PUBLISHED: NA\n"
        "DURATION_MIN: 0.0\n"
        "SOURCE_TYPE: sermon\n"
        "---\n\n"
    ).format(title, speaker, channel, channel, url, url)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(header + cleaned, encoding="utf-8")


_NA_VALUES = {"na", "none", "null", ""}


def _resolve_channel_name(ytdlp: str, url: str, channel_name: str) -> str:
    """If channel_name is missing or 'NA', fetch the real name from yt-dlp.
    Falls back to the original value if yt-dlp also returns nothing useful.
    """
    if channel_name.strip().lower() not in _NA_VALUES:
        return channel_name

    # yt-dlp single-video metadata fetch — no download
    cmd = [ytdlp, "-4", "--extractor-args", "youtube:player_client=android_vr,web_safari"]
    if COOKIES_PATH.exists():
        cmd += ["--cookies", str(COOKIES_PATH)]
    cmd += ["--flat-playlist", "--print", "%(channel)s", url]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
    except subprocess.TimeoutExpired:
        return channel_name

    for line in result.stdout.strip().splitlines():
        line = line.strip()
        if line and line.lower() not in _NA_VALUES:
            return line
    return channel_name


def _get_source_display_name(source_id: str) -> Optional[str]:
    try:
        result = (
            supabase.table("sources")
            .select("name")
            .eq("id", source_id)
            .limit(1)
            .execute()
        )
        if result.data:
            return result.data[0]["name"]
    except Exception:
        pass
    return None


# ── Speaker-first source resolution ──────────────────────────────────────────

def _resolve_speaker_source(speaker_name: str) -> Optional[str]:
    """Resolve source_id from a speaker name via source_aliases.alias_key.

    Normalizes via normalize_alias_key() from source_resolver — not re-implemented.
    Returns source_id on alias hit, None on miss.
    Prints the same ALIAS_MISS token the shared resolver uses, so `grep
    ALIAS_MISS` catches every miss in the pipeline (source_resolver.py's own
    misses, plus this preliminary speaker-lookup miss).
    Never raises.
    """
    if not speaker_name:
        return None
    key = normalize_alias_key(speaker_name)
    if not key:
        return None
    try:
        result = (
            supabase.table("source_aliases")
            .select("source_id")
            .eq("alias_key", key)
            .limit(1)
            .execute()
        )
        if result.data:
            return result.data[0]["source_id"]
    except Exception as exc:
        print("  WARNING: alias lookup for {!r} failed: {}".format(speaker_name, exc))
        return None
    # Same ALIAS_MISS token the shared resolver prints (source_resolver.py) so
    # `grep ALIAS_MISS` catches every miss in the pipeline, not just the final
    # fallback resolver call. This is a preliminary lookup, not necessarily a
    # final miss — ingest_video() tries channel_name next, then the full
    # resolver, before deciding the video needs_source.
    print("ALIAS_MISS  source_name=None  author={!r}  (speaker lookup, key={!r})".format(speaker_name, key))
    return None


# ── Per-video ingest ──────────────────────────────────────────────────────────

def ingest_video(
    ytdlp: str,
    url: str,
    video_title: str,
    channel_name: str,
    dry_run: bool = False,
) -> Tuple[str, str, str]:
    """
    Ingest one video end-to-end.

    Returns (status, resolved_source_display, log_reason).
    status is one of: "done" | "failed" | "needs_source" | "dry_run".
    Never raises — all exceptions captured and returned as ("failed", ..., reason).
    """

    # ── 1. Source resolution ──────────────────────────────────────────────────
    # channel_name may be "NA" when yt-dlp flat-playlist couldn't return it.
    # Fetch the real channel name from the video before trying alias lookup.
    if channel_name.strip().lower() in _NA_VALUES:
        print("  channel name is NA — fetching from yt-dlp...")
        channel_name = _resolve_channel_name(ytdlp, url, channel_name)
        print("    resolved channel: {!r}".format(channel_name))

    # Speaker-first resolution: extract name from title, try that alias first.
    # For whitelist-mode rows the triage stores the matched speaker name in
    # channel_name (v["wl_match"]), so the two lookups usually resolve the same
    # key.  The title-extracted path guards against a raw channel name (e.g.
    # "SermonIndex.net") ever slipping through from a future code path.
    extracted_speaker = _extract_speaker(video_title)
    source_id = _resolve_speaker_source(extracted_speaker) if extracted_speaker else None
    via = "title_speaker"
    norm_key = normalize_alias_key(extracted_speaker or "")
    if source_id is None:
        source_id = _resolve_speaker_source(channel_name)
        via = "channel_name"
        norm_key = normalize_alias_key(channel_name)
    if source_id is None:
        # Both lookups missed — fall back to the full resolver (tries channel_name
        # then author) so any existing aliases still resolve correctly.
        try:
            source_id, norm_key, via = resolve_source_id(
                supabase, channel_name, extracted_speaker or None
            )
        except Exception as exc:
            return "failed", "", "source resolution error: {}".format(exc)

    if source_id == SENTINEL_SOURCE_ID:
        return "needs_source", "", "no alias for speaker={!r} channel={!r}".format(
            extracted_speaker, channel_name
        )

    display_name = _get_source_display_name(source_id) or channel_name
    print("  source: {!r} → {!r} (via {})".format(norm_key, display_name, via))

    if dry_run:
        return "dry_run", display_name, "dry_run"

    # ── 2. Transcript fetch ───────────────────────────────────────────────────
    raw_text = None
    method = None

    with tempfile.TemporaryDirectory() as tmp_dir:
        print("  transcript: trying auto-captions...")
        raw_text = try_auto_captions(ytdlp, url, tmp_dir)
        if raw_text:
            method = "captions"
            print("    {:,} words from captions".format(len(raw_text.split())))

    if not raw_text:
        print("  transcript: no captions — falling back to Whisper medium...")
        with tempfile.TemporaryDirectory() as tmp_dir:
            try:
                raw_text = download_and_whisper(ytdlp, url, tmp_dir)
                if raw_text:
                    method = "whisper"
                    print("    {:,} words from Whisper".format(len(raw_text.split())))
            except Exception as exc:
                return "failed", display_name, "Whisper error: {}".format(exc)

    if not raw_text:
        return "failed", display_name, "no captions and Whisper returned nothing"

    # ── 3. Write temp .txt ────────────────────────────────────────────────────
    # No cleaning/rewriting step by design -- what was fetched is what gets
    # stored, verbatim. See the note where clean_transcript() used to live.
    cleaned = raw_text
    print("  storing {:,} words verbatim (method={})".format(
        len(cleaned.split()), method))
    video_id = extract_video_id(url)
    fname    = "{}_{}.txt".format(video_id, _slugify(video_title))
    tmp_path = CLEANED_DIR / fname
    # Reuse step 1's already-resolved extraction rather than recomputing it:
    # the recomputed value discarded `via`, which is exactly the evidence that
    # says whether this string is a real person or title noise.
    speaker  = _verified_speaker(extracted_speaker, via)
    if extracted_speaker and not speaker:
        print("  SPEAKER_UNVERIFIED  dropping author={!r} (resolved via={!r}); "
              "attribution falls back to the source name".format(
                  extracted_speaker, via))
    _write_transcript_file(tmp_path, video_title, speaker, channel_name, url, cleaned)
    print("  wrote: {}  (speaker={!r})".format(fname, speaker or "—"))

    # ── 4. Ingest through the full pipeline ───────────────────────────────────
    # Pass the gate-approved source_id through explicitly rather than letting
    # ingest_file re-resolve it from the SOURCE:/SPEAKER: headers just written
    # above. The two resolution orders differ (speaker-first here vs.
    # source_name-first in ingest_file), so a channel/speaker alias mismatch
    # could otherwise silently attribute the document differently than this
    # function's sentinel gate (step 1) already decided.
    ingest_status = None
    ingest_reason = None
    try:
        ingest_status, ingest_reason = ingest_file(
            tmp_path, is_copyrighted=True, skip_dedup=True, source_id_override=source_id
        )
    except Exception as exc:
        ingest_status = "error"
        ingest_reason = str(exc)

    # ── 5. Retain the local copy ──────────────────────────────────────────────
    # A stored document must never exist only in Supabase. On success the
    # transcript MOVES to ingested/ and stays there; only a failed ingest is
    # cleaned up, so ingested/ continues to mean "this is in the corpus".
    # Failing to retain is a real failure of the run, not a cosmetic one --
    # it is reported rather than swallowed, but it never invalidates a
    # database write that already succeeded.
    retain_note = ""
    if ingest_status in ("processed", "skipped"):
        try:
            INGESTED_DIR.mkdir(parents=True, exist_ok=True)
            final_path = INGESTED_DIR / fname
            tmp_path.replace(final_path)
            print("  retained: sources/youtube/ingested/{}".format(fname))
        except Exception as exc:
            retain_note = ", RETAIN FAILED: {}".format(exc)
            print("  !! RETAIN FAILED ({}) -- document is in the database but "
                  "has no local copy: {}".format(exc, fname))
    else:
        tmp_path.unlink(missing_ok=True)

    if ingest_status in ("processed", "skipped"):
        return "done", display_name, "method={}, ingest={}{}".format(
            method, ingest_status, retain_note)
    else:
        return "failed", display_name, "ingest_file: {}/{}".format(ingest_status, ingest_reason)


# ── Callable pipeline (used by CLI and queue orchestrator) ───────────────────

def ingest_sheet(
    wb: openpyxl.Workbook,
    ws,
    ytdlp: str,
    limit: Optional[int] = None,
    dry_run: bool = False,
) -> dict:
    """
    Ingest all ingest=TRUE AND status=triaged rows in ws.
    Called by CLI main() and by run_queue_ingest.py.
    Returns {done, failed, needs_source} counts.
    """
    candidates = []
    for r in range(2, ws.max_row + 1):
        ingest_val = str(gcell(ws, r, "ingest") or "").strip().upper()
        status_val = str(gcell(ws, r, "status") or "").strip().lower()
        url_val    = str(gcell(ws, r, "url") or "").strip()
        title_val  = str(gcell(ws, r, "video_title") or "").strip()
        if ingest_val == "TRUE" and status_val == "triaged" and url_val and title_val:
            candidates.append(r)

    if limit:
        candidates = candidates[:limit]

    mode = "(DRY-RUN) " if dry_run else ""
    print("\n── Stage 3: YouTube Ingest {}──────────────────────────────────".format(mode))
    print("  {} row(s) to process".format(len(candidates)))

    stats = {"done": 0, "failed": 0, "needs_source": 0, "blocked": 0}

    for row_idx in candidates:
        url          = str(gcell(ws, row_idx, "url")).strip()
        video_title  = str(gcell(ws, row_idx, "video_title")).strip()
        channel_name = str(gcell(ws, row_idx, "channel_name") or "").strip()

        print("\n{}".format("─" * 64))
        print("  {}".format(video_title[:72]))
        print("  {}".format(url))

        # Blocklist guard: skip URLs that were intentionally removed from the corpus.
        # removed_urls is written by DELETE /admin/document/{id} before the delete fires.
        try:
            blocked = supabase.table("removed_urls").select("url").eq("url", url).limit(1).execute()
            if blocked.data:
                print("  ✋  SKIPPED (BLOCKED) — this URL is in removed_urls: it was intentionally "
                      "removed from the corpus and must not be re-ingested.")
                print("  url: {}".format(url))
                if not dry_run:
                    scell(ws, row_idx, "status", "removed")
                    wb.save(QUEUE_PATH)
                stats["blocked"] += 1
                continue
        except Exception as bl_exc:
            print("  ⚠  blocklist check failed ({}); proceeding with ingest".format(bl_exc))

        try:
            final_status, display_name, log_reason = ingest_video(
                ytdlp, url, video_title, channel_name, dry_run=dry_run
            )
        except Exception as exc:
            final_status = "failed"
            display_name = ""
            log_reason   = "unhandled: {}".format(exc)

        if final_status == "dry_run":
            print("  [DRY-RUN] would ingest → source: {!r}".format(display_name))
            continue

        print("  → status={}  source={!r}  ({})".format(final_status, display_name, log_reason))
        stats[final_status] = stats.get(final_status, 0) + 1

        if not dry_run:
            if final_status == "done":
                scell(ws, row_idx, "status",          "done")
                scell(ws, row_idx, "resolved_source", display_name)
            elif final_status == "needs_source":
                scell(ws, row_idx, "status",          "needs_source")
                scell(ws, row_idx, "resolved_source", "⚠ {}".format(log_reason))
            else:
                scell(ws, row_idx, "status", "failed")
                if display_name:
                    scell(ws, row_idx, "resolved_source", display_name)
            wb.save(QUEUE_PATH)

    return stats


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="YouTube ingest — Stage 3 of the unified ingest pipeline"
    )
    parser.add_argument("--sheet",   metavar="NAME",
                        help="Tab name to operate on (required)")
    parser.add_argument("--limit",   metavar="N", type=int,
                        help="Process at most N rows (use 2-3 for demo)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Resolve sources only — no transcript fetch, no DB writes")
    args = parser.parse_args()

    ytdlp = find_ytdlp()
    if not ytdlp:
        print("ERROR: yt-dlp not found. Run: pip3 install yt-dlp")
        sys.exit(1)

    if not QUEUE_PATH.exists():
        print("ERROR: queue not found at {}".format(QUEUE_PATH))
        sys.exit(1)

    _wb_check = openpyxl.load_workbook(QUEUE_PATH, read_only=True)
    _available = _wb_check.sheetnames
    _wb_check.close()

    if not args.sheet:
        print("ERROR: --sheet is required. Available tabs: {}".format(_available))
        sys.exit(1)
    if args.sheet not in _available:
        print("ERROR: sheet {!r} not found. Available tabs: {}".format(args.sheet, _available))
        sys.exit(1)

    wb = openpyxl.load_workbook(QUEUE_PATH)
    ws = wb[args.sheet]

    # Collect eligible rows: ingest=TRUE AND status=triaged.
    # Terminal statuses (done, done_prior, failed, needs_source, expanded) all
    # have status != "triaged", so they can never satisfy the condition below.
    # done_prior rows additionally have ingest=FALSE — double-excluded.
    stats = ingest_sheet(wb, ws, ytdlp, limit=args.limit, dry_run=args.dry_run)

    print("\n── Results ──────────────────────────────────────────────────────")
    print("  done:         {}".format(stats.get("done", 0)))
    print("  failed:       {}".format(stats.get("failed", 0)))
    print("  needs_source: {}".format(stats.get("needs_source", 0)))


if __name__ == "__main__":
    main()
