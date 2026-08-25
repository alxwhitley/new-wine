#!/usr/bin/env python3
"""check_discovery_blog_links.py -- one-shot live check of Discovery-tab
candidates, labeling which ones actually look like a real blog/article
site versus which don't.

Why: Discovery's claimed_written_content_exists column is an automated
research pass's GUESS, never a real visit -- and in practice most claimed
candidates aren't real ongoing blogs (dead links, no article list,
PDF-only, a bio page with nothing to read, etc.). This script actually
fetches each unchecked candidate's link (the same SSRF-safe pinned fetch
source_ingest_queue.fetcher already uses) and runs the same post-link
detection site_ingest_crawler.py already uses on approved sites
(source_ingest_queue.link_discovery.discover_links) to see if the page
actually has real post-shaped links -- not just a claim.

Writes a new Discovery column, auto_link_check, one of:
  looks_like_blog  -- at least one real post-shaped link found.
  no_blog_detected -- fetched fine, found none.
  check_failed     -- couldn't fetch/check (network, blocked, etc.) --
                      NOT treated as "no blog". review_discovery_candidates.py
                      still shows these for manual review, since we
                      genuinely don't know either way.
Never touches verification_status -- that column stays Alex's own
decision, never an automated one. Only checks rows that don't already
have an auto_link_check value, so rerunning this later only costs
newly-added/unchecked rows.

review_discovery_candidates.py's queue skips auto_link_check ==
"no_blog_detected" -- so after this runs (with --apply), dead candidates
never surface in the interactive review tool.

Two modes, same convention as every other script that touches this
workbook:
  no flags   -- DRY RUN. Fetches and classifies every unchecked candidate,
                prints the plan, writes nothing.
  --apply    -- Recomputes the identical checks fresh, writes
                auto_link_check + auto_link_check_at to every checked row,
                saving every 10 candidates so an interrupted run keeps
                whatever it already found.

Refuses to run (or write) while
docs/ingestion/~$master_ingestion_queue.xlsx (Excel's own lock file)
exists.

Run: python3.12 scripts/check_discovery_blog_links.py [--apply] [--limit N]

Python 3.12 (Invariant 1).
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable, List, Optional, Tuple

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))

import review_discovery_candidates as review
from source_ingest_queue.fetcher import FetchRejected, FetchTransient, fetch_html
from source_ingest_queue.link_discovery import discover_links

MIN_POST_LINKS = 1
SAVE_EVERY = 10


def _ensure_column(ws, header_idx: dict, name: str) -> dict:
    if name not in header_idx:
        col = ws.max_column + 1
        ws.cell(row=1, column=col, value=name)
        header_idx[name] = col
    return header_idx


def classify(link: str, *, fetch: Callable = fetch_html) -> Tuple[str, str]:
    """(status, detail). status is looks_like_blog / no_blog_detected /
    check_failed. Pure apart from the injected fetch call -- tests pass a
    fake fetch to exercise this without the network."""
    try:
        fetched = fetch(link)
    except (FetchRejected, FetchTransient) as exc:
        return "check_failed", str(exc)
    result = discover_links(fetched.content, fetched.final_url)
    if len(result.post_urls) >= MIN_POST_LINKS:
        return "looks_like_blog", f"{len(result.post_urls)} post-shaped link(s) found"
    return "no_blog_detected", "fetched fine, no post-shaped links found"


def rows_needing_check(ws, header_idx: dict) -> List[Tuple[int, str, str]]:
    """(row_number, name, link) for every still-unverified Discovery row
    with a usable link and no auto_link_check value yet, in sheet order.
    Already-decided rows (Alex already said Yes/No) are skipped -- checking
    them would never change whether review_discovery_candidates.py shows
    them, so it's a pure waste of a network fetch."""
    pending = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=header_idx["name"]).value
        if not name:
            continue
        status = ws.cell(row=r, column=header_idx["verification_status"]).value
        if str(status or "").strip().lower() != "unverified":
            continue
        if ws.cell(row=r, column=header_idx["auto_link_check"]).value:
            continue
        row_dict = {h: ws.cell(row=r, column=c).value for h, c in header_idx.items()}
        link = review._candidate_link(row_dict)
        if not link:
            continue
        pending.append((r, str(name), link))
    return pending


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--apply", action="store_true", help="write results; omit for a dry run")
    parser.add_argument("--limit", type=int, default=None, help="cap how many new candidates to check this run")
    args = parser.parse_args(argv)

    if review.LOCK_PATH.exists():
        print(f"'{review.LOCK_PATH.name}' exists -- the workbook looks open in Excel. Close it, then rerun.")
        return 1

    wb = openpyxl.load_workbook(review.SHEET_PATH, data_only=False)
    ws = wb[review.DISCOVERY_TAB]
    idx = review._header_index(ws)
    _ensure_column(ws, idx, "auto_link_check")
    _ensure_column(ws, idx, "auto_link_check_at")

    to_check = rows_needing_check(ws, idx)
    if args.limit:
        to_check = to_check[: args.limit]

    print(f"{len(to_check)} candidate(s) need checking.")
    if not to_check:
        return 0

    counts = {"looks_like_blog": 0, "no_blog_detected": 0, "check_failed": 0}
    for i, (r, name, link) in enumerate(to_check, 1):
        status, detail = classify(link)
        counts[status] += 1
        print(f"[{i}/{len(to_check)}] {name} -> {status} ({detail})")
        if args.apply:
            ws.cell(row=r, column=idx["auto_link_check"], value=status)
            ws.cell(row=r, column=idx["auto_link_check_at"], value=datetime.now(timezone.utc).isoformat())
            if i % SAVE_EVERY == 0:
                wb.save(review.SHEET_PATH)
                print(f"  (progress saved: {i}/{len(to_check)})")

    if args.apply:
        wb.save(review.SHEET_PATH)
        print(
            f"\nSaved. {counts['looks_like_blog']} look like real blogs, "
            f"{counts['no_blog_detected']} auto-skipped (no blog structure found), "
            f"{counts['check_failed']} couldn't be checked (still shown for manual review)."
        )
    else:
        print(
            f"\nDRY RUN complete -- nothing written. Would label: "
            f"{counts['looks_like_blog']} looks_like_blog, "
            f"{counts['no_blog_detected']} no_blog_detected, "
            f"{counts['check_failed']} check_failed. Rerun with --apply to save."
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
