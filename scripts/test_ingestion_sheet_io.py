#!/usr/bin/env python3
"""Regression tests for ingestion_sheet_io.py -- proves the TSV encoding is
lossless, not just plausible. Two layers: targeted edge-case round trips
(embedded newline/tab/backslash/empty/None/bool/int), then a full round
trip of every real cell in the live docs/ingestion/master_ingestion_queue.xlsx
workbook (all four tabs) through write_tab()/read_tab(), compared cell by
cell against the original openpyxl-read values.

Run: python3.12 scripts/test_ingestion_sheet_io.py
"""
from __future__ import annotations

import shutil
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import ingestion_sheet_io as io

_checks = []
_failures = []


def check(label, condition):
    _checks.append(label)
    if not condition:
        _failures.append(label)
        print(f"FAIL: {label}")


# ---------------------------------------------------------------------------
# escape_cell / unescape_cell -- targeted edge cases
# ---------------------------------------------------------------------------
_EDGE_CASES = [
    None,
    "",
    "plain text",
    "text, with, commas",
    "line one\nline two",
    "line one\r\nline two (CRLF)",
    "a\ttab\tin\tthe\tmiddle",
    "trailing backslash\\",
    "already has \\n a literal backslash-n sequence",
    "mixed: back\\slash then\nreal newline then\ttab",
    True,
    False,
    0,
    42,
    -7,
]

for case in _EDGE_CASES:
    escaped = io.escape_cell(case)
    check(f"escape_cell({case!r}) contains no raw newline", "\n" not in escaped)
    check(f"escape_cell({case!r}) contains no raw tab", "\t" not in escaped)
    if isinstance(case, bool):
        expected = "TRUE" if case else "FALSE"
        check(f"escape_cell(bool {case!r}) -> {expected!r}", escaped == expected)
        check(f"round trip bool {case!r}", io.unescape_cell(escaped) == expected)
    elif case is None:
        check("escape_cell(None) -> ''", escaped == "")
        check("round trip None -> ''", io.unescape_cell(escaped) == "")
    else:
        expected_str = str(case)
        check(f"round trip {case!r}", io.unescape_cell(escaped) == expected_str)

# A value containing the literal two-character sequence backslash-n must
# round-trip distinctly from a value containing a real newline -- if the
# escaping were ambiguous these two would collide.
literal_backslash_n = "a\\nb"       # backslash, n, nothing else -- 4 chars: a \ n b
real_newline = "a\nb"               # a, real newline, b -- 3 chars
check(
    "literal backslash-n and real newline escape to different strings",
    io.escape_cell(literal_backslash_n) != io.escape_cell(real_newline),
)
check("literal backslash-n round trips exactly", io.unescape_cell(io.escape_cell(literal_backslash_n)) == literal_backslash_n)
check("real newline round trips exactly", io.unescape_cell(io.escape_cell(real_newline)) == real_newline)

# ---------------------------------------------------------------------------
# parse_bool_cell
# ---------------------------------------------------------------------------
check("parse_bool_cell(True) -> True (bool passthrough)", io.parse_bool_cell(True) is True)
check("parse_bool_cell(False) -> False (bool passthrough)", io.parse_bool_cell(False) is False)
check("parse_bool_cell(None) -> None", io.parse_bool_cell(None) is None)
check("parse_bool_cell('') -> None", io.parse_bool_cell("") is None)
check("parse_bool_cell('TRUE') -> True", io.parse_bool_cell("TRUE") is True)
check("parse_bool_cell('false') -> False (case-insensitive)", io.parse_bool_cell("false") is False)
check("parse_bool_cell('  TRUE  ') -> True (whitespace-tolerant)", io.parse_bool_cell("  TRUE  ") is True)
check("parse_bool_cell('garbage') -> None", io.parse_bool_cell("garbage") is None)

# ---------------------------------------------------------------------------
# write_tab / read_tab -- structural round trip against a temp file
# ---------------------------------------------------------------------------
_tmpdir = Path(tempfile.mkdtemp())
try:
    _path = _tmpdir / "roundtrip.tsv"
    _headers = ["name", "notes", "flag", "count", "blank_field"]
    _rows = [
        {"name": "Alpha", "notes": "simple", "flag": True, "count": 3, "blank_field": None},
        {"name": "Beta, With Comma", "notes": "line one\nline two\nline three", "flag": False, "count": 0, "blank_field": ""},
        {"name": "Gamma", "notes": "has a\\backslash and a\ttab", "flag": None, "count": -1, "blank_field": None},
    ]
    io.write_tab(_path, _headers, _rows)

    raw_lines = _path.read_text(encoding="utf-8").split("\n")
    check("write_tab produces one header line + N data lines + trailing newline", raw_lines[-1] == "" and len(raw_lines) - 1 == 1 + len(_rows))
    check("no raw tab characters leak into any single field beyond the delimiters", all(line.count("\t") == len(_headers) - 1 for line in raw_lines[:-1]))

    read_headers, read_rows = io.read_tab(_path)
    check("headers round trip exactly", read_headers == _headers)
    check("row count round trips exactly", len(read_rows) == len(_rows))

    for original, read_back in zip(_rows, read_rows):
        for h in _headers:
            orig_v = original.get(h)
            if isinstance(orig_v, bool):
                expected = "TRUE" if orig_v else "FALSE"
            elif orig_v is None:
                expected = ""
            else:
                expected = str(orig_v)
            check(f"{original['name']!r}.{h} round trips ({expected!r})", read_back.get(h) == expected)

    # A fully blank row (all cells "") must not silently vanish from a
    # normal write/read cycle if it's genuinely in the data -- but the
    # convention (matching the old openpyxl reader) is to skip fully-blank
    # lines. Confirm that's a deliberate, working behavior, not an accident.
    io.write_tab(_path, _headers, _rows + [{h: None for h in _headers}])
    _, read_rows_with_blank = io.read_tab(_path)
    check("a fully-blank trailing row is skipped, matching the old reader's behavior", len(read_rows_with_blank) == len(_rows))
finally:
    shutil.rmtree(_tmpdir, ignore_errors=True)

# ---------------------------------------------------------------------------
# Full round trip against the REAL live workbook -- every cell, all 4 tabs.
# This is the actual "prove it, don't assert it" evidence for the
# xlsx -> TSV conversion: every value openpyxl reads from the live file
# must come back identical (modulo the documented type normalization)
# after a write_tab()/read_tab() cycle.
# ---------------------------------------------------------------------------
_XLSX_PATH = io.ROOT / "docs" / "ingestion" / "master_ingestion_queue.xlsx"
if _XLSX_PATH.exists():
    import openpyxl

    wb = openpyxl.load_workbook(_XLSX_PATH, data_only=True)
    _tmpdir2 = Path(tempfile.mkdtemp())
    try:
        for tab_name in wb.sheetnames:
            ws = wb[tab_name]
            is_read_me = tab_name == "Read Me"
            if is_read_me:
                # Row 1 is title content, not a header; blank rows are
                # meaningful paragraph separators, not sheet noise.
                headers = ["text"]
                orig_rows = [{"text": row[0]} for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)]
            else:
                headers = [str(c.value) for c in ws[1]]
                orig_rows = []
                for raw in ws.iter_rows(min_row=2, values_only=True):
                    if all(v is None for v in raw):
                        continue
                    orig_rows.append(dict(zip(headers, raw)))

            path = _tmpdir2 / f"{tab_name.replace(' ', '_')}.tsv"
            io.write_tab(path, headers, orig_rows)
            _, read_back_rows = io.read_tab(path, skip_blank_rows=not is_read_me)

            check(f"[{tab_name}] row count survives round trip ({len(orig_rows)})", len(read_back_rows) == len(orig_rows))

            mismatches = []
            for i, (orig, back) in enumerate(zip(orig_rows, read_back_rows)):
                for h in headers:
                    ov = orig.get(h)
                    if isinstance(ov, bool):
                        expected = "TRUE" if ov else "FALSE"
                    elif ov is None:
                        expected = ""
                    else:
                        expected = str(ov)
                    if back.get(h) != expected:
                        mismatches.append((i, h, expected, back.get(h)))
            check(f"[{tab_name}] every real cell round trips byte-for-byte ({len(orig_rows)} rows x {len(headers)} cols)", not mismatches)
            if mismatches:
                for i, h, expected, actual in mismatches[:10]:
                    print(f"    row {i} field {h!r}: expected {expected!r}, got {actual!r}")
    finally:
        shutil.rmtree(_tmpdir2, ignore_errors=True)
else:
    print("NOTE: live xlsx not found -- skipped the full-workbook round-trip proof (edge-case + synthetic checks above still ran).")


print(f"\n{len(_checks) - len(_failures)}/{len(_checks)} checks passed")
if _failures:
    print("\nFAILED:")
    for f in _failures:
        print(f"  - {f}")
    raise SystemExit(1)
