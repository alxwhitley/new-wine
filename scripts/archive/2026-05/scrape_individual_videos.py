#!/usr/bin/env python3
"""
scrape_individual_videos.py — Ingest individual YouTube videos from an xlsx spreadsheet.

Pipeline per video:
  1. Try yt-dlp auto captions
  2. If no captions → Whisper medium transcription
  3. Clean via Groq Llama 3.3 70B
  4. Ingest via ingest.py (embed + Supabase)

Resume-safe: reads Status column from xlsx. Updates status after every step.
"""

import os
import re
import shutil
import subprocess
import sys
import tempfile
from datetime import datetime
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

ROOT = Path(__file__).resolve().parent.parent
load_dotenv(ROOT / "backend" / "app" / ".env")

XLSX_PATH   = ROOT / "sources" / "youtube" / "individual_videos.xlsx"
CLEANED_DIR = ROOT / "sources" / "youtube" / "cleaned"
COOKIES_PATH = ROOT / "scripts" / "youtube_cookies.txt"

GROQ_MODEL   = "llama-3.3-70b-versatile"
WHISPER_MODEL = "medium"
CHUNK_WORDS   = 6000

CLEANING_PROMPT = (
    "You are cleaning a YouTube sermon transcript for a theological research "
    "database. Remove ALL advertisement segments, sponsor reads, and promotional "
    "content. Remove non-speech markers like [music], [laughter], [applause], "
    "[clears throat], [cough]. Remove auto-caption filler artifacts like repeated "
    "phrases or mid-sentence restarts. Preserve ALL theological content verbatim. "
    "Return only the cleaned text with no commentary or preamble."
)

# ── Column indices (0-based) ─────────────────────────────────────────────────
COL_URL         = 1  # A
COL_TITLE       = 2  # B
COL_SPEAKER     = 3  # C
COL_CHANNEL     = 4  # D
COL_SOURCE_NAME = 5  # E
COL_STATUS      = 6  # F
COL_DATE        = 7  # G
COL_NOTES       = 8  # H


def slugify(title, max_len=60):
    s = re.sub(r"[^\w\s]", "", title.lower())
    s = re.sub(r"\s+", "_", s.strip())
    return s[:max_len].rstrip("_")


def find_ytdlp():
    candidates = [
        shutil.which("yt-dlp"),
        os.path.expanduser("~/Library/Python/3.9/bin/yt-dlp"),
        os.path.expanduser("~/Library/Python/3.10/bin/yt-dlp"),
        os.path.expanduser("~/Library/Python/3.11/bin/yt-dlp"),
        os.path.expanduser("~/Library/Python/3.12/bin/yt-dlp"),
        os.path.expanduser("~/.local/bin/yt-dlp"),
    ]
    for c in candidates:
        if c and os.path.isfile(c):
            return c
    return None


def _ytdlp_base_args(ytdlp):
    args = [ytdlp, "-4", "--extractor-args", "youtube:player_client=android_vr,web_safari"]
    if COOKIES_PATH.exists():
        args += ["--cookies", str(COOKIES_PATH)]
    return args


def try_auto_captions(ytdlp, url, tmp_dir):
    """Try to download auto-generated English captions via yt-dlp. Returns text or None."""
    out_template = os.path.join(tmp_dir, "%(id)s.%(ext)s")
    cmd = _ytdlp_base_args(ytdlp) + [
        "--write-auto-sub", "--sub-lang", "en",
        "--skip-download", "--convert-subs", "srt",
        "-o", out_template,
        url,
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)

    # Find the .srt file
    for f in os.listdir(tmp_dir):
        if f.endswith(".srt"):
            srt_path = os.path.join(tmp_dir, f)
            raw = Path(srt_path).read_text(encoding="utf-8", errors="replace")
            # Strip SRT formatting: timestamps, sequence numbers, blank lines
            lines = []
            for line in raw.splitlines():
                line = line.strip()
                if not line:
                    continue
                if re.match(r"^\d+$", line):
                    continue
                if re.match(r"\d{2}:\d{2}:\d{2}", line):
                    continue
                # Remove HTML tags from auto captions
                line = re.sub(r"<[^>]+>", "", line)
                if line:
                    lines.append(line)
            text = " ".join(lines)
            # Deduplicate repeated phrases (common in auto captions)
            text = re.sub(r"\b(\w+(?:\s+\w+){0,3})\s+\1\b", r"\1", text)
            if len(text.split()) > 100:
                return text
    return None


AUDIO_EXTENSIONS = {".m4a", ".mp3", ".opus", ".ogg", ".webm", ".wav"}


def download_and_whisper(ytdlp, url, tmp_dir):
    """Download audio and transcribe with Whisper medium. Returns text or None."""
    out_template = os.path.join(tmp_dir, "%(id)s.%(ext)s")
    cmd = _ytdlp_base_args(ytdlp) + ["-x", "-o", out_template, url]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        print("     yt-dlp audio error: {}".format(result.stderr.strip()[:200]))
        return None

    audio_path = None
    for f in os.listdir(tmp_dir):
        if os.path.splitext(f)[1].lower() in AUDIO_EXTENSIONS:
            audio_path = os.path.join(tmp_dir, f)
            break
    if not audio_path:
        return None

    import whisper
    print("     Loading Whisper model: {}".format(WHISPER_MODEL))
    model = whisper.load_model(WHISPER_MODEL)
    print("     Transcribing...")
    result = model.transcribe(audio_path, fp16=False, language="en")
    return result["text"].strip()


def clean_transcript(raw):
    """Clean transcript via Groq, chunking if needed."""
    from groq import Groq
    client = Groq(api_key=os.environ["GROQ_API_KEY"])
    words = raw.split()

    if len(words) <= CHUNK_WORDS:
        chunks = [raw]
    else:
        chunks = [
            " ".join(words[i:i + CHUNK_WORDS])
            for i in range(0, len(words), CHUNK_WORDS)
        ]

    print("     Cleaning via Groq ({} chunk(s), {:,} words)...".format(len(chunks), len(words)))
    cleaned_parts = []
    for i, chunk in enumerate(chunks, 1):
        if len(chunks) > 1:
            print("       Chunk {}/{}...".format(i, len(chunks)))
        response = client.chat.completions.create(
            model=GROQ_MODEL,
            max_tokens=8192,
            messages=[
                {"role": "system", "content": CLEANING_PROMPT},
                {"role": "user", "content": chunk},
            ],
        )
        cleaned_parts.append((response.choices[0].message.content or "").strip())

    return "\n\n".join(cleaned_parts)


def write_transcript(out_path, title, speaker, channel, source_name, url, cleaned):
    header = (
        "TITLE: {}\n"
        "SPEAKER: {}\n"
        "CHANNEL: {}\n"
        "SOURCE: {}\n"
        "URL: {}\n"
        "SOURCE_URL: {}\n"
        "PUBLISHED: NA\n"
        "DURATION_MIN: 0.0\n"
        "SOURCE_TYPE: sermon\n"
        "---\n\n"
    ).format(title, speaker, channel, source_name, url, url)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(header + cleaned, encoding="utf-8")


def update_xlsx(wb, ws, row_idx, status, notes=""):
    ws.cell(row=row_idx, column=COL_STATUS, value=status)
    ws.cell(row=row_idx, column=COL_DATE, value=datetime.now().strftime("%Y-%m-%d"))
    if notes:
        ws.cell(row=row_idx, column=COL_NOTES, value=notes)
    wb.save(XLSX_PATH)


def main():
    if not os.environ.get("GROQ_API_KEY"):
        print("ERROR: GROQ_API_KEY not set in backend/app/.env")
        sys.exit(1)

    ytdlp = find_ytdlp()
    if not ytdlp:
        print("ERROR: yt-dlp not found. Run: pip3 install yt-dlp")
        sys.exit(1)

    if not XLSX_PATH.exists():
        print("ERROR: Spreadsheet not found at {}".format(XLSX_PATH))
        sys.exit(1)

    CLEANED_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Videos"]

    print("yt-dlp:   {}".format(ytdlp))
    print("Cookies:  {}{}".format(COOKIES_PATH, " OK" if COOKIES_PATH.exists() else " NOT FOUND"))
    print("Output:   {}".format(CLEANED_DIR))
    print("Tracker:  {}\n".format(XLSX_PATH))

    pending = []
    for row_idx in range(2, ws.max_row + 1):
        status = str(ws.cell(row=row_idx, column=COL_STATUS).value or "").strip()
        if status == "Pending":
            pending.append(row_idx)

    if not pending:
        print("No pending videos found in spreadsheet.")
        return

    print("{} pending video(s)\n".format(len(pending)))

    stats = {"success": 0, "failed": 0}

    for row_idx in pending:
        url         = str(ws.cell(row=row_idx, column=COL_URL).value or "").strip()
        title       = str(ws.cell(row=row_idx, column=COL_TITLE).value or "").strip()
        speaker     = str(ws.cell(row=row_idx, column=COL_SPEAKER).value or "").strip()
        channel     = str(ws.cell(row=row_idx, column=COL_CHANNEL).value or "").strip()
        source_name = str(ws.cell(row=row_idx, column=COL_SOURCE_NAME).value or "Individual Videos").strip()

        if not url or not title:
            print("Row {}: missing URL or title — skipping".format(row_idx))
            update_xlsx(wb, ws, row_idx, "Failed", "Missing URL or title")
            stats["failed"] += 1
            continue

        print("{}".format("=" * 64))
        print("  {} — {}".format(speaker, title))
        print("  {}".format(url))
        print("{}".format("=" * 64))

        # Step 1: Try auto captions
        raw_text = None
        method = None
        with tempfile.TemporaryDirectory() as tmp_dir:
            print("  Step 1: Trying auto captions...")
            raw_text = try_auto_captions(ytdlp, url, tmp_dir)
            if raw_text:
                method = "captions"
                print("     Got {:,} words from auto captions".format(len(raw_text.split())))

        # Step 2: Whisper fallback
        if not raw_text:
            print("  Step 2: No captions — falling back to Whisper...")
            update_xlsx(wb, ws, row_idx, "Whisper", "No captions available")
            with tempfile.TemporaryDirectory() as tmp_dir:
                try:
                    raw_text = download_and_whisper(ytdlp, url, tmp_dir)
                    if raw_text:
                        method = "whisper"
                        print("     Got {:,} words from Whisper".format(len(raw_text.split())))
                except Exception as e:
                    print("     Whisper failed: {}".format(e))

        if not raw_text:
            print("  FAILED: Could not get transcript")
            update_xlsx(wb, ws, row_idx, "Failed", "No captions and Whisper failed")
            stats["failed"] += 1
            continue

        # Step 3: Clean via Groq
        print("  Step 3: Cleaning transcript...")
        update_xlsx(wb, ws, row_idx, "Cleaning")
        try:
            cleaned = clean_transcript(raw_text)
            print("     {:,} words after cleaning".format(len(cleaned.split())))
        except Exception as e:
            print("     Groq cleaning failed: {}".format(e))
            update_xlsx(wb, ws, row_idx, "Failed", "Groq cleaning failed: {}".format(str(e)[:100]))
            stats["failed"] += 1
            continue

        # Step 4: Write to cleaned/ for ingest.py
        fname = slugify(title) + ".txt"
        out_path = CLEANED_DIR / fname
        print("  Step 4: Writing {}".format(out_path.name))
        write_transcript(out_path, title, speaker, channel, source_name, url, cleaned)

        update_xlsx(wb, ws, row_idx, "Scraped", "Method: {} | {:,} words".format(method, len(cleaned.split())))
        stats["success"] += 1
        print("  Done ({})".format(method))

    # Run ingest.py to embed and push to Supabase
    if stats["success"] > 0:
        print("\n{}".format("=" * 64))
        print("Running ingest.py to embed and push to Supabase...")
        print("{}".format("=" * 64))
        result = subprocess.run(
            [sys.executable, str(ROOT / "scripts" / "ingest.py")],
            cwd=str(ROOT),
        )
        if result.returncode == 0:
            # Mark all Scraped rows as Ingested
            wb = openpyxl.load_workbook(XLSX_PATH)
            ws = wb["Videos"]
            for row_idx in range(2, ws.max_row + 1):
                status = str(ws.cell(row=row_idx, column=COL_STATUS).value or "").strip()
                if status == "Scraped":
                    ws.cell(row=row_idx, column=COL_STATUS, value="Ingested")
                    ws.cell(row=row_idx, column=COL_DATE, value=datetime.now().strftime("%Y-%m-%d"))
            wb.save(XLSX_PATH)
            print("Marked {} video(s) as Ingested".format(stats["success"]))
        else:
            print("WARNING: ingest.py exited with code {}".format(result.returncode))

    print("\n{}".format("=" * 64))
    print("Done. {} succeeded, {} failed.".format(stats["success"], stats["failed"]))


if __name__ == "__main__":
    main()
