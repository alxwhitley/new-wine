#!/usr/bin/env python3
"""Restructure docs/ingestion/master_ingestion_queue.xlsx from one flat
"Master Queue" tab into two tabs: Discovery (raw, unvetted candidates) and
Queue (vetted, ready to actually push to the database).

Alex's decision, 2026-08-19: two-tab flow, Discovery -> (manual promotion
by Alex) -> Queue. Discovery rows must never reach the database. The sync
script is updated separately to read only the Queue tab.

Reads the CURRENT workbook's 12 existing rows and re-splits them by their
original stage: the 6 research_candidate rows (the old hardcoded frontend
cards) go to Discovery under the new richer schema; the 6
already_queued/done rows (from the live database) go to Queue under a
trimmed version of the old schema. No row is dropped; nothing is merged.

Repo-file-only. No database access, no network access.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent.parent.parent
SHEET_PATH = ROOT / "docs" / "ingestion" / "master_ingestion_queue.xlsx"
IMPORT_DATE = "2026-08-19"

OLD_TAB = "Master Queue"

# ── Queue tab schema ─────────────────────────────────────────────────────
# Trimmed from the old flat schema: description/reference_urls/source_card_id
# dropped (those were discovery-only fields; no row here ever used them --
# verified before writing). promoted_from_discovery added, for traceability
# when Alex manually promotes a Discovery row here.
QUEUE_COLUMNS = [
    "stage",                  # ready_to_queue | already_queued | done -- research_candidate
                               # is no longer a valid value here; those rows live on Discovery.
    "name",
    "url",
    "source_format",
    "source_scope",
    "attribute_to",
    "attribution_mode",
    "on_unknown_author",
    "retain_original_text",
    "notes",
    "flag_reason",
    "cleared_to_run",
    "db_status",
    "db_execution_stage",
    "attempts",
    "max_attempts",
    "worker_id",
    "lease_expires_at",
    "run_after",
    "final_url",
    "content_sha256",
    "fetched_bytes",
    "attempted_documents",
    "stored_documents",
    "skipped_documents",
    "errored_documents",
    "result_document_id",
    "submitted_by",
    "source_db_id",
    "promoted_from_discovery",  # NEW -- free text; note which Discovery row this came from.
    "origin",
    "created_at",
    "updated_at",
]

# ── Discovery tab schema ─────────────────────────────────────────────────
DISCOVERY_COLUMNS = [
    "verification_status",              # unverified | in_progress | verified | rejected
    "already_in_corpus",                # TRUE/FALSE -- NEW, per Task 3's cross-check requirement
    "name",
    "organization",
    "location",
    "category",                         # practitioner_teacher | academic_scholar | historical_primary_source_archive
    "living_or_deceased",               # living | deceased | historical
    "main_url",
    "blog_or_articles_url",
    "archive_url",
    "other_urls",
    "claimed_written_content_exists",   # TRUE | FALSE | unknown -- name itself carries the caveat
    "claimed_licensing_status",
    "claimed_platform_size",
    "discovery_paths",                  # "round: seed/source" pairs, semicolon-joined; supports multiple
    "corpus_match_notes",               # NEW -- which existing corpus entry this seems to match, and how
    "claims_source",                    # NEW -- where the claimed_* fields' claims actually came from
    "notes",
    "date_added",                       # NEW -- bookkeeping
]

READ_ME_LINES = [
    ("Master Ingestion Queue -- Read Me", True),
    ("", False),
    ("This workbook is the single master source of truth for ingestion", False),
    ("candidates (Alex's decision, 2026-08-19). The admin panel's ingest", False),
    ("queue table is a read-only mirror, kept in sync by a sync script that", False),
    ("only ever reads the QUEUE tab. Edit here, not there.", False),
    ("", False),
    ("On any disagreement between the QUEUE tab and the database, the sheet", False),
    ("wins and overwrites, once the sync script is run with real writes on.", False),
    ("", False),
    ("The existing admin submission form is untouched and still works; rows", False),
    ("submitted through it live only in the database until they're added to", False),
    ("the Queue tab too.", False),
    ("", False),
    ("TWO-TAB FLOW (added 2026-08-19):", True),
    ("  DISCOVERY  -- raw, unvetted candidates. Anyone/anything that's come", False),
    ("               up in research but that Alex has not personally looked", False),
    ("               at yet. Never reaches the database, ever, under any", False),
    ("               circumstances -- the sync script physically cannot see", False),
    ("               this tab.", False),
    ("  QUEUE      -- vetted candidates, on their way into or already in the", False),
    ("               live database queue. Alex manually promotes a row from", False),
    ("               Discovery to Queue after investigating it himself --", False),
    ("               there is no automated promotion.", False),
    ("", False),
    ("QUEUE tab -- STAGE column (column A):", True),
    ("  ready_to_queue  -- fully specified (one URL, format, scope,", False),
    ("                     attribution decided) but not yet in the database", False),
    ("                     queue.", False),
    ("  already_queued  -- already exists as a row in the database queue and", False),
    ("                     is not yet marked done there.", False),
    ("  done            -- already exists in the database queue with a", False),
    ("                     'done' status.", False),
    ("  (research_candidate is not a valid value here -- those rows belong", False),
    ("   on the Discovery tab instead.)", False),
    ("", False),
    ("Blank vs FALSE (QUEUE tab): for the TRUE/FALSE columns, a blank cell", True),
    ("means 'not yet decided' -- it is NOT the same as FALSE.", False),
    ("", False),
    ("db_status / db_execution_stage vs this sheet's own stage column: these", True),
    ("two are copied verbatim from the database row's own internal", False),
    ("bookkeeping (its execution status and the runner's own internal", False),
    ("stage) for already_queued and done rows -- a different thing from this", False),
    ("sheet's own STAGE column, which is our triage/workflow category.", False),
    ("", False),
    ("DISCOVERY tab -- how to read it:", True),
    ("  verification_status defaults to 'unverified' for every row until", False),
    ("  Alex has personally looked at the candidate.", False),
    ("", False),
    ("  Every column named claimed_* was filled in by an automated research", False),
    ("  pass inferring from search results -- NOT by visiting the site.", False),
    ("  Treat these as asserted, not established, until verification_status", False),
    ("  says otherwise. The claims_source column records which research pass", False),
    ("  produced the claim.", False),
    ("", False),
    ("  already_in_corpus is set from a direct, read-only check against the", False),
    ("  live database's source list at the time a row was added -- it can go", False),
    ("  stale if the corpus changes later. corpus_match_notes records what it", False),
    ("  matched against.", False),
    ("", False),
    ("  discovery_paths records every research round and seed/source that", False),
    ("  surfaced this candidate, semicolon-separated -- several rounds often", False),
    ("  turn up the same name independently; this is one row per person or", False),
    ("  entity, not one row per mention.", False),
    ("", False),
    ("source_db_id / promoted_from_discovery: source_db_id (QUEUE tab) records", True),
    ("exactly which database queue row a Queue-tab row corresponds to, once", False),
    ("one exists -- the sync script's match key. promoted_from_discovery is a", False),
    ("free-text note for when Alex manually promotes a Discovery candidate --", False),
    ("write the candidate's name there so the two tabs stay traceable to each", False),
    ("other by hand.", False),
    ("", False),
    ("Layout: one flat table per tab, not the YouTube/magazine trackers'", True),
    ("one-tab-per-source pattern -- see git history for the original reasoning", False),
    ("(every row here is already one candidate; there's no natural per-source", False),
    ("grouping to split on).", False),
    ("", False),
    (f"Restructured into two tabs: {IMPORT_DATE}. Original single-tab build:", False),
    ("2026-08-19 (see git history). All 12 rows from the original tab were", False),
    ("carried over intact -- 6 to Queue, 6 to Discovery.", False),
]


def style_header(ws, columns):
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    for col_idx, name in enumerate(columns, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True)
        c.fill = header_fill
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def size_columns(ws, columns, wide: set, narrow: set):
    for col_idx, name in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        if name in wide:
            ws.column_dimensions[letter].width = 48
        elif name in narrow:
            ws.column_dimensions[letter].width = 20
        else:
            ws.column_dimensions[letter].width = 16


def add_dropdown(ws, columns, col_name: str, options: list[str]):
    col_idx = columns.index(col_name) + 1
    letter = get_column_letter(col_idx)
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(options) + '"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.add(f"{letter}2:{letter}2000")


def main() -> int:
    old_wb = openpyxl.load_workbook(SHEET_PATH, data_only=True)
    old_ws = old_wb[OLD_TAB]
    old_headers = [c.value for c in old_ws[1]]
    old_rows = []
    for raw in old_ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in raw):
            continue
        old_rows.append(dict(zip(old_headers, raw)))

    print(f"Read {len(old_rows)} existing rows from {OLD_TAB!r}.")
    assert len(old_rows) == 12, f"expected 12 existing rows, found {len(old_rows)}"

    discovery_source_rows = [r for r in old_rows if r["stage"] == "research_candidate"]
    queue_source_rows = [r for r in old_rows if r["stage"] != "research_candidate"]
    print(f"  -> {len(discovery_source_rows)} research_candidate rows go to Discovery")
    print(f"  -> {len(queue_source_rows)} already_queued/done rows go to Queue")
    assert len(discovery_source_rows) == 6
    assert len(queue_source_rows) == 6

    # ── Map old rows onto the new Queue schema ─────────────────────────────
    queue_rows = []
    for r in queue_source_rows:
        new_row = {col: r.get(col) for col in QUEUE_COLUMNS if col != "promoted_from_discovery"}
        new_row["promoted_from_discovery"] = None
        # Confirm no data silently lost: the three dropped columns must be
        # empty on every row that's leaving this schema behind.
        for dropped in ("description", "reference_urls", "source_card_id"):
            assert not r.get(dropped), f"queue row {r.get('name')!r} had non-empty {dropped!r} -- would lose data"
        queue_rows.append(new_row)

    # ── Map old research-candidate rows onto the new Discovery schema ──────
    discovery_rows = []
    for r in discovery_source_rows:
        new_row = {col: None for col in DISCOVERY_COLUMNS}
        new_row["verification_status"] = "unverified"
        new_row["already_in_corpus"] = False
        new_row["name"] = r.get("name")
        new_row["other_urls"] = r.get("reference_urls")
        new_row["discovery_paths"] = "pre-existing: frontend admin-panel research-target card (pre-dates the discovery-round tracking structure)"
        new_row["notes"] = r.get("description")
        new_row["claims_source"] = "carried over from the original hardcoded frontend research-target list, not from an automated research pass"
        new_row["date_added"] = IMPORT_DATE
        discovery_rows.append(new_row)

    # ── Write the new workbook ──────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws_readme = wb.active
    ws_readme.title = "Read Me"
    for i, (text, bold) in enumerate(READ_ME_LINES, start=1):
        cell = ws_readme.cell(row=i, column=1, value=text)
        if bold:
            cell.font = Font(bold=True)
    ws_readme.column_dimensions["A"].width = 90

    ws_disc = wb.create_sheet("Discovery")
    ws_disc.append(DISCOVERY_COLUMNS)
    style_header(ws_disc, DISCOVERY_COLUMNS)
    for row in discovery_rows:
        ws_disc.append([row[c] for c in DISCOVERY_COLUMNS])
    size_columns(
        ws_disc, DISCOVERY_COLUMNS,
        wide={"notes", "discovery_paths", "corpus_match_notes", "claims_source",
              "main_url", "blog_or_articles_url", "archive_url", "other_urls", "name",
              "claimed_licensing_status"},
        narrow={"date_added", "already_in_corpus", "verification_status"},
    )
    add_dropdown(ws_disc, DISCOVERY_COLUMNS, "verification_status", ["unverified", "in_progress", "verified", "rejected"])
    add_dropdown(ws_disc, DISCOVERY_COLUMNS, "already_in_corpus", ["TRUE", "FALSE"])
    add_dropdown(ws_disc, DISCOVERY_COLUMNS, "category", ["practitioner_teacher", "academic_scholar", "historical_primary_source_archive"])
    add_dropdown(ws_disc, DISCOVERY_COLUMNS, "living_or_deceased", ["living", "deceased", "historical"])
    add_dropdown(ws_disc, DISCOVERY_COLUMNS, "claimed_written_content_exists", ["TRUE", "FALSE", "unknown"])

    ws_queue = wb.create_sheet("Queue")
    ws_queue.append(QUEUE_COLUMNS)
    style_header(ws_queue, QUEUE_COLUMNS)
    for row in queue_rows:
        ws_queue.append([row[c] for c in QUEUE_COLUMNS])
    size_columns(
        ws_queue, QUEUE_COLUMNS,
        wide={"name", "url", "final_url", "notes", "flag_reason", "promoted_from_discovery"},
        narrow={"source_db_id", "submitted_by", "result_document_id", "content_sha256",
                "worker_id", "lease_expires_at", "run_after", "created_at", "updated_at"},
    )
    add_dropdown(ws_queue, QUEUE_COLUMNS, "stage", ["ready_to_queue", "already_queued", "done"])
    add_dropdown(ws_queue, QUEUE_COLUMNS, "source_format", ["web_page", "pdf"])
    add_dropdown(ws_queue, QUEUE_COLUMNS, "source_scope", ["single", "collection"])
    add_dropdown(ws_queue, QUEUE_COLUMNS, "attribution_mode", ["declared", "per_item"])
    add_dropdown(ws_queue, QUEUE_COLUMNS, "on_unknown_author", ["flag", "skip"])
    add_dropdown(ws_queue, QUEUE_COLUMNS, "retain_original_text", ["TRUE", "FALSE"])
    add_dropdown(ws_queue, QUEUE_COLUMNS, "cleared_to_run", ["TRUE", "FALSE"])

    wb.save(str(SHEET_PATH))
    print(f"Wrote restructured workbook: {SHEET_PATH}")
    print(f"  Discovery tab: {len(discovery_rows)} rows, {len(DISCOVERY_COLUMNS)} columns")
    print(f"  Queue tab: {len(queue_rows)} rows, {len(QUEUE_COLUMNS)} columns")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
