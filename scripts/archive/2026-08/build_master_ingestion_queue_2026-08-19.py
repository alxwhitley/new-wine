#!/usr/bin/env python3
"""Build docs/ingestion/master_ingestion_queue.xlsx — the one-time initial
population of the new master ingestion-candidate spreadsheet.

Session scope (Alex, 2026-08-19): a spreadsheet checked into the repo
becomes the single master source of truth for ingestion candidates; the
admin panel becomes a read-only mirror fed by a LATER sync script (not
built here); on disagreement the spreadsheet wins. This script only reads
source_ingest_queue (SELECT only, explicit read-only transaction) and reads
the hardcoded frontend research-target card list. It writes the workbook
and nothing else -- no database writes, no edits to the frontend file, no
sync script.

Run once. Re-running would duplicate rows -- this is a build record, not a
reusable importer.
"""
from __future__ import annotations

import json
import os
import sys
from pathlib import Path

import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent.parent.parent
OUT_PATH = ROOT / "docs" / "ingestion" / "master_ingestion_queue.xlsx"
IMPORT_DATE = "2026-08-19"

load_dotenv(ROOT / "backend" / "app" / ".env")

# ── Master Queue column schema ──────────────────────────────────────────────
# Superset of source_ingest_queue's own columns (so a later sync script can
# write a DB row from any "ready_to_queue" row without losing information)
# plus loose research-candidate columns. One flat tab, not per-source detail
# tabs like the YouTube/magazine trackers -- see the Read Me tab for why.
COLUMNS = [
    "stage",                  # research_candidate | ready_to_queue | already_queued | done
    "name",
    "description",
    "url",                    # single confirmed URL, once decided
    "reference_urls",         # research-stage leads; semicolon-separated
    "source_format",          # web_page | pdf
    "source_scope",           # single | collection
    "attribute_to",
    "attribution_mode",       # declared | per_item
    "on_unknown_author",      # flag | skip
    "retain_original_text",   # TRUE/FALSE; blank = not yet decided
    "notes",
    "flag_reason",
    "cleared_to_run",         # TRUE/FALSE; blank = not applicable yet
    "db_status",              # source_ingest_queue.status, verbatim, DB rows only
    "db_execution_stage",     # source_ingest_queue.stage (runner's internal execution
                               # stage), verbatim, DB rows only -- distinct from our own
                               # "stage" column in column A
    "attempts",
    "max_attempts",
    "worker_id",
    "lease_expires_at",
    "run_after",
    "final_url",
    "content_sha256",
    "fetched_bytes",
    "attempted_documents",
    "stored_documents",
    "skipped_documents",
    "errored_documents",
    "result_document_id",
    "submitted_by",            # raw auth user id, DB rows only
    "source_db_id",            # source_ingest_queue.id -- sync match key for DB-derived rows
    "source_card_id",          # frontend card id -- prevents re-importing the same card
    "origin",
    "created_at",
    "updated_at",
]

READ_ME_LINES = [
    ("Master Ingestion Queue -- Read Me", True),
    ("", False),
    ("This workbook is the single master source of truth for ingestion", False),
    ("candidates (Alex's decision, 2026-08-19). The admin panel's ingest", False),
    ("queue table is a read-only mirror, kept in sync by a LATER script that", False),
    ("does not exist yet. Until that sync script is built, this workbook and", False),
    ("the database queue can drift -- edit here, not there.", False),
    ("", False),
    ("On any disagreement between this sheet and the database, this sheet", False),
    ("wins and silently overwrites, once the sync script exists.", False),
    ("", False),
    ("The existing admin submission form is untouched and still works; rows", False),
    ("submitted through it live only in the database until the sync script", False),
    ("is built, at which point they should be pulled into this sheet too.", False),
    ("", False),
    ("STAGE column (Master Queue tab, column A) -- the one column that", True),
    ("matters most for triage:", True),
    ("  research_candidate  -- a loose idea. May have several possible URLs,", False),
    ("                         nothing confirmed yet (format, scope, who to", False),
    ("                         credit). Not ready to submit to the queue.", False),
    ("  ready_to_queue      -- fully specified (one URL, format, scope,", False),
    ("                         attribution decided) but not yet in the", False),
    ("                         database queue. This is what the future sync", False),
    ("                         script will push. No rows start in this state", False),
    ("                         -- promote a research_candidate row to it by", False),
    ("                         hand once it's fleshed out.", False),
    ("  already_queued      -- already exists as a row in the database queue", False),
    ("                         (source_db_id is filled in) and is not yet", False),
    ("                         marked done there.", False),
    ("  done                -- already exists in the database queue with a", False),
    ("                         'done' status.", False),
    ("", False),
    ("Blank vs FALSE: for the TRUE/FALSE columns (retain_original_text,", True),
    ("cleared_to_run), a blank cell means 'not yet decided' -- it is NOT the", False),
    ("same as FALSE. Only research_candidate and freshly-promoted", False),
    ("ready_to_queue rows should ever have these blank.", False),
    ("", False),
    ("db_status / db_execution_stage vs this sheet's own stage column:", True),
    ("db_status and db_execution_stage are copied verbatim from the database", False),
    ("queue row's own status/stage fields (its internal run bookkeeping) for", False),
    ("already_queued and done rows. They are a different thing from this", False),
    ("sheet's own STAGE column in the Master Queue tab, which is our", False),
    ("triage/workflow category. Both are kept so no information is lost.", False),
    ("", False),
    ("source_db_id / source_card_id: whichever of these is filled in on a", True),
    ("row records exactly where it came from -- a database queue row id, or", False),
    ("a frontend research-card id -- so a later sync script (or a human) can", False),
    ("tell whether a row already has a database counterpart, and never", False),
    ("double-import the same candidate.", False),
    ("", False),
    ("Layout: one flat Master Queue tab holds every candidate, rather than", True),
    ("the YouTube/magazine trackers' one-tab-per-source pattern. Those", False),
    ("trackers split by source because each source has many individual video", False),
    ("rows needing separate per-channel filter settings. Here every row", False),
    ("already IS one candidate (one URL, one idea) with no natural per-", False),
    ("source grouping, so a single sortable/filterable table serves better", False),
    ("than splitting a handful of rows across many tabs.", False),
    ("", False),
    (f"Initial population: {IMPORT_DATE}. Built by reading every row then in", False),
    ("the database ingest queue, and every card then in the frontend's", False),
    ("hardcoded research-target list. See git history for the build record.", False),
]


def fetch_queue_rows():
    conn = psycopg2.connect(os.environ["SUPABASE_DB_URL"])
    conn.autocommit = False
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("BEGIN READ ONLY")
        cur.execute("SELECT * FROM source_ingest_queue ORDER BY created_at ASC")
        rows = [dict(r) for r in cur.fetchall()]
    finally:
        conn.rollback()  # read-only transaction; nothing to commit
        conn.close()
    return rows


# Hand-written readable names for the 6 live queue rows (source_ingest_queue
# has no "name"/"title" field of its own). Verified against each row's url +
# notes at build time (2026-08-19) -- not derived from any DB column.
DB_ROW_NAMES = {
    "8e8f23e0-7dc6-4057-aa4d-c07f1b607c99": "John Charles Ryle — Worldly Conformity (1878, PDF)",
    "85962adf-f4d6-440a-bd32-de414dbc4605": "Vlad Savchuk — How to Develop Your Prayer Language in Private",
    "fd16372d-aa6a-42f2-80be-094c7396f3b1": "Vlad Savchuk / Lana Savchuk — Intrusive Thoughts: Demon, Stronghold, or Just Your Mind? (byline mismatch, flagged)",
    "2f18306f-b910-4493-a87a-39cfd1454f2f": "Vlad Savchuk — 10 Ways to Know the Holy Spirit Better",
    "c2f52424-125a-43a0-9c3e-1f635eb57bfb": "Vlad Savchuk — Planted Not Buried: What God Is Doing While You Wait",
    "fbcc5a42-2866-4a28-909e-c66ffeda90dc": "Vlad Savchuk — Signs the Enemy Is Attacking Your Mind and How to Fight Back",
}


def db_row_to_sheet_row(r: dict) -> dict:
    status = r["status"]
    if status == "done":
        stage = "done"
    else:
        # waiting / running / failed / needs_attention -- still sitting in
        # the database queue, not finished.
        stage = "already_queued"

    return {
        "stage": stage,
        "name": DB_ROW_NAMES.get(r["id"], ""),
        "description": "",
        "url": r["url"],
        "reference_urls": "",
        "source_format": r["source_format"],
        "source_scope": r["source_scope"],
        "attribute_to": r["attribute_to"] or "",
        "attribution_mode": r["attribution_mode"],
        "on_unknown_author": r["on_unknown_author"],
        "retain_original_text": r["retain_original_text"],
        "notes": r["notes"] or "",
        "flag_reason": r["flag_reason"] or "",
        "cleared_to_run": r["cleared_to_run"],
        "db_status": r["status"],
        "db_execution_stage": r["stage"],
        "attempts": r["attempts"],
        "max_attempts": r["max_attempts"],
        "worker_id": r["worker_id"] or "",
        "lease_expires_at": str(r["lease_expires_at"]) if r["lease_expires_at"] else "",
        "run_after": str(r["run_after"]) if r["run_after"] else "",
        "final_url": r["final_url"] or "",
        "content_sha256": r["content_sha256"] or "",
        "fetched_bytes": r["fetched_bytes"],
        "attempted_documents": r["attempted_documents"],
        "stored_documents": r["stored_documents"],
        "skipped_documents": r["skipped_documents"],
        "errored_documents": r["errored_documents"],
        "result_document_id": str(r["result_document_id"]) if r["result_document_id"] else "",
        "submitted_by": str(r["submitted_by"]) if r["submitted_by"] else "",
        "source_db_id": str(r["id"]),
        "source_card_id": "",
        "origin": f"admin_ingest_queue_{IMPORT_DATE}",
        "created_at": str(r["created_at"]),
        "updated_at": str(r["updated_at"]),
    }


# Exact card content, copied verbatim from
# frontend/components/admin/corpus-data.ts FUTURE_TARGETS at build time
# (2026-08-19) -- that file was only read, never modified.
CARDS = [
    {
        "id": "andrew-murray-extra",
        "name": "Andrew Murray — Additional Titles",
        "description": "More Andrew Murray works beyond the 4 already ingested. Available on olddeadguys.com and Internet Archive.",
        "urls": [
            "https://www.olddeadguys.com/andrew-murray",
            "https://archive.org/search?query=andrew+murray&mediatype=texts",
        ],
    },
    {
        "id": "azusa-apostolic-faith",
        "name": "Azusa Street — Apostolic Faith Magazine",
        "description": "13 issues of the original Apostolic Faith newsletter from the Azusa Street Revival (1906–1908). Primary source documents of the Pentecostal movement.",
        "urls": ["https://place.asburyseminary.edu/apostolicfaith/"],
    },
    {
        "id": "pentecostal-archives",
        "name": "Consortium of Pentecostal Archives",
        "description": "Digital archives from Pentecostal and charismatic institutions. Explore for freely downloadable primary source documents.",
        "urls": ["https://www.pentecostalarchives.org"],
    },
    {
        "id": "frank-bartleman",
        "name": "Frank Bartleman — Azusa Street Writings",
        "description": "Frank Bartleman’s firsthand accounts of the Azusa Street Revival. Multiple titles available on Internet Archive.",
        "urls": ["https://archive.org/search?query=frank+bartleman&mediatype=texts"],
    },
    {
        "id": "stepbible-tipnr",
        "name": "STEPBible — TIPNR (Proper Names)",
        "description": "Every proper noun in the Bible with exhaustive references, family relationships, geolocation, and descriptions. CC BY 4.0.",
        "urls": ["https://github.com/STEPBible/STEPBible-Data"],
    },
    {
        "id": "stepbible-tahot-ref",
        "name": "STEPBible — TAHOT (Hebrew OT) Additional Books",
        "description": "Already ingested full OT. This card is for reference if re-ingestion or updates are needed.",
        "urls": [
            "https://github.com/STEPBible/STEPBible-Data/tree/master/Translators%20Amalgamated%20OT%2BNT",
        ],
    },
]


def card_to_sheet_row(c: dict) -> dict:
    return {
        "stage": "research_candidate",
        "name": c["name"],
        "description": c["description"],
        "url": "",
        "reference_urls": "; ".join(c["urls"]),
        "source_format": "",
        "source_scope": "",
        "attribute_to": "",
        "attribution_mode": "",
        "on_unknown_author": "",
        "retain_original_text": None,
        "notes": "",
        "flag_reason": "",
        "cleared_to_run": None,
        "db_status": "",
        "db_execution_stage": "",
        "attempts": None,
        "max_attempts": None,
        "worker_id": "",
        "lease_expires_at": "",
        "run_after": "",
        "final_url": "",
        "content_sha256": "",
        "fetched_bytes": None,
        "attempted_documents": None,
        "stored_documents": None,
        "skipped_documents": None,
        "errored_documents": None,
        "result_document_id": "",
        "submitted_by": "",
        "source_db_id": "",
        "source_card_id": c["id"],
        "origin": f"frontend_research_card_{IMPORT_DATE}",
        "created_at": IMPORT_DATE,
        "updated_at": IMPORT_DATE,
    }


def build_workbook(sheet_rows: list[dict]) -> Workbook:
    wb = Workbook()

    # ── Read Me tab ──────────────────────────────────────────────────────
    ws_readme = wb.active
    ws_readme.title = "Read Me"
    for i, (text, bold) in enumerate(READ_ME_LINES, start=1):
        cell = ws_readme.cell(row=i, column=1, value=text)
        if bold:
            cell.font = Font(bold=True)
    ws_readme.column_dimensions["A"].width = 90

    # ── Master Queue tab ─────────────────────────────────────────────────
    ws = wb.create_sheet("Master Queue")
    ws.append(COLUMNS)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    for col_idx, name in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True)
        c.fill = header_fill
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for row in sheet_rows:
        ws.append([row[col] for col in COLUMNS])

    # Reasonable column widths -- narrow id/flag columns, wide text columns.
    WIDE = {"description", "notes", "flag_reason", "reference_urls", "url", "final_url", "name"}
    NARROW = {
        "source_db_id", "source_card_id", "submitted_by", "result_document_id",
        "content_sha256", "worker_id", "lease_expires_at", "run_after",
        "created_at", "updated_at",
    }
    for col_idx, name in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        if name in WIDE:
            ws.column_dimensions[letter].width = 48
        elif name in NARROW:
            ws.column_dimensions[letter].width = 20
        else:
            ws.column_dimensions[letter].width = 16

    # Dropdown validation on the columns most prone to hand-typo drift.
    def add_dropdown(col_name: str, options: list[str]):
        col_idx = COLUMNS.index(col_name) + 1
        letter = get_column_letter(col_idx)
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(options) + '"',
            allow_blank=True,
            showDropDown=False,
        )
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}2000")

    add_dropdown("stage", ["research_candidate", "ready_to_queue", "already_queued", "done"])
    add_dropdown("source_format", ["web_page", "pdf"])
    add_dropdown("source_scope", ["single", "collection"])
    add_dropdown("attribution_mode", ["declared", "per_item"])
    add_dropdown("on_unknown_author", ["flag", "skip"])
    add_dropdown("retain_original_text", ["TRUE", "FALSE"])
    add_dropdown("cleared_to_run", ["TRUE", "FALSE"])

    return wb


def main() -> int:
    db_rows = fetch_queue_rows()
    print(f"Database rows read: {len(db_rows)}")
    for r in db_rows:
        print(f"  - {r['id']}  status={r['status']}  url={r['url']}")

    print(f"Card entries read: {len(CARDS)}")
    for c in CARDS:
        print(f"  - {c['id']}")

    sheet_rows = [db_row_to_sheet_row(r) for r in db_rows] + [card_to_sheet_row(c) for c in CARDS]

    stage_counts: dict[str, int] = {}
    for row in sheet_rows:
        stage_counts[row["stage"]] = stage_counts.get(row["stage"], 0) + 1

    print(f"Spreadsheet rows to write: {len(sheet_rows)}")
    print(f"Stage breakdown: {json.dumps(stage_counts, indent=2)}")

    assert len(sheet_rows) == len(db_rows) + len(CARDS), "row count mismatch before write"

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook(sheet_rows)
    wb.save(str(OUT_PATH))
    print(f"Wrote: {OUT_PATH}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
