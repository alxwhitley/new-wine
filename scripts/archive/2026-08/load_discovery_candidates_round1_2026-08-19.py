#!/usr/bin/env python3
"""Load the Round 1/2/3 research-candidate list (Send Conference network map,
grassroots obscure pass, scholarly Pentecostal/charismatic academics pass)
into the Discovery tab of docs/ingestion/master_ingestion_queue.xlsx.

Source: Alex's pasted research output, 2026-08-19. Every candidate is loaded
unfiltered -- no relevance judgment, nothing dropped. The 6 original
research-target cards already on the Discovery tab (Round 0) are left
untouched and not duplicated.

Two schema additions made in this pass, beyond what Task 2 originally
specified -- see the Read Me tab and the session report for why:
  - living_or_deceased gains a fourth option, "unknown". Most of this list
    states neither a death date nor "deceased" -- guessing "living" for
    someone with zero stated vital status would itself be an unverified
    claim dressed as a fact, which is exactly what this tab exists to avoid.
  - category gains a fourth option, "worship_leader_musician", for the small
    number of candidates whose primary public role is worship/music rather
    than teaching, and who are explicitly noted as having no written content
    at all -- neither "practitioner_teacher" nor the other two existing
    options describe them honestly.

Corpus cross-check: reads the live `sources` table (read-only, via the
dedicated read-only analysis role) and matches by first+last name token,
case-insensitive, so "Craig S. Keener" still matches the corpus's
"Craig Keener". A name/organization match is recorded as a note but does
NOT set already_in_corpus -- e.g. Michael/Jessica Koulianos's ministry
"Jesus Image" already exists as a corpus source, but neither of them does
as an individual, and conflating the two would misrepresent them as already
present when they are not.

Repo-file-write plus one read-only database query. No database writes.
"""
from __future__ import annotations

import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from dotenv import load_dotenv

ROOT = Path(__file__).resolve().parent.parent.parent.parent
SHEET_PATH = ROOT / "docs" / "ingestion" / "master_ingestion_queue.xlsx"
IMPORT_DATE = "2026-08-19"

load_dotenv(ROOT / "backend" / "app" / ".env.readonly-analysis")

DISCOVERY_COLUMNS = [
    "verification_status",
    "already_in_corpus",
    "name",
    "organization",
    "location",
    "category",
    "living_or_deceased",
    "main_url",
    "blog_or_articles_url",
    "archive_url",
    "other_urls",
    "claimed_written_content_exists",
    "claimed_licensing_status",
    "claimed_platform_size",
    "discovery_paths",
    "corpus_match_notes",
    "claims_source",
    "notes",
    "date_added",
]

CLAIMS_SOURCE_R1R2 = (
    "Automated research tool inferring from search results (Round 1-2 "
    "Send Conference network-mapping / grassroots pass, 2026-08-19) -- "
    "content type, licensing, and platform size are inferred claims, not "
    "confirmed by visiting any site. No URLs were captured in this pass."
)
CLAIMS_SOURCE_R3 = (
    "Automated research tool compiling academic-press bibliographic "
    "information (Round 3 scholarly Pentecostal/charismatic academics "
    "pass, 2026-08-19). Licensing assumed as standard copyright per the "
    "research pass's own stated instruction ('assume standard copyright "
    "unless verified otherwise') -- not independently verified. Platform "
    "size and URLs were not part of this pass's fields."
)

DECEASED = {
    "loren cunningham", "reinhard bonnke", "steve hill",
    "gordon d. fee", "roger stronstad", "stanley m. horton",
}


def r1(name, org, loc, content, size, license_, seed, notes=""):
    return {
        "name": name, "organization": org, "location": loc,
        "category": "practitioner_teacher",
        "claimed_written_content_exists": False if "NO written content" in content else True,
        "claimed_licensing_status": license_,
        "claimed_platform_size": size,
        "discovery_paths": f"Round 1, Seed {seed} (Send Conference network map)",
        "notes": (content + (" -- " + notes if notes else "")),
        "claims_source": CLAIMS_SOURCE_R1R2,
    }


def r2(name, org, loc, content, footprint, note):
    return {
        "name": name, "organization": org, "location": loc,
        "category": "practitioner_teacher",
        "claimed_written_content_exists": True,
        "claimed_licensing_status": "unstated",
        "claimed_platform_size": footprint,
        "discovery_paths": "Round 2 (obscure/non-mainstream grassroots pass)",
        "notes": content + " -- " + note,
        "claims_source": CLAIMS_SOURCE_R1R2,
    }


def r3(name, org, works, seed_tag=""):
    return {
        "name": name, "organization": org, "location": "",
        "category": "academic_scholar",
        "claimed_written_content_exists": True,
        "claimed_licensing_status": "standard copyright (assumed -- academic-press published, not independently verified)",
        "claimed_platform_size": "",
        "discovery_paths": "Round 3 (scholarly Pentecostal/charismatic academics pass)",
        "notes": ("Works: " + works) + (f" [{seed_tag}]" if seed_tag else ""),
        "claims_source": CLAIMS_SOURCE_R3,
    }


CANDIDATES = [
    # ── Round 1, Seed A: Bill Johnson / Ben Fitzgerald ──────────────────────
    r1("Bill Johnson", "Bethel Church", "Redding CA", "has articles/books/devotionals", "large global", "standard copyright", "A"),
    r1("Ben Fitzgerald", "Awakening Europe / Awakening Church", "Eimeldingen Germany", "ministry updates, essays", "SMALL LOCAL CHURCH", "standard copyright", "A"),
    r1("Kris Vallotton", "Bethel Church / BSSM", "", "active personal blog", "large", "standard copyright", "A"),
    r1("Shawn Bolz", "Bolz Ministries / iCreate", "", "ebooks, written devotionals", "large", "standard copyright", "A"),
    r1("Dan McCollam", "Prophetic Company / Bethel", "", "articles, training manuals", "small/medium", "standard copyright", "A"),
    r1("Randy Clark", "Global Awakening", "", "articles, academic papers", "large", "standard copyright", "A"),
    r1("Heidi Baker", "Iris Global", "", "missions letters, articles", "large", "standard copyright", "A"),
    r1("Ché Ahn", "Harvest International Ministry", "", "articles, apostolic letters", "large", "standard copyright", "A"),
    r1("Billy Wilson", "Empowered21 / Oral Roberts University", "", "academic & ministry essays", "large", "standard copyright", "A"),
    r1("Ken & Jen Hodges", "Christalignment", "Australia", "NO written content, audio/video only", "small local", "standard copyright", "A"),
    r1("Sean Feucht", "Let Us Worship / Bold", "", "blog posts, op-eds", "large", "standard copyright", "A",
       notes="Source list tags this as '[ALSO APPEARS IN SEED D]' -- could not independently confirm a second appearance in Seed D as pasted; flagged for Alex to check original research notes."),
    r1("Brian & Jenn Johnson", "Bethel Music", "", "no written content, lyrics/video", "large", "CCLI royalty", "A"),
    r1("Bob Hasson", "Bolz Ministries / author", "", "business & faith articles", "small/medium", "standard copyright", "A"),
    r1("Todd Lollar", "Mobilize Ministries", "Ojai CA & DFW TX", "extensive devotional blog", "~21,000 followers", "OPEN LICENSE / YouVersion partner", "A",
       notes="Source list tags this as '[ALSO APPEARS IN ROUND 2]' -- could not independently confirm a second appearance in the Round 2 list as pasted; flagged for Alex to check original research notes."),
    r1("Johannes Hartl", "Gebetshaus Augsburg", "Germany", "theological essays, blog", "medium European", "standard copyright", "A"),
    r1("Steve Cuss", "CapCity Church", "Austin TX", "leadership articles, blog", "SMALL LOCAL CHURCH", "standard copyright", "A"),
    r1("Lance Wallnau", "Lance Learning Group", "", "commentary articles, blog", "large", "standard copyright", "A"),
    r1("Eric Metaxas", "Metaxas Media / Socrates", "", "articles, transcripts", "large", "standard copyright", "A"),
    r1("Lou Engle", "Lou Engle Ministries / TheCall", "", "teaching notes, ebooks", "large", "standard copyright / some free ebooks", "A"),
    r1("Jonathan Cahn", "Hope of the World Ministries", "", "articles, prophetic writings", "large", "standard copyright", "A"),
    r1("Dean Briggs", "Lou Engle Ministries", "", "blog, prayer guides", "small/medium", "standard copyright", "A"),
    r1("Matt Lockett", "Justice House of Prayer DC", "Washington DC", "articles, historic essays", "small local", "standard copyright", "A"),
    r1("Dick Eastman", "Every Home for Christ", "", "articles, prayer guides", "large", "standard copyright", "A"),
    r1("Chris Berglund", "Ascendance Ministries", "", "teaching blogs, dream notes", "small local", "standard copyright", "A"),
    r1("Will Ford", "818 Texas / JHOP DC", "", "historical articles, blog", "small local", "standard copyright", "A"),
    r1("David Kim", "Contend Global", "", "fast guides, articles", "small youth prayer ministry", "standard copyright", "A"),
    r1("Scott Volk", "Together for Israel", "", "articles, ministry letters", "small local", "standard copyright", "A"),

    # ── Round 1, Seed B: Francis Chan ────────────────────────────────────────
    r1("Francis Chan", "Crazy Love / We Are Church", "", "books, articles, study guides", "small house church network", "standard copyright, some open study PDFs", "B"),
    r1("Lisa Chan", "Crazy Love Ministries", "", "co-authored books, blog articles", "small house church network", "unstated", "B"),
    r1("Nathan J'Diim", "Mission Connexion", "", "missions articles, guides", "small/medium", "standard copyright", "B"),
    r1("Carey Nieuwhof", "Carey Nieuwhof Leadership", "", "extensive leadership blog, transcripts", "large media platform", "standard copyright", "B"),
    r1("Mike Bickle", "IHOPKC", "", "PDFs, teaching outlines, articles", "large", "OPEN ACCESS ARCHIVES / standard copyright", "B"),
    r1("John Mark Comer", "Practicing the Way / Bridgetown", "Portland OR", "essays, formation guides, blog", "SMALL LOCAL CHURCH", "standard copyright, some free open guides", "B"),
    r1("Christine Caine", "A21 / Propel Women", "", "blog articles, books", "large", "standard copyright", "B"),
    r1("Craig Groeschel", "Life.Church", "", "leadership blog, devotionals", "large megachurch", "OPEN LICENSE (Open Network resources)", "B"),
    r1("Jon Tyson", "Church of the City New York", "", "articles, cultural essays", "SMALL/MEDIUM LOCAL CHURCH", "standard copyright", "B"),
    r1("Lisa Harper", "Lisa Harper Ministries", "", "devotional articles", "large", "standard copyright", "B"),
    r1("Mark Sayers", "Red Church", "Melbourne, Australia", "cultural analysis articles, blog", "SMALL LOCAL CHURCH", "standard copyright", "B"),
    r1("Pete Greig", "24-7 Prayer / Emmaus Road", "Guildford UK", "prayer guides, articles, blog", "small/medium local", "CREATIVE COMMONS / open prayer guides", "B"),
    r1("Alain Emerson", "24-7 Prayer Ireland / Emmanuel Church", "Lurgan NI", "blog articles, pastoral letters", "SMALL LOCAL CHURCH", "standard copyright", "B"),
    r1("Emma Stark", "Powerhouse Church / Glasgow Prophetic Centre", "Scotland", "prophetic articles, blog", "SMALL LOCAL CHURCH", "standard copyright", "B"),

    # ── Round 1, Seed C: Circuit Riders ──────────────────────────────────────
    r1("Circuit Riders Music", "Circuit Riders / YWAM", "", "training blogs, tour manuals", "large youth movement", "standard copyright", "C",
       notes="This is a ministry/brand entity, not an individual person."),
    r1("Brian & Christy Brennt", "Circuit Riders", "", "books, articles, prophetic papers", "large", "standard copyright", "C"),
    r1("Andy Byrd", "YWAM Kona / Circuit Riders", "", "missions articles, blog", "large", "standard copyright", "C",
       notes="Source list tags this as '[ALSO APPEARS IN SEED D]' -- could not independently confirm a second appearance in Seed D as pasted; flagged for Alex to check original research notes."),
    r1("Amy Sollars", "Circuit Riders / YWAM", "", "historical vision blog posts", "large", "standard copyright", "C"),
    r1("Brian Barcelona", "Gen Z For Jesus / One Voice Student Missions", "", "articles, outreach manuals", "large youth movement", "standard copyright", "C"),
    r1("Teo Hayashi", "Dunamis Movement / Zion Church", "São Paulo Brazil", "leadership articles, blog", "small/medium local network", "standard copyright", "C"),
    r1("Michael Miller", "UPPERROOM Dallas", "", "devotional articles, teaching notes", "SMALL/MEDIUM LOCAL CHURCH", "standard copyright", "C"),
    r1("Junia Hayashi", "Dunamis Movement", "", "articles, women's blog", "small/medium", "unstated", "C"),
    r1("Lorisa Miller", "UPPERROOM Dallas", "", "prayer guides, articles", "small/medium local", "unstated", "C"),
    r1("Brandon Hampton", "UPPERROOM Music", "", "NO written content, audio/video", "small/medium", "standard copyright", "C"),
    r1("Peter Mattis", "Grace Center", "Franklin TN", "pastoral blogs, teaching notes", "SMALL LOCAL CHURCH", "standard copyright", "C"),
    r1("Ray Hughes", "Selah Ministries", "", "articles, historical worship studies", "small/medium", "standard copyright", "C"),
    r1("Joel Richardson", "Joel's Trumpet", "", "extensive theological blog, articles", "small/medium", "open online teaching / standard copyright", "C"),

    # ── Round 1, Seed D: Fire and Fragrance ──────────────────────────────────
    r1("Fire and Fragrance", "YWAM Kona", "", "missions blogs, school manuals", "large", "standard copyright", "D",
       notes="This is a ministry/brand entity, not an individual person."),
    r1("Amy Ward", "Fire & Fragrance Kona", "", "leadership articles, blog", "large", "standard copyright", "D"),
    r1("Loren Cunningham", "YWAM International", "", "books, global missions essays", "large", "standard copyright", "D",
       notes="Source list marks this person deceased."),
    r1("Lindy Cofer", "Lindy & The Circuit Riders", "", "NO written content, music/video", "large", "standard copyright", "D"),
    r1("Darlene Cunningham", "YWAM International", "", "leadership letters, articles", "large", "standard copyright", "D"),
    r1("David Hamilton", "YWAM University of the Nations", "", "academic papers, biblical studies", "large academic", "standard copyright", "D"),
    r1("Paul Eshleman", "Jesus Film Project / Finishing the Task", "", "global strategy reports, articles", "large coalition", "PUBLIC DOMAIN / open missional access", "D"),
    r1("Steve Douglass", "Cru (Campus Crusade)", "", "leadership articles, books", "large", "standard copyright", "D"),
    r1("Mark Gauthier", "Cru USA", "", "ministry articles, op-eds", "large", "standard copyright", "D"),

    # ── Round 1, Seed E: Send core lineup ───────────────────────────────────
    r1("Michael Koulianos", "Jesus Image / Jesus School", "Orlando FL", "newsletter, books, articles", "large", "standard copyright", "E",
       notes="Ministry organization 'Jesus Image' already exists as a corpus source; see corpus_match_notes."),
    r1("Daniel Kolenda", "CfaN / Nations Church", "Orlando FL", "extensive blog, books, press", "large + local church", "standard copyright", "E"),
    r1("Todd White", "Lifestyle Christianity", "", "training manuals, devotionals", "large", "standard copyright", "E"),
    r1("Jessica Koulianos", "Jesus Image", "", "ministry essays", "large", "standard copyright", "E",
       notes="Ministry organization 'Jesus Image' already exists as a corpus source; see corpus_match_notes."),
    r1("Benny Hinn", "Benny Hinn Ministries", "", "teaching articles, books", "large", "standard copyright", "E"),
    r1("Reinhard Bonnke", "CfaN", "", "articles, books, legacy blog", "large", "standard copyright", "E",
       notes="Source list marks this person deceased."),
    r1("Steffany Gretzinger", "Jesus School / worship", "", "NO written content", "large", "standard copyright", "E"),
    r1("Brother Yun", "Back to Jerusalem", "", "missions letters, articles", "large", "standard copyright", "E"),
    r1("Russell Benson", "CfaN Global", "", "global campaign reports", "large", "standard copyright", "E"),
    r1("Peter Vandenberg", "CfaN Global", "", "executive articles, news", "large", "standard copyright", "E"),
    r1("Peter Youngren", "World Impact Ministries", "", "theological blog, articles", "large", "standard copyright", "E"),
    r1("Mario Murillo", "Mario Murillo Ministries", "", "active political/faith blog", "large", "standard copyright", "E"),
    r1("Steve Hill", "Steve Hill Ministries", "", "legacy revival articles", "large", "standard copyright", "E",
       notes="Source list marks this person deceased."),
    r1("John Kilpatrick", "Church of His Presence", "Daphne AL", "sermon transcripts, blog", "SMALL LOCAL CHURCH", "standard copyright", "E"),
    r1("Nathan Morris", "Shake the Nations", "", "revival articles, blog", "small/medium global", "standard copyright", "E"),
    r1("Lydia Stanley Morris", "Shake the Nations", "", "ministry essays", "small/medium", "standard copyright", "E"),
    r1("Jean-Luc Trachsel", "Europe Shall Be Saved", "Switzerland", "articles, European reports", "medium European", "standard copyright", "E"),

    # ── Round 2: obscure / non-mainstream grassroots ────────────────────────
    r2("Dave Ward", "Twin View Campus, Bethel Church", "Redding CA", "profile & articles", "negligible personal platform", "Bethel staff, not independent"),
    r2("Jamila Page", "Bethel Church", "Redding CA", "features on diversity & culture", "low public footprint", "Bethel staff, not independent"),
    r2("Jason Chin", "Love Says Go Ministries", "Redding CA / regional", "blog entries via Destiny Image", "low profile regional", "regional"),
    r2("Adam Shepski", "YFC Canada / regional prayer network", "Peterborough ON Canada", "ministry essays", "regional coordinator", "regional coordinator"),
    r2("Stephanie Schmidt", "American Faith & Family / Jesus School satellites", "Ghana & DFW TX", "\"The Reluctant Missionary\" blog", "grassroots missionary, small readership", "grassroots missionary"),
    r2("Chris Ryan", "Daystar Worship Center", "Gordonsville VA", "pastoral blog via regional denominational platform", "SINGLE-SITE LOCAL CHURCH, under 1,000 followers", "single-site local church"),

    # ── Round 3: scholarly Pentecostal/charismatic academics ───────────────
    r3("Gordon D. Fee", "Regent College", "God's Empowering Presence; Gospel and Spirit; NICNT 1 Corinthians", "SEED"),
    r3("Craig S. Keener", "Asbury Theological Seminary", "Acts commentary (4 vols); Miracles (2 vols); Spirit Hermeneutics", "SEED, alive/active"),
    r3("Roger Stronstad", "Summit Pacific College", "The Charismatic Theology of St. Luke; The Prophethood of All Believers", "SEED"),
    r3("Rick Wadholm Jr.", "AGTS", "A Theology of the Spirit in the Former Prophets; A Pentecostal Reading of Daniel"),
    r3("Paul Elbert", "Journal for Biblical and Pneumatological Research", "Essays on Apostolic Themes; The Lukan Gift of the Holy Spirit"),
    r3("Stanley M. Horton", "AGTS", "Systematic Theology: A Pentecostal Perspective"),
    r3("Byron D. Klaus", "AGTS (former President)", "Called and Empowered"),
    r3("Murray Dempster", "Professor of Social Ethics (institution not stated in source)", "Called and Empowered; Essays on Pentecostal Social Ethics"),
    r3("Amos Yong", "Fuller Seminary", "Spirit-Word-Community; The Spirit Poured Out on All Flesh; Disability and the Gifts of the Spirit"),
    r3("John Christopher Thomas", "Pentecostal Theological Seminary", "The Spirit of the New Testament; Pentecostal Commentary on Revelation"),
    r3("Martin W. Mittelstadt", "Evangel University", "The Spirit and Suffering in Luke-Acts"),
    r3("C. Bill Oliverio", "AGTS", "Theological Hermeneutics in the Classical Pentecostal Tradition"),
    r3("Robert P. Menzies", "Synergy (Director)", "Empowered for Witness; Pentecost: This Story is Our Story; Speaking in Tongues"),
    r3("Wonsuk Ma", "Oral Roberts University", "Until the Spirit Comes; Asian Journal of Pentecostal Studies"),
    r3("Mark J. Cartledge", "Professor of Practical Ministry / Renewal Studies (institution not stated in source)", "Encountering the Spirit; Testimony in the Spirit"),
    r3("Dale M. Coulter", "Pentecostal Theological Seminary", "Holiness: The Beauty of Perfection"),
    r3("Christopher A. Stephenson", "Lee University", "Types of Pentecostal Theology (Oxford UP)"),
    r3("Kenneth J. Archer", "Southeastern University", "A Pentecostal Hermeneutic for the Twenty-First Century"),
    r3("Frank D. Macchia", "Vanguard University", "Baptized in the Spirit; Justified in the Spirit; Jesus the Spirit Baptizer"),
    r3("Paul Nathan Alexander", "Professor of Social Ethics (institution not stated in source)", "Peace to All Who Are Far Off; Black Fire"),
    r3("Estrelda Y. Alexander", "William Joseph Seymour Institute", "Black Fire; The Women of Azusa Street"),
    r3("Lee Roy Martin", "Pentecostal Theological Seminary", "Pentecostal Hermeneutics; The Book of Judges"),
    r3("Taehyun Lee", "Ascent College", "dissertations/essays on Pauline corpus and renewal pneumatology"),
    r3("Todd Korpi", "Fuller / Ascent", "The Missional Vision of Digital Ministry"),
    r3("Lois E. Olena", "AGTS", "Children of the Calling"),
    r3("Paul Lewis", "AGTS", "papers on early church mission in China, historical pneumatology"),
]

MUSICIAN_OVERRIDE = {"Brian & Jenn Johnson", "Brandon Hampton", "Lindy Cofer", "Steffany Gretzinger"}
for c in CANDIDATES:
    if c["name"] in MUSICIAN_OVERRIDE:
        c["category"] = "worship_leader_musician"
    c["living_or_deceased"] = "deceased" if c["name"].lower() in DECEASED else "unknown"
    c["verification_status"] = "unverified"
    c["already_in_corpus"] = False
    c["corpus_match_notes"] = ""
    c["date_added"] = IMPORT_DATE
    for url_field in ("main_url", "blog_or_articles_url", "archive_url", "other_urls"):
        c.setdefault(url_field, "")
    if not c.get("notes", "").strip():
        c["notes"] = ""


def _name_tokens(full_name: str):
    parts = [p.strip(".,") for p in full_name.split() if p.strip(".,")]
    if not parts:
        return None
    return (parts[0].lower(), parts[-1].lower())


def cross_check_corpus(candidates: list) -> tuple:
    conn_url = os.environ["READONLY_ANALYSIS_DB_URL"]
    import psycopg2
    conn = psycopg2.connect(conn_url)
    conn.set_session(readonly=True, autocommit=True)
    cur = conn.cursor()
    cur.execute("SELECT id, name FROM sources ORDER BY name")
    corpus = cur.fetchall()
    conn.close()

    corpus_by_tokens = {}
    for cid, cname in corpus:
        t = _name_tokens(cname)
        if t:
            corpus_by_tokens.setdefault(t, []).append((cid, cname))

    matches = []
    for c in candidates:
        t = _name_tokens(c["name"])
        if t and t in corpus_by_tokens:
            hits = corpus_by_tokens[t]
            cid, cname = hits[0]
            c["already_in_corpus"] = True
            c["corpus_match_notes"] = f"Matches existing corpus source {cname!r} (id {cid})."
            matches.append((c["name"], cname))
    return matches, len(corpus)


def style_header(ws, columns):
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def size_columns(ws, columns, wide: set, narrow: set):
    for col_idx, name in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        if name in wide:
            ws.column_dimensions[letter].width = 48
        elif name in narrow:
            ws.column_dimensions[letter].width = 20
        else:
            ws.column_dimensions[letter].width = 16


def add_dropdown(ws, columns, col_name, options):
    col_idx = columns.index(col_name) + 1
    letter = get_column_letter(col_idx)
    dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"{letter}2:{letter}2000")


def main() -> int:
    wb = openpyxl.load_workbook(SHEET_PATH, data_only=True)
    disc_ws_old = wb["Discovery"]
    headers = [c.value for c in disc_ws_old[1]]
    assert headers == DISCOVERY_COLUMNS, "Discovery schema drifted since last build -- stopping"
    existing_rows = [dict(zip(headers, r)) for r in disc_ws_old.iter_rows(min_row=2, values_only=True) if any(v is not None for v in r)]
    print(f"Existing Discovery rows (Round 0, preserved as-is): {len(existing_rows)}")
    assert len(existing_rows) == 6

    existing_names = {r["name"] for r in existing_rows}
    for c in CANDIDATES:
        assert c["name"] not in existing_names, f"{c['name']!r} would duplicate a Round 0 row"

    # Within-batch duplicate name check.
    seen = {}
    for c in CANDIDATES:
        seen.setdefault(c["name"], []).append(c)
    dupes = {name: rows for name, rows in seen.items() if len(rows) > 1}
    if dupes:
        print(f"Duplicate names found within this batch, collapsing: {list(dupes.keys())}")
        collapsed = []
        used = set()
        for c in CANDIDATES:
            if c["name"] in used:
                continue
            group = seen[c["name"]]
            if len(group) > 1:
                merged = dict(group[0])
                merged["discovery_paths"] = "; ".join(dict.fromkeys(g["discovery_paths"] for g in group))
                merged["notes"] = " | ".join(dict.fromkeys(g["notes"] for g in group if g["notes"]))
                collapsed.append(merged)
            else:
                collapsed.append(c)
            used.add(c["name"])
        CANDIDATES[:] = collapsed
    else:
        print("No duplicate names within this batch.")

    unique_count = len(CANDIDATES)
    duplicates_collapsed = sum(len(v) - 1 for v in dupes.values())

    matches, corpus_size = cross_check_corpus(CANDIDATES)
    print(f"Corpus cross-check against {corpus_size} live sources:")
    for cand_name, corpus_name in matches:
        print(f"  MATCH: {cand_name!r} -> corpus source {corpus_name!r}")
    print(f"Total already-in-corpus flags: {len(matches)}")

    queue_ws = wb["Queue"]
    queue_headers = [c.value for c in queue_ws[1]]
    queue_rows = [dict(zip(queue_headers, r)) for r in queue_ws.iter_rows(min_row=2, values_only=True) if any(v is not None for v in r)]
    print(f"Queue tab rows (must stay untouched): {len(queue_rows)}")
    assert len(queue_rows) == 6

    read_me_ws = wb["Read Me"]
    read_me_lines = [c[0].value for c in read_me_ws.iter_rows(max_col=1) if c[0].value is not None or True]
    read_me_lines = [row[0].value for row in read_me_ws.iter_rows(max_col=1)]

    new_wb = openpyxl.Workbook()
    new_readme = new_wb.active
    new_readme.title = "Read Me"
    for i, val in enumerate(read_me_lines, start=1):
        cell = new_readme.cell(row=i, column=1, value=val)
    addendum_start = len(read_me_lines) + 2
    addendum = [
        ("Discovery-tab schema addendum, 2026-08-19 (Round 1-3 candidate load):", True),
        ("  living_or_deceased gained a fourth option, 'unknown' -- most", False),
        ("  candidates in this load have no stated vital status either way;", False),
        ("  guessing 'living' would itself be an unverified claim.", False),
        ("  category gained a fourth option, 'worship_leader_musician', for", False),
        ("  the handful of candidates whose public role is worship/music with", False),
        ("  explicitly no written content -- neither teacher, scholar, nor", False),
        ("  archive describes them honestly.", False),
    ]
    for j, (text, bold) in enumerate(addendum):
        cell = new_readme.cell(row=addendum_start + j, column=1, value=text)
        if bold:
            cell.font = Font(bold=True)
    new_readme.column_dimensions["A"].width = 90

    new_disc = new_wb.create_sheet("Discovery")
    new_disc.append(DISCOVERY_COLUMNS)
    style_header(new_disc, DISCOVERY_COLUMNS)
    all_disc_rows = existing_rows + CANDIDATES
    for row in all_disc_rows:
        new_disc.append([row.get(c) for c in DISCOVERY_COLUMNS])
    size_columns(
        new_disc, DISCOVERY_COLUMNS,
        wide={"notes", "discovery_paths", "corpus_match_notes", "claims_source",
              "main_url", "blog_or_articles_url", "archive_url", "other_urls", "name",
              "claimed_licensing_status", "organization"},
        narrow={"date_added", "already_in_corpus", "verification_status", "location"},
    )
    add_dropdown(new_disc, DISCOVERY_COLUMNS, "verification_status", ["unverified", "in_progress", "verified", "rejected"])
    add_dropdown(new_disc, DISCOVERY_COLUMNS, "already_in_corpus", ["TRUE", "FALSE"])
    add_dropdown(new_disc, DISCOVERY_COLUMNS, "category", ["practitioner_teacher", "academic_scholar", "historical_primary_source_archive", "worship_leader_musician"])
    add_dropdown(new_disc, DISCOVERY_COLUMNS, "living_or_deceased", ["living", "deceased", "historical", "unknown"])
    add_dropdown(new_disc, DISCOVERY_COLUMNS, "claimed_written_content_exists", ["TRUE", "FALSE", "unknown"])

    new_queue = new_wb.create_sheet("Queue")
    new_queue.append(queue_headers)
    style_header(new_queue, queue_headers)
    for row in queue_rows:
        new_queue.append([row.get(c) for c in queue_headers])
    size_columns(
        new_queue, queue_headers,
        wide={"name", "url", "final_url", "notes", "flag_reason", "promoted_from_discovery"},
        narrow={"source_db_id", "submitted_by", "result_document_id", "content_sha256",
                "worker_id", "lease_expires_at", "run_after", "created_at", "updated_at"},
    )
    add_dropdown(new_queue, queue_headers, "stage", ["ready_to_queue", "already_queued", "done"])
    add_dropdown(new_queue, queue_headers, "source_format", ["web_page", "pdf"])
    add_dropdown(new_queue, queue_headers, "source_scope", ["single", "collection"])
    add_dropdown(new_queue, queue_headers, "attribution_mode", ["declared", "per_item"])
    add_dropdown(new_queue, queue_headers, "on_unknown_author", ["flag", "skip"])
    add_dropdown(new_queue, queue_headers, "retain_original_text", ["TRUE", "FALSE"])
    add_dropdown(new_queue, queue_headers, "cleared_to_run", ["TRUE", "FALSE"])

    new_wb.save(str(SHEET_PATH))

    print()
    print("=" * 90)
    print("RECONCILIATION")
    print("=" * 90)
    print(f"Unique new candidates loaded: {unique_count}")
    print(f"Duplicate rows collapsed within this batch: {duplicates_collapsed}")
    print(f"Flagged as already in corpus: {len(matches)}")
    print(f"Discovery tab total after load: {len(all_disc_rows)} (6 pre-existing + {unique_count} new)")
    print(f"Queue tab rows (unchanged): {len(queue_rows)}")
    print(f"Wrote: {SHEET_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
