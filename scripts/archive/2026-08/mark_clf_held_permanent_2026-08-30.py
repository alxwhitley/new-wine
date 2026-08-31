"""Mark the 15 held CLF recordings as permanently ruled (Alex, 2026-08-30).

Sets status='held_permanent' on the CLF Church rows at guess='sermon',
ingest='FALSE', status='triaged', and records the reason in a new `notes`
column. Dry-run by default; --apply required to write.

Why status, not just a note: youtube_ingest.py's gate is
`ingest == "TRUE" AND status == "triaged"`, and the file already uses terminal
statuses as the durable protection ("done_prior rows additionally have
ingest=FALSE -- double-excluded"). A terminal status double-excludes these the
same way and also removes them from any future triaged-based sweep.

The `notes` column is column 8. youtube_triage/youtube_ingest map columns
positionally over the first 7 (COLUMNS/COL), so an 8th column is ignored by
both and cannot shift any existing read.
"""
import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

QUEUE = Path("/Users/alexwhitley/rhemata/sources/youtube/ingest_queue.xlsx")
SHEET = "CLF Church"
NEW_STATUS = "held_permanent"
NOTE = ("held permanently 2026-08-30 (Alex) — whole-service recording carrying "
        "named-congregant pastoral material; ruled on content shape, not runtime. "
        "See docs/audits/2026-08/clf_held_recordings_review_2026-08-30.md")

# The 7 columns both scripts map positionally. Guarded, not assumed.
EXPECTED = ["url", "video_title", "channel_name", "guess", "ingest", "status",
            "resolved_source"]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="write (default: dry run)")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(QUEUE)
    ws = wb[SHEET]

    header = [str(ws.cell(row=1, column=i + 1).value or "") for i in range(7)]
    if header != EXPECTED:
        sys.exit(f"ABORT: unexpected header layout {header!r}; refusing to write.")

    notes_col = 8
    existing_h8 = str(ws.cell(row=1, column=notes_col).value or "").strip()
    if existing_h8 and existing_h8 != "notes":
        sys.exit(f"ABORT: column 8 already holds {existing_h8!r}; refusing to overwrite.")

    targets = []
    for r in range(2, ws.max_row + 1):
        guess = str(ws.cell(row=r, column=4).value or "").strip()
        ingest = str(ws.cell(row=r, column=5).value or "").strip()
        status = str(ws.cell(row=r, column=6).value or "").strip()
        if guess == "sermon" and ingest == "FALSE" and status == "triaged":
            targets.append(r)

    print(f"{len(targets)} row(s) match guess=sermon, ingest=FALSE, status=triaged")
    for r in targets:
        title = str(ws.cell(row=r, column=2).value or "")[:58]
        print(f"  row {r:>3}  triaged -> {NEW_STATUS}   {title}")

    if len(targets) != 15:
        sys.exit(f"ABORT: expected exactly 15 rows, found {len(targets)}.")

    if not args.apply:
        print("\nDRY RUN — nothing written. Re-run with --apply.")
        return

    backup = QUEUE.with_suffix(
        f".pre_hold_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    shutil.copy2(QUEUE, backup)
    print(f"\nbackup: {backup.name}")

    ws.cell(row=1, column=notes_col, value="notes")
    for r in targets:
        ws.cell(row=r, column=6, value=NEW_STATUS)
        ws.cell(row=r, column=notes_col, value=NOTE)
    wb.save(QUEUE)
    print(f"wrote {len(targets)} row(s) + notes header")


if __name__ == "__main__":
    main()
