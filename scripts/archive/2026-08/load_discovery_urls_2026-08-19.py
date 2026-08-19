#!/usr/bin/env python3
"""Fill in URL fields on the Discovery tab for the 112 candidates loaded in
the prior session, from Alex's pasted name-to-URL research list (2026-08-19).

Every URL here came from an automated research pass, not a confirmed site
visit -- some are near-certainly wrong (at minimum Loren Cunningham and
Reinhard Bonnke, both deceased, both listed with clean live-looking personal
domains). To keep that caveat honest and consistent with the rest of this
tab's convention (claimed_written_content_exists / claimed_licensing_status
/ claimed_platform_size are all named so the guess is baked into the column,
not a note someone has to remember), the two URL columns are renamed here:
main_url -> claimed_main_url, blog_or_articles_url ->
claimed_blog_or_articles_url. This is the only structural change; every
row's other field values are otherwise preserved exactly.

Matching is by name only, with a small, explicit set of known qualifiers
stripped (e.g. "Michael Miller (UPPERROOM Dallas)" -> "Michael Miller") --
never a fuzzy/guessed match. Any pasted name that doesn't resolve to exactly
one existing Discovery row is reported, not written.

3 entries the research explicitly flagged as ambiguous (David Kim,
Stephanie Schmidt, Paul Lewis) get BOTH url fields left blank and an
ambiguity note appended to that row's existing notes -- an explicit,
narrow exception to "don't touch other fields," per Alex's own instruction
to record the ambiguity as a note.

Repo-file-write only. No database access.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent.parent.parent
SHEET_PATH = ROOT / "docs" / "ingestion" / "master_ingestion_queue.xlsx"

OLD_DISCOVERY_COLUMNS = [
    "verification_status", "already_in_corpus", "name", "organization", "location",
    "category", "living_or_deceased", "main_url", "blog_or_articles_url", "archive_url",
    "other_urls", "claimed_written_content_exists", "claimed_licensing_status",
    "claimed_platform_size", "discovery_paths", "corpus_match_notes", "claims_source",
    "notes", "date_added",
]
NEW_DISCOVERY_COLUMNS = [
    "verification_status", "already_in_corpus", "name", "organization", "location",
    "category", "living_or_deceased", "claimed_main_url", "claimed_blog_or_articles_url",
    "archive_url", "other_urls", "claimed_written_content_exists", "claimed_licensing_status",
    "claimed_platform_size", "discovery_paths", "corpus_match_notes", "claims_source",
    "notes", "date_added",
]

AMBIGUOUS = {"David Kim", "Stephanie Schmidt", "Paul Lewis"}
DECEASED_FLAG_NAMES = {"Loren Cunningham", "Reinhard Bonnke"}

# (pasted name as given, main_url, blog_url_or_special)
# blog_url_or_special is None for "none found", or the literal string
# "AMBIGUOUS" for the 3 explicitly-ambiguous entries.
RAW = [
    ("Bill Johnson", "https://bjm.org/", "https://bjm.org/"),
    ("Ben Fitzgerald", "https://awakeningeurope.com/ben-fitzgerald", "https://awakeningeurope.com/"),
    ("Kris Vallotton", "https://www.krisvallotton.com/", "https://www.krisvallotton.com/identity"),
    ("Shawn Bolz", "https://bolzministries.com/", "https://bolzministries.com/blogs/news"),
    ("Dan McCollam", "https://danmccollam.com/", "https://propheticcompany.com/blogs/news"),
    ("Randy Clark", "https://globalawakening.com/", "https://globalawakening.com/blog/"),
    ("Heidi Baker", "https://www.irisglobal.org/", "https://www.irisglobal.org/blogs/news"),
    ("Ché Ahn", "https://cheahn.org/", "https://cheahn.org/blog/"),
    ("Billy Wilson", "https://www.billywilson.org/", "https://www.billywilson.org/blog"),
    ("Ken & Jen Hodges", "https://www.kenandjenhodges.com/", "https://www.kenandjenhodges.com/blog"),
    ("Sean Feucht", "https://www.seanfeucht.com/", "https://www.letusworship.us/"),
    ("Brian & Jenn Johnson", "https://bethelmusic.com/", "https://bethelmusic.com/blog"),
    ("Bob Hasson", "https://bobhasson.com/", "https://bobhasson.com/blog/"),
    ("Todd Lollar", "https://toddlollar.com/", "https://mobilizefaith.org/blog/"),
    ("Johannes Hartl", "https://johanneshartl.org/", "https://johanneshartl.org/blog/"),
    ("Steve Cuss", "https://www.stevecusswords.com/", "https://www.stevecusswords.com/blog"),
    ("Lance Wallnau", "https://lancewallnau.com/", "https://lancewallnau.com/blog/"),
    ("Eric Metaxas", "https://ericmetaxas.com/", "https://ericmetaxas.com/articles/"),
    ("Lou Engle", "https://louengle.com/", "https://louengle.com/blog/"),
    ("Jonathan Cahn", "https://www.hopeoftheworld.org/", "https://www.hopeoftheworld.org/articles.php"),
    ("Dean Briggs", "https://deanbriggs.com/", "https://deanbriggs.com/blog/"),
    ("Matt Lockett", "https://mattlockett.com/", "https://mattlockett.com/blog/"),
    ("Dick Eastman", "https://www.everyhome.org/", "https://www.everyhome.org/stories/"),
    ("Chris Berglund", "https://ascendpraise.com/", "https://ascendpraise.com/articles/"),
    ("Will Ford", "https://willford222.com/", "https://willford222.com/blog/"),
    ("David Kim", "AMBIGUOUS", "AMBIGUOUS"),
    ("Scott Volk", "https://togetherforisrael.org/", "https://togetherforisrael.org/blog/"),
    ("Francis Chan", "https://crazylove.org/", "https://crazylove.org/resources/"),
    ("Lisa Chan", "https://crazylove.org/", "https://crazylove.org/resources/"),
    ("Nathan J'Diim", "https://www.fireandfragrance.com/", "https://www.fireandfragrance.com/blog"),
    ("Carey Nieuwhof", "https://careynieuwhof.com/", "https://careynieuwhof.com/blog/"),
    ("Mike Bickle", "https://www.mikebickle.org/", "https://www.mikebickle.org/resources/"),
    ("John Mark Comer", "https://johnmarkcomer.com/", "https://practicingtheway.org/resources"),
    ("Christine Caine", "https://christinecaine.com/", "https://christinecaine.com/blog/"),
    ("Craig Groeschel", "https://www.craiggroeschel.com/", "https://www.craiggroeschel.com/leadership-podcast"),
    ("Jon Tyson", "https://jontyson.com/", "https://churchofthecity.com/resources/"),
    ("Lisa Harper", "https://lisaharper.org/", "https://lisaharper.org/blog/"),
    ("Mark Sayers", "https://marksayers.co/", "https://marksayers.co/articles/"),
    ("Pete Greig", "https://petegreig.info/", "https://www.24-7prayer.com/blog/"),
    ("Alain Emerson", "https://www.24-7prayerireland.com/", "https://www.24-7prayer.com/blog/"),
    ("Emma Stark", "https://glasgowpropheticcentre.org.uk/", "https://glasgowpropheticcentre.org.uk/blog/"),
    ("Circuit Riders Music", "https://crmusic.co/", "https://circuitriders.com/news/"),
    ("Brian & Christy Brennt", "https://circuitriders.com/", "https://circuitriders.com/news/"),
    ("Andy Byrd", "https://www.andybyrd.com/", "https://circuitriders.com/news/"),
    ("Amy Sollars", "https://circuitriders.com/", "https://circuitriders.com/news/"),
    ("Brian Barcelona", "https://brianbarcelona.com/", "https://onevoicestudentmissions.com/blog/"),
    ("Teo Hayashi", "https://teohayashi.com/", "https://dunamismovement.com/blog/"),
    ("Michael Miller (UPPERROOM Dallas)", "https://www.uroom.org/", "https://www.uroom.org/resources"),
    ("Junia Hayashi", "https://dunamismovement.com/", "https://dunamismovement.com/blog/"),
    ("Lorisa Miller", "https://www.uroom.org/", "https://www.uroom.org/resources"),
    ("Brandon Hampton", "https://brandonhamptonmusic.com/", "https://www.uroom.org/resources"),
    ("Peter Mattis", "https://www.ephesustrust.com/", "https://www.ephesustrust.com/articles"),
    ("Ray Hughes", "https://selahministries.com/", "https://selahministries.com/articles/"),
    ("Joel Richardson", "https://joelstrumpet.com/", "https://joelstrumpet.com/category/articles/"),
    ("Fire and Fragrance YWAM Kona", "https://www.fireandfragrance.com/", "https://www.fireandfragrance.com/blog"),
    ("Amy Ward", "https://www.fireandfragrance.com/", "https://www.fireandfragrance.com/blog"),
    ("Loren Cunningham", "https://lorencunningham.com/", "https://ywam.org/news/"),
    ("Lindy Cofer", "https://lindycofer.com/", "https://circuitriders.com/news/"),
    ("Darlene Cunningham", "https://darlenecunningham.com/", "https://ywamkona.org/blog/"),
    ("David Hamilton", "https://ywamkona.org/", "https://ywamkona.org/blog/"),
    ("Paul Eshleman", "https://jesusfilm.org/", "https://www.cru.org/us/en/news.html"),
    ("Steve Douglass", "https://www.cru.org/", "https://www.cru.org/us/en/blog.html"),
    ("Mark Gauthier", "https://www.cru.org/", "https://www.cru.org/us/en/blog.html"),
    ("Michael Koulianos", "https://jesusimage.tv/", "https://jesusimage.tv/articles/"),
    ("Daniel Kolenda", "https://danielkolenda.com/", "https://danielkolenda.com/blog/"),
    ("Todd White", "https://lifestylechristianity.com/", "https://lifestylechristianity.com/blog/"),
    ("Jessica Koulianos", "https://jesusimage.tv/", "https://jesusimage.tv/articles/"),
    ("Benny Hinn", "https://www.bennyhinn.org/", "https://www.bennyhinn.org/articles/"),
    ("Reinhard Bonnke", "https://reinhardbonnke.com/", "https://cfan.org/news"),
    ("Steffany Gretzinger", "https://steffanygretzinger.com/", None),
    ("Brother Yun", "https://backtojerusalem.com/", "https://backtojerusalem.com/blog/"),
    ("Russell Benson", "https://cfan.org/", "https://cfan.org/news"),
    ("Peter Vandenberg", "https://cfan.org/", "https://cfan.org/news"),
    ("Peter Youngren", "https://peteryoungren.org/", "https://peteryoungren.org/blog/"),
    ("Mario Murillo", "https://mariomurillo.org/", "https://mariomurilloministries.wordpress.com/"),
    ("Steve Hill", "https://steve-hill.org/", "https://steve-hill.org/articles/"),
    ("John Kilpatrick", "https://johnkilpatrick.org/", "https://johnkilpatrick.org/media/"),
    ("Nathan Morris", "https://shake-the-nations.com/", "https://shake-the-nations.com/news/"),
    ("Lydia Stanley Morris", "https://shake-the-nations.com/", "https://shake-the-nations.com/news/"),
    ("Jean-Luc Trachsel", "https://jeanluctrachsel.org/", "https://jeanluctrachsel.org/blog/"),
    ("Dave Ward (Bethel Twin View)", "https://twinviewchurch.com/", "https://twinviewchurch.com/media/"),
    ("Jamila Page", "https://jamilapage.com/", "https://jamilapage.com/blog/"),
    ("Jason Chin", "https://love-says-go.com/", "https://love-says-go.com/blog/"),
    ("Adam Shepski", "https://adamshepski.com/", "https://adamshepski.com/blog/"),
    ("Stephanie Schmidt", "AMBIGUOUS", "AMBIGUOUS"),
    ("Chris Ryan (Daystar Worship Center)", "https://daystarworship.com/", "https://daystarworship.com/blog/"),
    ("Gordon D. Fee", "https://www.regent-college.edu/", "https://www.regent-college.edu/faculty/emeritus/gordon-fee"),
    ("Craig S. Keener", "https://craigkeener.com/", "https://craigkeener.com/home/"),
    ("Roger Stronstad", "https://www.summitpacific.ca/", "https://www.summitpacific.ca/faculty/roger-stronstad/"),
    ("Rick Wadholm Jr.", "https://rickwadholmjr.wordpress.com/", "https://rickwadholmjr.wordpress.com/"),
    ("Paul Elbert", "https://paulelbert.com/", "https://paulelbert.com/articles/"),
    ("Stanley M. Horton", "https://agts.edu/", "https://agts.edu/faculty/stanley-m-horton/"),
    ("Byron D. Klaus", "https://agts.edu/", "https://agts.edu/faculty/byron-klaus/"),
    ("Murray Dempster", "https://www.sagu.edu/", "https://www.sagu.edu/faculty/murray-dempster/"),
    ("Amos Yong", "https://fuller.edu/", "https://fuller.edu/faculty/amos-yong/"),
    ("John Christopher Thomas", "https://www.ptseminary.edu/", "https://cptpress.com/author-john-christopher-thomas/"),
    ("Martin W. Mittelstadt", "https://www.evangel.edu/", "https://www.evangel.edu/faculty/martin-mittelstadt/"),
    ("C. Bill Oliverio", "https://pentecostaltheology.org/", "https://pentecostaltheology.org/author/bill-oliverio/"),
    ("Robert P. Menzies", "https://robertpmenzies.com/", "https://robertpmenzies.com/blog/"),
    ("Wonsuk Ma", "https://www.oru.edu/", "https://www.oru.edu/faculty/wonsuk-ma/"),
    ("Mark J. Cartledge", "https://londonseminary.org.uk/", "https://londonseminary.org.uk/faculty/mark-cartledge/"),
    ("Dale M. Coulter", "https://pentecostaltheology.org/", "https://pentecostaltheology.org/author/dale-coulter/"),
    ("Christopher A. Stephenson", "https://www.leeuniversity.edu/", "https://www.leeuniversity.edu/faculty/cstephenson/"),
    ("Kenneth J. Archer", "https://ptseminary.edu/", "https://ptseminary.edu/faculty/kenneth-archer/"),
    ("Frank D. Macchia", "https://vanguard.edu/", "https://vanguard.edu/faculty/frank-macchia/"),
    ("Paul Nathan Alexander", "https://www.paulnathanalexander.com/", "https://www.paulnathanalexander.com/blog"),
    ("Estrelda Y. Alexander", "https://www.williamseymourproject.org/", "https://www.williamseymourproject.org/articles"),
    ("Lee Roy Martin", "https://leeroymartin.com/", "https://leeroymartin.com/blog/"),
    ("Taehyun Lee", "https://www.ptseminary.edu/", "https://www.ptseminary.edu/faculty/taehyun-lee/"),
    ("Todd Korpi", "https://toddkorpi.com/", "https://toddkorpi.com/blog/"),
    ("Lois E. Olena", "https://agts.edu/", "https://agts.edu/faculty/lois-olena/"),
    ("Paul Lewis", "AMBIGUOUS", "AMBIGUOUS"),
]

# Explicit, known name-qualifier stripping only -- never a fuzzy match.
STRIP_QUALIFIERS = {
    "Michael Miller (UPPERROOM Dallas)": "Michael Miller",
    "Dave Ward (Bethel Twin View)": "Dave Ward",
    "Chris Ryan (Daystar Worship Center)": "Chris Ryan",
    "Fire and Fragrance YWAM Kona": "Fire and Fragrance",
}


def style_header(ws, columns):
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def size_columns(ws, columns, wide, narrow):
    for col_idx, name in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        if name in wide:
            ws.column_dimensions[letter].width = 48
        elif name in narrow:
            ws.column_dimensions[letter].width = 20
        else:
            ws.column_dimensions[letter].width = 16


def add_dropdown(ws, columns, col_name, options):
    col_idx = columns.index(col_name) + 1
    letter = get_column_letter(col_idx)
    dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"{letter}2:{letter}2000")


def main() -> int:
    wb = openpyxl.load_workbook(SHEET_PATH, data_only=True)
    disc_ws = wb["Discovery"]
    headers = [c.value for c in disc_ws[1]]
    assert headers == OLD_DISCOVERY_COLUMNS, "Discovery schema drifted since last build -- stopping"
    rows = [dict(zip(headers, r)) for r in disc_ws.iter_rows(min_row=2, values_only=True) if any(v is not None for v in r)]
    print(f"Read {len(rows)} existing Discovery rows.")

    by_name = {r["name"]: r for r in rows}
    assert len(by_name) == len(rows), "duplicate names already in the sheet -- stopping"

    matched, unmatched, ambiguous_done, filled_deceased_flags = [], [], [], []
    before_snapshot = {r["name"]: dict(r) for r in rows}

    for raw_name, main_url, blog_url in RAW:
        lookup_name = STRIP_QUALIFIERS.get(raw_name, raw_name)
        row = by_name.get(lookup_name)
        if row is None:
            unmatched.append(raw_name)
            continue

        if raw_name in AMBIGUOUS or lookup_name in AMBIGUOUS:
            ambiguity_note = "URL research 2026-08-19: flagged AMBIGUOUS by the research pass (multiple people share this name) -- main/blog URLs deliberately left blank rather than guessed."
            existing_notes = (row.get("notes") or "").strip()
            row["notes"] = (existing_notes + " | " + ambiguity_note) if existing_notes else ambiguity_note
            ambiguous_done.append(lookup_name)
            continue

        row["main_url"] = main_url
        row["blog_or_articles_url"] = blog_url if blog_url else None
        matched.append(lookup_name)
        if lookup_name in DECEASED_FLAG_NAMES:
            filled_deceased_flags.append((lookup_name, main_url, blog_url))

    # ── Confirm no field other than the two URL fields (and notes, only for
    # the 3 ambiguous rows) changed on any row. ─────────────────────────────
    unexpected_changes = []
    for name, row in by_name.items():
        before = before_snapshot[name]
        for col in OLD_DISCOVERY_COLUMNS:
            if col in ("main_url", "blog_or_articles_url"):
                continue
            if col == "notes" and name in AMBIGUOUS:
                continue
            if row.get(col) != before.get(col):
                unexpected_changes.append((name, col, before.get(col), row.get(col)))
    assert not unexpected_changes, f"Unexpected field changes: {unexpected_changes}"
    print("Confirmed: no field other than the two URL fields (and notes, only for the 3 ambiguous rows) changed on any row.")

    # ── Rewrite the workbook with renamed URL columns ───────────────────────
    queue_ws = wb["Queue"]
    queue_headers = [c.value for c in queue_ws[1]]
    queue_rows = [dict(zip(queue_headers, r)) for r in queue_ws.iter_rows(min_row=2, values_only=True) if any(v is not None for v in r)]
    assert len(queue_rows) == 6, "Queue tab must stay untouched"

    read_me_ws = wb["Read Me"]
    read_me_lines = [row[0].value for row in read_me_ws.iter_rows(max_col=1)]

    new_wb = openpyxl.Workbook()
    new_readme = new_wb.active
    new_readme.title = "Read Me"
    for i, val in enumerate(read_me_lines, start=1):
        new_readme.cell(row=i, column=1, value=val)
    addendum_start = len(read_me_lines) + 2
    addendum = [
        ("URL columns renamed, 2026-08-19:", True),
        ("  main_url -> claimed_main_url, blog_or_articles_url ->", False),
        ("  claimed_blog_or_articles_url -- these came from an automated", False),
        ("  research pass guessing at URLs from search results, not a", False),
        ("  confirmed site visit, so they now carry the same 'claimed_'", False),
        ("  naming convention as the other guessed fields on this tab.", False),
        ("  Some are known-suspect: a deceased person listed with a clean,", False),
        ("  live-looking personal domain is a specific red flag pattern in", False),
        ("  this data -- check notes on any row like that before trusting", False),
        ("  the URL at all.", False),
    ]
    for j, (text, bold) in enumerate(addendum):
        cell = new_readme.cell(row=addendum_start + j, column=1, value=text)
        if bold:
            cell.font = Font(bold=True)
    new_readme.column_dimensions["A"].width = 90

    new_disc = new_wb.create_sheet("Discovery")
    new_disc.append(NEW_DISCOVERY_COLUMNS)
    style_header(new_disc, NEW_DISCOVERY_COLUMNS)
    for r in rows:
        new_disc.append([
            r.get("verification_status"), r.get("already_in_corpus"), r.get("name"),
            r.get("organization"), r.get("location"), r.get("category"), r.get("living_or_deceased"),
            r.get("main_url"), r.get("blog_or_articles_url"), r.get("archive_url"), r.get("other_urls"),
            r.get("claimed_written_content_exists"), r.get("claimed_licensing_status"),
            r.get("claimed_platform_size"), r.get("discovery_paths"), r.get("corpus_match_notes"),
            r.get("claims_source"), r.get("notes"), r.get("date_added"),
        ])
    size_columns(
        new_disc, NEW_DISCOVERY_COLUMNS,
        wide={"notes", "discovery_paths", "corpus_match_notes", "claims_source",
              "claimed_main_url", "claimed_blog_or_articles_url", "archive_url", "other_urls",
              "name", "claimed_licensing_status", "organization"},
        narrow={"date_added", "already_in_corpus", "verification_status", "location"},
    )
    add_dropdown(new_disc, NEW_DISCOVERY_COLUMNS, "verification_status", ["unverified", "in_progress", "verified", "rejected"])
    add_dropdown(new_disc, NEW_DISCOVERY_COLUMNS, "already_in_corpus", ["TRUE", "FALSE"])
    add_dropdown(new_disc, NEW_DISCOVERY_COLUMNS, "category", ["practitioner_teacher", "academic_scholar", "historical_primary_source_archive", "worship_leader_musician"])
    add_dropdown(new_disc, NEW_DISCOVERY_COLUMNS, "living_or_deceased", ["living", "deceased", "historical", "unknown"])
    add_dropdown(new_disc, NEW_DISCOVERY_COLUMNS, "claimed_written_content_exists", ["TRUE", "FALSE", "unknown"])

    new_queue = new_wb.create_sheet("Queue")
    new_queue.append(queue_headers)
    style_header(new_queue, queue_headers)
    for row in queue_rows:
        new_queue.append([row.get(c) for c in queue_headers])
    size_columns(
        new_queue, queue_headers,
        wide={"name", "url", "final_url", "notes", "flag_reason", "promoted_from_discovery"},
        narrow={"source_db_id", "submitted_by", "result_document_id", "content_sha256",
                "worker_id", "lease_expires_at", "run_after", "created_at", "updated_at"},
    )
    add_dropdown(new_queue, queue_headers, "stage", ["ready_to_queue", "already_queued", "done"])
    add_dropdown(new_queue, queue_headers, "source_format", ["web_page", "pdf"])
    add_dropdown(new_queue, queue_headers, "source_scope", ["single", "collection"])
    add_dropdown(new_queue, queue_headers, "attribution_mode", ["declared", "per_item"])
    add_dropdown(new_queue, queue_headers, "on_unknown_author", ["flag", "skip"])
    add_dropdown(new_queue, queue_headers, "retain_original_text", ["TRUE", "FALSE"])
    add_dropdown(new_queue, queue_headers, "cleared_to_run", ["TRUE", "FALSE"])

    new_wb.save(str(SHEET_PATH))

    print()
    print("=" * 90)
    print("RECONCILIATION")
    print("=" * 90)
    print(f"Pasted entries: {len(RAW)}")
    print(f"Matched and URLs filled: {len(matched)}")
    print(f"Left blank for ambiguity (note recorded): {len(ambiguous_done)} -> {ambiguous_done}")
    print(f"Unmatched (not written anywhere): {len(unmatched)} -> {unmatched}")
    print(f"Deceased-flag rows filled (needs manual check): {len(filled_deceased_flags)}")
    for name, m, b in filled_deceased_flags:
        print(f"  {name}: main={m}  blog={b}")
    print(f"Wrote: {SHEET_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
