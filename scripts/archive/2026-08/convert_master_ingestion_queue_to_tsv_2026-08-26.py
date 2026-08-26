#!/usr/bin/env python3
"""One-time conversion: docs/ingestion/master_ingestion_queue.xlsx (4 tabs)
-> four plain-text TSV files (see ingestion_sheet_io.py). Run once, then
archived -- never meant to be re-run against a live xlsx that no longer
exists after the cutover commit.

Also adds the 7 new Discovery columns this session's task required
(clearance checklist x4 bools + 1 date, clearance-cost lane, blog index
url), appended after the existing 25 columns, all blank/FALSE on every
existing row -- nothing is retroactively marked as checked or classified.
The Discovery "declined state" requirement (task step 2) is NOT a new
column: verification_status='rejected' + reviewed_at + review_notes
already does exactly this (6 existing rows already use it) -- see the
session report for why no new column was added for that step.

Run: python3.12 scripts/archive/2026-08/convert_master_ingestion_queue_to_tsv_2026-08-26.py
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

import ingestion_sheet_io as io

XLSX_PATH = io.ROOT / "docs" / "ingestion" / "master_ingestion_queue.xlsx"

NEW_DISCOVERY_COLUMNS = [
    "site_visited_by_human",
    "author_identity_confirmed",
    "licensing_posture_confirmed",
    "content_type_confirmed",
    "clearance_checked_at",
    "clearance_cost_lane",
    "blog_index_url",
]
NEW_DISCOVERY_BOOL_DEFAULTS = {
    "site_visited_by_human": False,
    "author_identity_confirmed": False,
    "licensing_posture_confirmed": False,
    "content_type_confirmed": False,
}


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    for tab_name, out_path in io.TAB_FILES.items():
        if tab_name not in wb.sheetnames:
            raise SystemExit(f"Expected tab {tab_name!r} not found in {XLSX_PATH}")
        ws = wb[tab_name]

        if tab_name == "Read Me":
            # Row 1 is the title line of the actual content, not a column
            # name -- treating it as a header (as every other tab does)
            # would silently swallow it as metadata. Use a neutral header
            # and keep every row 1..max_row, blanks included: blank lines
            # here are meaningful paragraph separators, not sheet noise.
            headers = ["text"]
            rows = [{"text": row[0]} for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)]
            io.write_tab(out_path, headers, rows)
            print(f"Wrote {out_path.relative_to(io.ROOT)}: {len(headers)} columns, {len(rows)} rows (blank separators preserved)")
            continue

        headers = [str(c.value) for c in ws[1]]
        rows = []
        for raw in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in raw):
                continue
            rows.append(dict(zip(headers, raw)))

        if tab_name == "Discovery":
            headers = headers + NEW_DISCOVERY_COLUMNS
            for row in rows:
                for col in NEW_DISCOVERY_COLUMNS:
                    row[col] = NEW_DISCOVERY_BOOL_DEFAULTS.get(col, None)

        io.write_tab(out_path, headers, rows)
        print(f"Wrote {out_path.relative_to(io.ROOT)}: {len(headers)} columns, {len(rows)} rows")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
